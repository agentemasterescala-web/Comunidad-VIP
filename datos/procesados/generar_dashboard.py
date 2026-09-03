#!/usr/bin/env python3
"""
Dashboard del Panel Comunidad VIP — Iván Caicedo.
Layout tipo panel ejecutivo con tabs, charts y métricas en vivo.
"""
import os, json, html, re, hashlib
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)
ORIG = os.path.join(PARENT, "originales")
RAW = os.path.join(HERE, "ghl_contacts_raw.json")
MAESTRO = os.path.join(HERE, "maestro_emails.xlsx")
CLAS = os.path.join(HERE, "clasificacion_usuarios.xlsx")
OUT = os.path.join(HERE, "dashboard.html")

# IMPORTANTE: lista (no set) para garantizar orden de iteración estable entre
# ejecuciones. Un set daría orden distinto en cada corrida por hash randomization,
# lo que provocaba que `paises` cambiara de orden en cada regeneración del
# dashboard.html y rompía el change detection del publish-light.
TIENDA_IDS = [
    "CQk3UpeEwUnbegiqR2Q3","P3jZOcralEFKIg4XpYho","p7TjCy0lVm6fP9xEbS3l",
    "CZSUn21ycO4tr4LkNrbj","Mir5XAqxPoCrfT3fkgRF","riVtpCpiQPJvASPdEkKd",
    "ThOpku1erpbHGCCBei6Z","83RmxTxcA8gkkWUsADls","2bl6kY6oQEIbPRRKpmaq",
    "75tlJ2SQeSdymOTxoScY"
]
TIENDA_PAIS_IDS = {
    "CQk3UpeEwUnbegiqR2Q3":"0HWT1wbaaadgxxBPODUH","P3jZOcralEFKIg4XpYho":"yJyX6eZUzkgnpBmAslde",
    "p7TjCy0lVm6fP9xEbS3l":"IsI0hZEBHczVZ0itmPmV","CZSUn21ycO4tr4LkNrbj":"CqZvpz0gtfu4bvCZwNqA",
    "Mir5XAqxPoCrfT3fkgRF":"yWAvExtJrOJXTdjDnQuj","riVtpCpiQPJvASPdEkKd":"u7iiKtYTqJSKiFpioMdv",
    "ThOpku1erpbHGCCBei6Z":"28jQePKJQGIZ198U0R6Z","83RmxTxcA8gkkWUsADls":"gCaIQVieS9PqEU5AI8Uh",
    "2bl6kY6oQEIbPRRKpmaq":"pEkrMm5ahV6PPow8PYQW","75tlJ2SQeSdymOTxoScY":"cnShQcBSUMbU1WAFVqQx",
}
PROG_ID = "TwrGqT8nj3jJmVMUlVFq"
TAG_MASTER = "escala"
TAG_INICIACION = "iniciacion"

TIER_ORDER = ["Diamante","Platino","Oro","Plata","Bronce","Sin clasificar"]

ES_MONTHS_SHORT_FULL = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
def mes_label(yyyy_mm):
    y, m = yyyy_mm.split("-")
    return f"{ES_MONTHS_SHORT_FULL[int(m)-1]} {y}"


# ============================================================
# Normalización de nombres pegados (típico en data Dropi)
# Ej: "nicolasmanrique" -> "Nicolas Manrique"
#     "juanavalentinalopezsanchez" -> "Juana Valentina Lopezsanchez"
# ============================================================
import unicodedata
NOMBRES_HISPANOS = {
    # Masculinos
    "abraham","abrahan","adalberto","adan","adolfo","adrian","agustin","alejandro","alberto","albino",
    "alex","alexander","alexis","alfonso","alfredo","alirio","alvaro","amilcar","anderson","andres",
    "andy","angel","anibal","antonio","arcadio","arcadio","armando","arnaldo","arnulfo","arthur","arturo",
    "augusto","aurelio","axel","baltazar","benjamin","bernardo","bismark","boris","brandon","breiner",
    "brian","bruno","byron","camilo","carlos","cesar","christian","christhian","cipriano","cristian",
    "cristhian","cristobal","cruz","daniel","danny","dario","david","deivis","deivys","deyvis","deyson",
    "diego","diomedes","domingo","duvan","eder","edgar","edinson","eduardo","edwin","efrain","efren",
    "einer","elias","eliecer","elkin","emanuel","emiliano","emilio","enrique","eric","ernesto","esteban",
    "estiven","ever","ezequiel","fabian","fabio","federico","felipe","felix","fermin","fernando",
    "francisco","franklin","franco","fredy","freddy","gabriel","geison","geovanny","geovany","gerardo",
    "german","gerson","gilberto","giovanni","giovani","gonzalo","gregorio","guillermo","gustavo",
    "hamilton","harold","hector","henry","heriberto","hernan","hernando","hilario","homero","horacio",
    "hugo","humberto","ibarra","ignacio","ildefonso","isaac","isaias","ismael","ivan","jacinto","jacobo",
    "jaime","jair","jairo","james","javier","jean","jefferson","jeremy","jeronimo","jesus","jeyder",
    "jeyfer","jhojan","jhon","jhonatan","jhonny","joaquin","joel","johan","johnatan","johnny","jonathan",
    "jordan","jorge","jose","joshua","juan","julian","julio","kenny","kevin","leandro","leider","leiver",
    "leonardo","leonel","leyder","lisandro","lorenzo","luciano","lucas","luis","manuel","marcelo","marco",
    "marcos","mariano","mario","martin","marvin","matheus","mateo","mateus","matias","mauricio","maximo",
    "michael","miguel","milton","misael","moises","nelson","nestor","nicanor","nicolas","noel","octavio",
    "olmer","omar","orlando","oscar","osvaldo","pablo","patricio","paul","pedro","pio","rafael","ramiro",
    "ramon","raul","rene","reinaldo","ricardo","richard","roberto","robinson","rodolfo","rodrigo",
    "rogelio","rolando","ronald","ronaldo","ruben","salomon","samuel","santiago","saul","sebastian",
    "sergio","silvio","simon","steven","stiven","teodoro","tobias","tomas","ubaldo","valentin","vicente",
    "victor","vinicio","vladimir","walter","wilber","wilberto","wilfredo","wilfrido","william","willmar",
    "wilmer","wilmar","wilson","wladimir","yamil","yefferson","yefri","yeison","yender","yesid","yhojan",
    "yhonatan","yoel","yonatan","yonny","yordan",
    # Femeninos
    "adelina","adriana","alba","alejandra","alexandra","alicia","alondra","amalia","amelia","amparo",
    "ana","andrea","anet","angie","angela","angelica","anita","antonia","aracely","araceli","ariana",
    "ariadna","astrid","aura","aurora","azucena","barbara","beatriz","berenice","betty","bianca","blanca",
    "brenda","camila","carla","carmen","carolina","catalina","cecilia","celeste","celia","cesarina",
    "chela","cinthia","claudia","clemencia","constanza","consuelo","cristina","dana","daniela","dayana",
    "delia","diana","dolores","dora","dorothy","edith","elba","eliana","elena","elia","elisa","eliza",
    "elizabeth","elsa","elvira","emilia","emma","enith","erika","esmeralda","esperanza","estefania",
    "estela","ester","esther","estrella","eugenia","eulalia","eva","fabiana","fabiola","fanny","felipa",
    "fernanda","flor","flora","florencia","florinda","francisca","gabriela","gala","geraldine","gladis",
    "gladys","gloria","graciela","greicell","greisy","guadalupe","henny","hilda","ingrid","inocencia",
    "irene","iris","irma","isabel","isabela","ivanna","jackeline","jacqueline","janeth","jennifer",
    "jenny","jessenia","jessica","joana","joanna","johana","johanna","josefa","josefina","juana","judith",
    "julia","juliana","julieta","karen","karina","karla","katherine","katia","kelly","kim","kimberly",
    "laura","leidy","lesly","leticia","lidia","lilia","liliana","lilibeth","lina","linda","lisbeth",
    "lizbeth","lizeth","lorena","lourdes","lucia","lucrecia","ludy","luisa","luz","macarena","magaly",
    "magdalena","maite","manuela","marcela","margarita","maria","mariana","maricela","marilin","marilyn",
    "marina","marisa","marisol","marlene","marta","martha","mary","matilde","mayerlin","mayra","melissa",
    "mercedes","milagros","milena","mildred","mireya","monica","monserrat","nancy","natalia","nayeli",
    "nelida","nelly","nereida","nidia","nieves","noemi","nora","norma","ofelia","olga","olivia","olivia",
    "ondina","pamela","paola","patricia","paula","paulina","paz","perla","piedad","pilar","pricila",
    "priscila","raquel","ramona","rebeca","regina","rita","rocio","rosa","rosalia","rosalina","rosario",
    "rosaura","roxana","rubi","ruth","sandra","sara","sarah","sayda","selena","selenia","senith","shirley",
    "silvia","sofia","soledad","sonia","stefania","stefany","stephania","stephany","susana","tania",
    "tatiana","tere","teresa","trinidad","ursula","valentina","valeria","vanesa","vanessa","vera",
    "veronica","victoria","violeta","viviana","wendy","ximena","yadira","yajaira","yamila","yamile",
    "yamileth","yaneth","yarisbel","yenifer","yenny","yesenia","yesica","yolanda","yorleny","yuli",
    "yulieth","yuliana","yuri","zaida","zaira","zenaida","zoila","zoraida",
}

def _strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

# Precomputado: map de forma sin tildes → forma canónica
_NOMBRES_MAP = {_strip_accents(n): n for n in NOMBRES_HISPANOS}
_NOMBRES_SORTED = sorted(_NOMBRES_MAP.keys(), key=lambda x: -len(x))

# Valores basura que mejor mostrar vacíos
_NOMBRE_BASURA = {"#n/a","n/a","na","null","none","-","--","sin nombre","sin_nombre","x","xxx"}

def normalizar_nombre_dropi(s):
    """Si el nombre viene pegado (sin espacios), intenta separar nombre + apellido.
    Maneja tildes y filtra valores basura."""
    if not s:
        return ""
    s = str(s).strip()
    if not s or s.lower() in _NOMBRE_BASURA:
        return ""
    # Si ya tiene espacios → title-case por palabra (preserva tildes)
    if " " in s:
        return " ".join(w.capitalize() for w in s.split() if w)
    s_no_acc = _strip_accents(s.lower())
    s_low = s.lower()
    parts = []                # nombres detectados (con tildes)
    rest = s_low              # cola con tildes
    rest_no_acc = s_no_acc    # cola sin tildes para matching
    while rest_no_acc and len(parts) < 2:
        match_key = None
        for name_key in _NOMBRES_SORTED:
            if rest_no_acc.startswith(name_key):
                match_key = name_key
                break
        if match_key and len(rest_no_acc) - len(match_key) >= 3:
            # tomar la porción de la cola con tildes (misma longitud)
            parts.append(rest[:len(match_key)])
            rest = rest[len(match_key):]
            rest_no_acc = rest_no_acc[len(match_key):]
        else:
            break
    if not parts:
        return s.capitalize()
    out = [p.capitalize() for p in parts]
    if rest:
        out.append(rest.capitalize())
    return " ".join(out)


def load_maestro_window():
    wb = openpyxl.load_workbook(MAESTRO, read_only=True, data_only=True)
    ws = wb["MAESTRO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]; idx = {n:i for i,n in enumerate(hdr)}
    db = defaultdict(list)
    months_set = set()
    has_nombre = "nombre" in idx
    has_tel = "telefono" in idx
    for r in rows[1:]:
        em = r[idx["email"]]
        if not em: continue
        mes = r[idx["mes"]]; months_set.add(mes)
        db[em].append({
            "pais": r[idx["pais"]], "mes": mes,
            "pedidos": r[idx["pedidos"]] or 0,
            "entregados": r[idx["entregados"]] or 0,
            "devoluciones": r[idx["devoluciones"]] or 0,
            "nombre": (r[idx["nombre"]] if has_nombre else "") or "",
            "telefono": (r[idx["telefono"]] if has_tel else "") or "",
        })
    months = sorted(months_set)[-5:]
    return db, months


_PAIS_CANONICO = {
    "colombia":"Colombia","ecuador":"Ecuador","peru":"Perú","mexico":"México",
    "guatemala":"Guatemala","panama":"Panamá","chile":"Chile","argentina":"Argentina",
    "paraguay":"Paraguay","uruguay":"Uruguay","bolivia":"Bolivia","venezuela":"Venezuela",
    "costarica":"Costa Rica","honduras":"Honduras","elsalvador":"El Salvador",
    "nicaragua":"Nicaragua","republicadominicana":"República Dominicana",
    "brasil":"Brasil","brazil":"Brasil","españa":"España","espana":"España",
    "estadosunidos":"Estados Unidos","usa":"Estados Unidos",
}
def pais_canonico(p):
    """Colapsa variantes ('COLOMBIA'/'colombia'/'Colombia') a forma canónica.
    Fallback: title case sin acentos si no está en el mapa."""
    if not p: return ""
    import unicodedata as _ud
    k = "".join(c for c in _ud.normalize("NFD", str(p)) if not _ud.combining(c))
    k = k.lower().strip().replace(" ", "")
    if k in _PAIS_CANONICO: return _PAIS_CANONICO[k]
    return " ".join(w.capitalize() for w in str(p).strip().split())

def extract_tiendas(contact):
    """Extrae las tiendas de un contacto GHL. Dedupea por (email, país) — así un
    contacto puede tener el MISMO email registrado en dos slots con países
    distintos (ej. una cuenta Dropi que opera en Colombia y Guatemala) y ambas
    tiendas aparecen. Solo se filtra la duplicación exacta (mismo email + mismo país)."""
    cf = {f["id"]: f.get("value") for f in contact.get("customFields", [])}
    seen = {}   # key: (email, pais_normalizado) → tienda
    for fid in TIENDA_IDS:
        em = cf.get(fid)
        if em and isinstance(em, str) and "@" in em:
            em_low = em.strip().lower()
            pais = (cf.get(TIENDA_PAIS_IDS[fid]) or "").strip()
            key = (em_low, pais.lower())
            if key not in seen:
                seen[key] = {"email": em_low, "pais": pais}
    return list(seen.values())


def derivar_paises(contact, maestro):
    """Países de un contacto, derivados del Excel por país de Dropi.
    Combina: el campo 'país' de cada tienda del contacto + los países de las ventas
    Dropi asociadas (buscando por el email de cada tienda Y por el email principal).
    Devuelve lista única ordenada. (Mismo criterio que la clasificación VIP.)"""
    paises = set()
    tienda_emails = set()
    for t in extract_tiendas(contact):
        tienda_emails.add(t["email"])
        if t["pais"]:
            paises.add(t["pais"])
        for mr in maestro.get(t["email"], []):
            if mr.get("pais"):
                paises.add(mr["pais"])
    em_principal = (contact.get("email") or "").strip().lower()
    if em_principal and em_principal not in tienda_emails:
        for mr in maestro.get(em_principal, []):
            if mr.get("pais"):
                paises.add(mr["pais"])
    return sorted(paises)


def first_seen_month(maestro, email):
    """Devuelve el mes (YYYY-MM) más antiguo en que aparece este correo en el maestro."""
    rows = maestro.get(email, [])
    if not rows: return ""
    return min(r["mes"] for r in rows if r.get("mes"))


def clasificar_nivel(ped_values):
    """Aplica las reglas de escalafón a los pedidos mensuales de un usuario.
    `ped_values` es un iterable de pedidos por mes (ej. ped_mes.values()).
    Devuelve: Diamante / Platino / Oro / Plata / Bronce / Sin clasificar.

    Reglas (idénticas a las del loop de clasificación VIP):
      - <2 meses con ventas o suma top-2 < 60  → Sin clasificar
      - exactamente 2 meses con ventas         → Bronce
      - 3+ meses con ventas, según suma top-3:
          >= 15000 Diamante · >= 3000 Platino · >= 900 Oro · >= 300 Plata · resto Bronce
    """
    vals = list(ped_values)
    active = sum(1 for v in vals if v > 0)
    sv = sorted(vals, reverse=True)
    top1 = sv[0] if sv else 0
    top2 = sv[1] if len(sv) > 1 else 0
    top3 = sv[2] if len(sv) > 2 else 0
    suma_top2 = top1 + top2
    suma_top3 = top1 + top2 + top3
    if active < 2 or suma_top2 < 60:
        return "Sin clasificar"
    if active < 3:
        return "Bronce"
    if suma_top3 >= 15000: return "Diamante"
    if suma_top3 >= 3000:  return "Platino"
    if suma_top3 >= 900:   return "Oro"
    if suma_top3 >= 300:   return "Plata"
    return "Bronce"


VIP_TAG = "comunidad vip new"
SINFORM_TAG = "vip sin form"

def audiencias_de(tags_lower):
    """Devuelve la lista de audiencias a las que pertenece un contacto, según
    sus tags. Categorías SaaS del dashboard."""
    has_vip = VIP_TAG in tags_lower
    has_sf  = SINFORM_TAG in tags_lower
    has_m   = TAG_MASTER in tags_lower
    has_i   = TAG_INICIACION in tags_lower
    aud = []
    if has_vip: aud.append("vip")
    if has_m or has_i: aud.append("estudiantes")
    if has_sf: aud.append("pendientes")
    if has_vip or has_sf: aud.append("todos_vip")   # "Todos VIP" = new o sin form
    return aud

def compute_all():
    all_contacts = json.load(open(RAW))
    # Universo de trabajo: cualquier contacto que pertenezca a alguna audiencia
    # del dashboard (VIP new, Estudiantes escala/iniciación, o VIP sin form).
    def _auds(c): return audiencias_de([t.lower() for t in (c.get("tags") or [])])
    contacts = [c for c in all_contacts if _auds(c)]
    excluidos = len(all_contacts) - len(contacts)
    print(f"Contactos GHL totales: {len(all_contacts)}  ·  En alguna audiencia: {len(contacts)}  ·  Excluidos: {excluidos}")
    maestro, months = load_maestro_window()

    # Fuentes
    excels_total = 0
    excels_breakdown = []
    if os.path.isdir(ORIG):
        # Walk recursivo: la fuente puede ser un Drive sincronizado con
        # estructura <Año>/<MesAño>/, no plana. Solo carpetas hoja con .xlsx
        # cuentan, deduplicadas por nombre de carpeta para evitar contar
        # el mismo Mes-Año dos veces si aparece en varias ramas.
        seen_folders = set()
        for dirpath, dirnames, filenames in os.walk(ORIG, followlinks=True):
            dirnames.sort()
            n = len([f for f in filenames if f.endswith(".xlsx") and not f.startswith("~$")])
            if n == 0:
                continue
            sub = os.path.basename(dirpath)
            if sub in seen_folders:
                continue
            seen_folders.add(sub)
            excels_breakdown.append({"carpeta": sub, "n": n}); excels_total += n
    paises_set = set()
    correos_maestro = set()
    pedidos_comunidad_total = 0   # toda la comunidad Dropi (para Fuentes de datos)
    for em, rows in maestro.items():
        correos_maestro.add(em)
        for r in rows:
            if r["pais"]: paises_set.add(r["pais"])
            pedidos_comunidad_total += r["pedidos"]
    paises_lista = sorted(paises_set)

    # Procesar contactos
    LATEST = months[-1]
    PREVIOUS = months[-2] if len(months) > 1 else None
    ACTIVE_RECENT = months[-2:]  # 2 últimos meses

    # Programa formativo derivado de los TAGS, no del custom field.
    # "escala" → Master Escala, "iniciacion" → Iniciación Escala.
    # Posibles: solo Master, solo Iniciación, Ambos, Sin programa.
    PROGRAMA = {}
    sin_programa_ids = []
    for c in contacts:
        tags = [t.lower() for t in (c.get("tags") or [])]
        has_m = TAG_MASTER in tags
        has_i = TAG_INICIACION in tags
        if has_m and has_i: PROGRAMA[c["id"]] = "Ambos"
        elif has_m: PROGRAMA[c["id"]] = "Master Escala"
        elif has_i: PROGRAMA[c["id"]] = "Iniciación Escala"
        else:
            PROGRAMA[c["id"]] = "Sin programa"
            sin_programa_ids.append({
                "contact_id": c["id"],
                "nombre": c.get("contactName") or "",
                "email": (c.get("email") or "").lower(),
                "tags": [t for t in (c.get("tags") or []) if not t.lower().startswith("wa:")],
            })

    users_with_tienda = []
    multipais_count = 0
    activos_2m = 0
    desaparecidos = 0
    recuperados = 0
    pedidos_vip_total = 0   # SOLO pedidos de los VIPs (lo que va en el card)
    for c in contacts:
        tiendas = extract_tiendas(c)
        # Si no hay tiendas registradas, igual incluirlo como "Sin tienda" para que el equipo lo vea.
        sin_tienda = not tiendas
        ped_mes = {m: 0 for m in months}
        ent_mes = {m: 0 for m in months}
        dev_mes = {m: 0 for m in months}
        paises_lista = []   # una entrada por cada tienda (puede repetirse país si tiene varios correos en el mismo país)
        paises_set = set()  # para flag "multi-país" (países únicos)
        tienda_emails = set()
        # Desglose por país (canónico) → {mes: valor}, para permitir filtrar
        # la vista por país y ver solo las ventas realizadas en ese país.
        ped_mes_pais = {}   # {pais_canonico: {mes: pedidos}}
        ent_mes_pais = {}
        dev_mes_pais = {}
        # Fase 1: recolectar países y emails únicos (extract_tiendas puede devolver
        # el MISMO email en dos slots con países distintos — ej. un correo Dropi que
        # opera en Colombia y Ecuador — y en ese caso ambas tiendas deben aparecer
        # en la lista de países, pero maestro.get(email) devuelve TODAS las filas
        # de ese email así que sumarlo dos veces es doble conteo).
        for t in tiendas:
            tienda_emails.add(t["email"])
            if t["pais"]:
                paises_lista.append(t["pais"])
                paises_set.add(t["pais"])
        # Fase 2: sumar el maestro UNA SOLA VEZ por email único.
        for em in tienda_emails:
            for mr in maestro.get(em, []):
                if mr["mes"] in ped_mes:
                    ped_mes[mr["mes"]] += mr["pedidos"]
                    ent_mes[mr["mes"]] += mr["entregados"]
                    dev_mes[mr["mes"]] += mr["devoluciones"]
                    # Acumular por país canónico (colapsa 'COLOMBIA'/'Colombia'/'colombia')
                    pais_c = pais_canonico(mr.get("pais") or "")
                    if pais_c:
                        ped_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["pedidos"]
                        ent_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["entregados"]
                        dev_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["devoluciones"]
        # También las ventas asociadas al EMAIL PRINCIPAL si está en el maestro
        # (caso VIP SIN FORM: vendedores creados sin tiendas, su email vende en Dropi).
        em_principal = (c.get("email") or "").strip().lower()
        if em_principal and em_principal not in tienda_emails and em_principal in maestro:
            for mr in maestro.get(em_principal, []):
                if mr["mes"] in ped_mes:
                    ped_mes[mr["mes"]] += mr["pedidos"]
                    ent_mes[mr["mes"]] += mr["entregados"]
                    dev_mes[mr["mes"]] += mr["devoluciones"]
                    pais_c = pais_canonico(mr.get("pais") or "")
                    if pais_c:
                        ped_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["pedidos"]
                        ent_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["entregados"]
                        dev_mes_pais.setdefault(pais_c, {m:0 for m in months})[mr["mes"]] += mr["devoluciones"]
                if mr.get("pais") and mr["pais"] not in paises_set:
                    paises_lista.append(mr["pais"]); paises_set.add(mr["pais"])
            if maestro.get(em_principal):
                sin_tienda = False   # tiene ventas vía email principal
        pedidos_vip_total += sum(ped_mes.values())
        active = sum(1 for v in ped_mes.values() if v > 0)
        sv = sorted(ped_mes.values(), reverse=True)
        top1 = sv[0]; top2 = sv[1] if len(sv) > 1 else 0; top3 = sv[2] if len(sv) > 2 else 0
        suma_top2 = top1 + top2
        suma_top3 = top1 + top2 + top3
        nivel = clasificar_nivel(ped_mes.values())

        # alertas
        en_riesgo = (
            len(months) >= 3
            and all(ped_mes[m] == 0 for m in months[-3:])
            and any(ped_mes[m] > 0 for m in months[:-3])
        )
        recientes_count = sum(1 for m in ACTIVE_RECENT if ped_mes[m] > 0)
        if recientes_count > 0:
            activos_2m += 1
        # desaparecido: tenía pedidos en algún mes previo PERO 0 en ACTIVE_RECENT
        if (any(ped_mes[m] > 0 for m in months[:-2]) if len(months) > 2 else False) \
           and all(ped_mes[m] == 0 for m in ACTIVE_RECENT):
            desaparecidos += 1
        # recuperado: 0 en penúltimo mes pero >0 en último
        if PREVIOUS and ped_mes.get(PREVIOUS, 0) == 0 and ped_mes.get(LATEST, 0) > 0:
            recuperados += 1
        if len(paises_set) > 1:
            multipais_count += 1

        total_ped = sum(ped_mes.values())
        total_dev = sum(dev_mes.values())
        pct_dev = (total_dev / total_ped * 100) if total_ped > 0 else 0
        # Alerta tipo (excluyentes en orden de severidad):
        #   "Eliminado"     → últimos 3+ meses en 0 (con actividad previa)
        #   "Riesgo"        → últimos 2 meses en 0 (con actividad previa)
        #   "Desaparecido"  → último mes en 0 (con actividad previa)
        # Y crítica (puede coexistir):
        #   "Crítica"       → % devolución > 50%
        # Países donde el email vende (según Dropi) pero que NO están declarados
        # como tienda en GHL. Indica que hay que actualizar el formulario del
        # contacto o que Dropi tiene el país mal clasificado.
        _paises_declarados_set = {pais_canonico(t["pais"]) for t in tiendas if t["pais"]}
        paises_no_declarados = sorted(
            p for p, pm in ped_mes_pais.items()
            if p not in _paises_declarados_set and any(v > 0 for v in pm.values())
        )
        tiene_no_declarada = bool(paises_no_declarados)
        alerta_tipo = None
        # 1. Huérfana: tiene tienda pero NUNCA vendió en toda la ventana
        if not sin_tienda and total_ped == 0:
            alerta_tipo = "Huérfana"
        # 2-4. Eliminado / Riesgo / Desaparecido (requieren actividad previa)
        if alerta_tipo is None and len(months) >= 3:
            last3 = months[-3:]
            if all(ped_mes[m] == 0 for m in last3) and any(ped_mes[m] > 0 for m in months[:-3]):
                alerta_tipo = "Eliminado"
        if alerta_tipo is None and len(months) >= 2:
            last2 = months[-2:]
            if all(ped_mes[m] == 0 for m in last2) and any(ped_mes[m] > 0 for m in months[:-2]):
                alerta_tipo = "Riesgo"
        if alerta_tipo is None and len(months) >= 1:
            if ped_mes[months[-1]] == 0 and any(ped_mes[m] > 0 for m in months[:-1]):
                alerta_tipo = "Desaparecido"
        # 5. Crítica (puede sumarse a las otras, pero solo es etiqueta primaria si no hay otra)
        es_critica = pct_dev > 50 and total_ped > 0
        if alerta_tipo is None and es_critica:
            alerta_tipo = "Crítica"
        # 6. Tienda no declarada (solo si no hay otra alerta más grave; el flag
        # separado `tiene_no_declarada` permite marcarla en la UI aunque exista otra).
        if alerta_tipo is None and tiene_no_declarada:
            alerta_tipo = "Tienda no declarada"

        # Semáforo: verde=ok, amarillo=desap/crit, naranja=riesgo, rojo=eliminado/sin nivel/sin tienda
        if sin_tienda: semaforo = "rojo"
        elif alerta_tipo == "Eliminado": semaforo = "rojo"
        elif alerta_tipo == "Riesgo": semaforo = "naranja"
        elif alerta_tipo in ("Desaparecido","Crítica"): semaforo = "amarillo"
        elif nivel == "Sin clasificar": semaforo = "gris"
        else: semaforo = "verde"
        users_with_tienda.append({
            "cid": c["id"],
            "nombre": c.get("contactName") or "",
            "email": (c.get("email") or "").lower(),
            "telefono": c.get("phone") or "",
            "n_tiendas": len(tiendas),
            "paises": paises_lista,
            "paises_unicos": sorted(paises_set),
            # Países DECLARADOS en GHL (via slots país de tiendas). Se usa para
            # marcar ⚠ ventas en países no declarados (donde el correo aparece
            # en Dropi pero el usuario no tiene tienda registrada allí).
            "paises_declarados": sorted({pais_canonico(t["pais"]) for t in tiendas if t["pais"]}),
            # Países donde el email vende en Dropi pero NO están declarados en GHL.
            "paises_no_declarados": paises_no_declarados,
            "tiene_no_declarada": tiene_no_declarada,
            "tiendas_detalle": [
                {"email": t["email"], "pais": t["pais"], "primera_vez": first_seen_month(maestro, t["email"])}
                for t in tiendas
            ],
            "ped_mes": ped_mes,
            "ent_mes": ent_mes,
            "dev_mes": dev_mes,
            # Desglose por país canónico → {mes: valor}. Se usa cuando el
            # usuario filtra por país en el dashboard, para mostrar SOLO
            # las ventas realizadas en ese país en vez del total del usuario.
            "ped_mes_pais": ped_mes_pais,
            "ent_mes_pais": ent_mes_pais,
            "dev_mes_pais": dev_mes_pais,
            "total_pedidos": total_ped,
            "pct_dev": round(pct_dev, 1),
            "active": active,
            "top1": top1, "top2": top2, "top3": top3,
            "suma_top2": suma_top2,
            "suma_top3": suma_top3,
            "nivel": nivel,
            "programa": PROGRAMA.get(c["id"], "Sin programa"),
            "en_riesgo": en_riesgo,
            "semaforo": semaforo,
            "sin_tienda": sin_tienda,
            "alerta_tipo": alerta_tipo,
            "es_critica": es_critica,
            # Audiencias SaaS: ["vip","estudiantes","pendientes","todos_vip"]
            "aud": audiencias_de([t.lower() for t in (c.get("tags") or [])]),
        })

    # Programa formativo (derivado de tags)
    prog_counts = Counter(PROGRAMA.values())
    solo_master = prog_counts.get("Master Escala", 0)
    solo_iniciacion = prog_counts.get("Iniciación Escala", 0)
    ambos = prog_counts.get("Ambos", 0)
    sin_programa = prog_counts.get("Sin programa", 0)

    # Distribución por nivel
    dist = Counter(u["nivel"] for u in users_with_tienda)
    clasificados = sum(dist[t] for t in TIER_ORDER if t != "Sin clasificar")
    nivel_pedidos = defaultdict(int)
    for u in users_with_tienda:
        nivel_pedidos[u["nivel"]] += sum(u["ped_mes"].values())

    # Semáforo (4 categorías, alineadas con los nuevos tipos de alerta)
    verde     = sum(1 for u in users_with_tienda if u["semaforo"] == "verde")
    amarillo  = sum(1 for u in users_with_tienda if u["semaforo"] in ("amarillo","naranja"))
    rojo      = sum(1 for u in users_with_tienda if u["semaforo"] == "rojo")
    sin_activ = sum(1 for u in users_with_tienda if u["semaforo"] == "gris")

    sin_alertas = verde

    # Series para charts mensuales
    pedidos_por_mes = {m: 0 for m in months}
    activos_por_mes = {m: 0 for m in months}
    for u in users_with_tienda:
        for m in months:
            v = u["ped_mes"].get(m, 0)
            pedidos_por_mes[m] += v
            if v > 0:
                activos_por_mes[m] += 1

    # Diagnóstico (del Excel)
    diagnostico = {"huerfanas":[],"duplicados":[],"riesgo":[],"capeados":[]}
    if os.path.isfile(CLAS):
        wb = openpyxl.load_workbook(CLAS, read_only=True, data_only=True)
        if "TIENDAS_NO_ENCONTRADAS" in wb.sheetnames:
            for r in list(wb["TIENDAS_NO_ENCONTRADAS"].iter_rows(values_only=True))[1:]:
                if not r[0]: continue
                diagnostico["huerfanas"].append({"contact_id":r[0],"contact_email":r[1] or "","nombre":r[2] or "",
                                  "label":r[3] or "","tienda_email":r[4] or "","pais":r[5] or ""})
        if "CORREOS_DUPLICADOS_GHL" in wb.sheetnames:
            for r in list(wb["CORREOS_DUPLICADOS_GHL"].iter_rows(values_only=True))[1:]:
                if not r[0]: continue
                diagnostico["duplicados"].append({"contact_id":r[0],"correo":r[3] or "","pais":r[4] or "",
                                                   "nombre":r[2] or ""})
        if "RIESGO_ELIMINACION" in wb.sheetnames:
            for r in list(wb["RIESGO_ELIMINACION"].iter_rows(values_only=True))[1:]:
                if not r[0]: continue
                diagnostico["riesgo"].append({"contact_id":r[0],"nombre":r[1] or "","nivel":r[3] or ""})
        if "CAPEADOS_2_MESES" in wb.sheetnames:
            for r in list(wb["CAPEADOS_2_MESES"].iter_rows(values_only=True))[1:]:
                if not r[0]: continue
                diagnostico["capeados"].append({"contact_id":r[0],"nombre":r[1] or "",
                                                "meses":r[3] or 0,"top3":r[4] or 0,"hubiera_sido":r[5] or ""})

    # Por país
    por_pais = Counter()
    for u in users_with_tienda:
        for p in u["paises"]:
            por_pais[p] += 1

    # ============================================================
    # MÉTRICAS · datasets independientes del VIP
    # No tocan la lógica de niveles ni los stats de VIP.
    # ============================================================

    # Universo total de emails GHL: principal + emails de las 10 tiendas
    # + sufijos de teléfono (para deduplicar también por teléfono).
    def _tel_suffix(raw):
        """Últimos 8 dígitos del teléfono (número nacional, dedup country-agnostic).
        Se usan 8 y no 10 porque GHL guarda E.164 (+57..., +56...) y Dropi guarda
        el número local sin código país; los últimos 8 dígitos (la parte de
        suscriptor) coinciden entre ambos formatos. '' si <8 díg."""
        if not raw: return ""
        d = re.sub(r"\D", "", str(raw))
        return d[-8:] if len(d) >= 8 else ""
    ghl_emails_set = set()
    ghl_phone_suffixes = set()
    for c in all_contacts:
        em_p = (c.get("email") or "").strip().lower()
        if em_p: ghl_emails_set.add(em_p)
        suf_c = _tel_suffix(c.get("phone"))
        if suf_c: ghl_phone_suffixes.add(suf_c)
        cf_all = {f["id"]: f.get("value") for f in c.get("customFields", [])}
        for fid in TIENDA_IDS:
            v = cf_all.get(fid)
            if v and isinstance(v, str) and "@" in v:
                ghl_emails_set.add(v.strip().lower())

    # Vista 1: contactos con tag escala/iniciacion que NO tienen "comunidad vip new"
    # Vista 2: clasificación de TODOS los contactos GHL por programa
    met_sin_vip = []
    met_programas = []
    met_estudiantes = []   # estudiantes (tag Escala/Iniciación) + sus pedidos Dropi
    cnt_master = 0
    cnt_iniciacion = 0
    cnt_ambos = 0
    cnt_sin_prog = 0
    for c in all_contacts:
        tags_l = [t.lower() for t in (c.get("tags") or [])]
        has_m = TAG_MASTER in tags_l
        has_i = TAG_INICIACION in tags_l
        has_vip = VIP_TAG in tags_l
        if has_m and has_i:
            prog = "Ambos"; cnt_ambos += 1
        elif has_m:
            prog = "Master Escala"; cnt_master += 1
        elif has_i:
            prog = "Iniciación Escala"; cnt_iniciacion += 1
        else:
            prog = "Sin programa"; cnt_sin_prog += 1
        paises_c = derivar_paises(c, maestro)
        item = {
            "cid": c["id"],
            "nombre": c.get("contactName") or "",
            "email": (c.get("email") or "").lower(),
            "telefono": c.get("phone") or "",
            "programa": prog,
            "paises": paises_c,
            "tiene_vip_new": has_vip,
        }
        # Vista 2 incluye TODOS (incluyendo Sin programa para tener conteo completo)
        met_programas.append(item)
        # Vista 1 solo los que tienen programa pero NO tienen VIP new
        if (has_m or has_i) and not has_vip:
            met_sin_vip.append(item)
        # Vista "Estudiantes": cualquiera con programa formativo (Escala/Iniciación/Ambos),
        # enriquecido con sus pedidos de Dropi (vía tiendas cruzadas con el maestro).
        if has_m or has_i:
            tiendas_e = extract_tiendas(c)
            ped_mes_e = {mo: 0 for mo in months}
            ped_mes_pais_e = {}   # {pais_canonico: {mes: pedidos}} para filtro
            # Dedup por email único (ver fix del mismo bug arriba: extract_tiendas
            # puede devolver un correo en dos slots con países distintos → sumar
            # maestro.get(email) dos veces es doble conteo).
            for em in {t["email"] for t in tiendas_e}:
                for mr in maestro.get(em, []):
                    if mr["mes"] in ped_mes_e:
                        ped_mes_e[mr["mes"]] += mr["pedidos"]
                        pais_c = pais_canonico(mr.get("pais") or "")
                        if pais_c:
                            ped_mes_pais_e.setdefault(pais_c, {mo:0 for mo in months})[mr["mes"]] += mr["pedidos"]
            total_ped_e = sum(ped_mes_e.values())
            met_estudiantes.append({
                "cid": c["id"],
                "nombre": c.get("contactName") or "",
                "email": (c.get("email") or "").lower(),
                "telefono": c.get("phone") or "",
                "programa": prog,
                "tiene_vip_new": has_vip,
                "paises": paises_c,
                "n_tiendas": len(tiendas_e),
                "ped_mes": ped_mes_e,
                "ped_mes_pais": ped_mes_pais_e,
                "total_pedidos": total_ped_e,
                "tiene_ventas": total_ped_e > 0,
                "nivel": clasificar_nivel(ped_mes_e.values()),
            })
    met_estudiantes.sort(key=lambda x: -x["total_pedidos"])

    # Vista 3: emails Dropi que NO están en GHL.
    # Excluye los que su EMAIL ya está en GHL Y los que su TELÉFONO ya está en GHL
    # (estos existen en GHL bajo otro email; no son realmente "sin GHL").
    met_dropi_sin_ghl = []
    excluidos_por_tel = 0
    for em, rows in maestro.items():
        if not em or em.strip().lower() in ghl_emails_set:
            continue
        tel_em = next((r.get("telefono") for r in rows if r.get("telefono")), "") or ""
        if _tel_suffix(tel_em) in ghl_phone_suffixes and _tel_suffix(tel_em):
            excluidos_por_tel += 1
            continue
        em_low = em.strip().lower()
        total_ped = sum(r["pedidos"] for r in rows)
        total_ent = sum(r["entregados"] for r in rows)
        total_dev = sum(r["devoluciones"] for r in rows)
        paises_em = sorted({r["pais"] for r in rows if r.get("pais")})
        meses_activos_em = sorted({r["mes"] for r in rows if r.get("pedidos", 0) > 0})
        nombre_raw = next((r.get("nombre") for r in rows if r.get("nombre")), "") or ""
        nombre_em = normalizar_nombre_dropi(nombre_raw)
        ped_mes_em = {m: 0 for m in months}
        for r in rows:
            if r["mes"] in ped_mes_em:
                ped_mes_em[r["mes"]] += r["pedidos"]
        met_dropi_sin_ghl.append({
            "email": em_low,
            "nombre": nombre_em,
            "telefono": tel_em,
            "paises": paises_em,
            # No están en GHL → no tienen tags de programa formativo.
            "programa": "Sin programa",
            "total_pedidos": total_ped,
            "total_entregados": total_ent,
            "total_devoluciones": total_dev,
            "n_meses_activos": len(meses_activos_em),
            "tiene_ventas": total_ped > 0,
            "ped_mes": ped_mes_em,
            # Escalafón calculado con las mismas reglas VIP, aunque no estén en GHL.
            "nivel": clasificar_nivel(ped_mes_em.values()),
        })
    # Orden: primero por escalafón (Diamante→…→Sin clasificar), luego por ventas desc.
    met_dropi_sin_ghl.sort(key=lambda x: (TIER_ORDER.index(x["nivel"]) if x["nivel"] in TIER_ORDER else 99, -x["total_pedidos"]))

    # Vista 4: duplicados potenciales
    # Detecta tiendas cuyo correo coincide con el email principal de OTRO contacto.
    # Ej: Contacto A tiene email 'diego@gmail.com'.
    #     Contacto B (Diego Adolfo) tiene email principal 'diegoaforero@gmail.com'
    #     y como Tienda 3 'diego@gmail.com'.
    #     -> A y B son posibles duplicados o cuentas compartidas.
    email_to_contact = {}
    for c in all_contacts:
        em_p = (c.get("email") or "").strip().lower()
        if em_p:
            # Si hay varios contactos con el mismo email principal (raro), nos quedamos con el primero
            if em_p not in email_to_contact:
                email_to_contact[em_p] = {
                    "cid": c["id"],
                    "nombre": c.get("contactName") or "",
                    "telefono": c.get("phone") or "",
                }
    met_duplicados = []
    SLOT_LABEL = {fid: f"Tienda {i+1}" for i, fid in enumerate(sorted(TIENDA_IDS))}
    # Mantener orden estable
    TIENDA_IDS_ORDER = list(TIENDA_IDS)
    for c in all_contacts:
        cid = c["id"]
        em_principal = (c.get("email") or "").strip().lower()
        nombre = c.get("contactName") or ""
        telefono = c.get("phone") or ""
        cf_all = {f["id"]: f.get("value") for f in c.get("customFields", [])}
        for i, fid in enumerate(TIENDA_IDS_ORDER):
            v = cf_all.get(fid)
            if not v or not isinstance(v, str) or "@" not in v:
                continue
            t_em = v.strip().lower()
            other = email_to_contact.get(t_em)
            if other and other["cid"] != cid:
                pais = cf_all.get(TIENDA_PAIS_IDS.get(fid, "")) or ""
                met_duplicados.append({
                    "cid": cid,
                    "nombre": nombre,
                    "email_principal": em_principal,
                    "telefono": telefono,
                    "tienda_slot": f"Tienda {i+1}",
                    "tienda_email": t_em,
                    "tienda_pais": pais,
                    "otro_cid": other["cid"],
                    "otro_nombre": other["nombre"],
                    "otro_telefono": other["telefono"],
                })
    # Ordenar por nombre del contacto B (el que tiene la tienda)
    met_duplicados.sort(key=lambda x: (x["nombre"] or "").lower())

    return {
        "meta": {
            "ultimo_mes": LATEST,
            "ultimo_mes_label": mes_label(LATEST),
            "generated": datetime.now().strftime("%d de %B de %Y, %H:%M"),
            "generated_iso": datetime.now().isoformat(),
            "ventana": months,
            "ventana_labels": [mes_label(m) for m in months],
        },
        "fuentes": {
            "excels_total": excels_total,
            "excels_breakdown": excels_breakdown,
            "meses": months,
            "paises": paises_lista,
            "correos_maestro": len(correos_maestro),
            "pedidos_total": pedidos_comunidad_total,
            "contactos_ghl": len(contacts),
            "contactos_con_tienda": sum(1 for u in users_with_tienda if not u["sin_tienda"]),
        },
        "stats": {
            "usuarios_totales": len(contacts),
            "clasificados_vip": clasificados,
            "total_pedidos": pedidos_comunidad_total,
            "total_pedidos_vip": pedidos_vip_total,
            "multi_pais": multipais_count,
            "sin_alertas": sin_alertas,
            "activos_2_meses": activos_2m,
            "desaparecidos": desaparecidos,
            "recuperados": recuperados,
        },
        "programa": {
            "master": solo_master,
            "iniciacion": solo_iniciacion,
            "ambos": ambos,
            "sin_programa": sin_programa,
            "sin_programa_ids": sin_programa_ids,
        },
        "distribucion": {t: {"n": dist.get(t,0), "pedidos": nivel_pedidos.get(t,0)} for t in TIER_ORDER},
        "semaforo": {"verde": verde, "amarillo": amarillo, "rojo": rojo, "sin_actividad": sin_activ},
        "pedidos_por_mes": pedidos_por_mes,
        "activos_por_mes": activos_por_mes,
        "por_pais": dict(sorted(por_pais.items(), key=lambda x:-x[1])),
        "usuarios": users_with_tienda,
        "diagnostico": diagnostico,
        "metricas": {
            "ghl_total": len(all_contacts),
            "ghl_emails_universo": len(ghl_emails_set),
            "dropi_emails_total": len(maestro),
            "master_total": cnt_master,
            "iniciacion_total": cnt_iniciacion,
            "ambos_total": cnt_ambos,
            "sin_programa_total": cnt_sin_prog,
            "sin_vip_total": len(met_sin_vip),
            "dropi_sin_ghl_total": len(met_dropi_sin_ghl),
            "dropi_sin_ghl_con_ventas": sum(1 for x in met_dropi_sin_ghl if x["tiene_ventas"]),
            "dropi_sin_ghl_sin_ventas": sum(1 for x in met_dropi_sin_ghl if not x["tiene_ventas"]),
            "sin_comunidad_vip": met_sin_vip,
            "programas": met_programas,
            "estudiantes": met_estudiantes,
            "estudiantes_total": len(met_estudiantes),
            "estudiantes_vip": sum(1 for x in met_estudiantes if x["tiene_vip_new"]),
            "estudiantes_no_vip": sum(1 for x in met_estudiantes if not x["tiene_vip_new"]),
            "estudiantes_con_ventas": sum(1 for x in met_estudiantes if x["tiene_ventas"]),
            "dropi_sin_ghl": met_dropi_sin_ghl,
            "duplicados": met_duplicados,
            "duplicados_total": len(met_duplicados),
            "duplicados_contactos_unicos": len({d["cid"] for d in met_duplicados}),
        },
        "pagos": load_pagos(),
    }


def load_pagos():
    """Carga pagos.json (generado por extraer_pagos.py). Devuelve [] si no existe
    aún (el backfill inicial puede no haber terminado)."""
    path = os.path.join(HERE, "pagos.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def render_html(data):
    j = json.dumps(data, ensure_ascii=False, default=str)
    return """<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Panel Comunidad VIP — Iván Caicedo</title>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body { background: #06091a; }
  .card { background: linear-gradient(180deg, rgba(255,255,255,0.02), rgba(255,255,255,0.01));
          border: 1px solid rgba(255,255,255,0.06); border-radius: 12px; backdrop-filter: blur(8px); }
  .card:hover { border-color: rgba(255,255,255,0.10); }
  .tab { padding: 8px 14px; font-size: 13px; font-weight: 500; color: #94a3b8; border-bottom: 2px solid transparent; transition: all .15s; }
  .tab:hover { color: #e2e8f0; }
  .tab.active { color: #38bdf8; border-bottom-color: #38bdf8; }
  .cat-btn { padding: 7px 18px; font-size: 13px; font-weight: 600; color: #94a3b8; border-radius: 8px; transition: all .15s;
             background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.08); }
  .cat-btn:hover { color: #e2e8f0; background: rgba(255,255,255,0.06); }
  .cat-btn.active { color: #ffffff; background: linear-gradient(135deg, rgba(34,211,238,0.25) 0%, rgba(59,130,246,0.25) 100%); border-color: rgba(34,211,238,0.55); box-shadow: 0 0 14px rgba(34,211,238,0.15); }
  /* Sidebar SaaS */
  .sidebar { width: 248px; flex-shrink: 0; background: rgba(8,11,26,0.75); border-right: 1px solid rgba(255,255,255,0.06);
             position: sticky; top: 0; height: 100vh; overflow-y: auto; backdrop-filter: blur(8px); }
  .side-cat { display: flex; align-items: center; gap: 10px; width: 100%; text-align: left; padding: 11px 13px;
              font-size: 13px; font-weight: 600; color: #94a3b8; border-radius: 9px; border: 1px solid transparent;
              transition: all .15s; cursor: pointer; line-height: 1.2; }
  .side-cat:hover { color: #e2e8f0; background: rgba(255,255,255,0.06); }
  .side-cat.active { color: #ffffff; background: linear-gradient(135deg, rgba(34,211,238,0.22) 0%, rgba(59,130,246,0.22) 100%);
                     border-color: rgba(34,211,238,0.5); box-shadow: 0 0 14px rgba(34,211,238,0.12); }
  .side-cat .cat-count { margin-left: auto; font-size: 10px; font-weight: 600; color: #64748b;
                         background: rgba(255,255,255,0.05); padding: 1px 7px; border-radius: 999px; }
  @media (max-width: 768px) { .sidebar { width: 64px; } .side-cat .cat-text, .side-cat .cat-count { display: none; }
                              .sidebar .brand-text { display:none; } }
  @keyframes pulse-warn {
    0%, 100% { box-shadow: 0 0 0 0 rgba(248,113,113,0); }
    50%      { box-shadow: 0 0 0 6px rgba(248,113,113,0.25); }
  }
  .pill { padding: 2px 10px; border-radius: 9999px; font-size: 11px; font-weight: 600; border: 1px solid; display: inline-block; }
  .neon-cyan { color: #22d3ee; }
  .neon-yellow { color: #facc15; }
  .neon-green { color: #4ade80; }
  .neon-pink { color: #f472b6; }
  .neon-red { color: #f87171; }
  .neon-violet { color: #a78bfa; }
  .neon-orange { color: #fb923c; }
  .neon-blue { color: #60a5fa; }
  table { border-collapse: collapse; }
  tr.hover-row:hover { background: rgba(255,255,255,0.02); }
  .scrollable { max-height: 600px; overflow-y: auto; }
  .scrollable::-webkit-scrollbar { width: 8px; }
  .scrollable::-webkit-scrollbar-track { background: transparent; }
  .scrollable::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.1); border-radius: 4px; }
</style>
</head>
<body class="text-slate-200 min-h-screen font-sans">

<div class="flex min-h-screen">
  <!-- SIDEBAR -->
  <aside class="sidebar flex flex-col">
    <div class="px-5 py-5 flex items-center gap-3 border-b border-white/5">
      <div class="w-10 h-10 rounded-lg bg-gradient-to-br from-cyan-500/30 to-blue-700/30 border border-cyan-500/30 flex items-center justify-center text-xs font-bold flex-shrink-0">VIP</div>
      <div class="brand-text min-w-0">
        <div class="text-sm font-bold leading-tight truncate">Comunidad <span class="neon-cyan">VIP</span></div>
        <div class="text-[10px] text-slate-500">Iván Caicedo</div>
      </div>
    </div>
    <nav class="flex-1 p-3 flex flex-col gap-1.5" id="categories"></nav>
    <div class="p-3 border-t border-white/5">
      <button id="btn-refresh" onclick="location.reload()"
              class="cat-btn w-full flex items-center justify-center gap-2"
              title="Recargar para obtener los datos más recientes">🔄 <span class="cat-text">Actualizar</span></button>
      <div class="text-[10px] text-slate-600 mt-2 px-1 leading-snug" id="header-meta"></div>
    </div>
  </aside>

  <!-- MAIN COLUMN -->
  <div class="flex-1 min-w-0 flex flex-col">
    <!-- TABS -->
    <nav class="border-b border-white/5 px-6 flex flex-wrap gap-1 bg-[#070b1a]/40 sticky top-0 z-20 backdrop-blur" id="tabs"></nav>
    <!-- CONTENT -->
    <main class="p-6 flex-1" id="main-content"></main>
    <footer class="text-center text-xs text-slate-700 py-8">
      Generado por <code>generar_dashboard.py</code> · datos en vivo de GHL
    </footer>
  </div><!-- /main column -->
</div><!-- /flex -->

<!-- MODAL FICHA -->
<div id="ficha-modal" class="hidden fixed inset-0 z-50 bg-black/70 backdrop-blur-sm overflow-y-auto" onclick="if(event.target===this)cerrarFicha()">
  <div class="max-w-3xl mx-auto my-6 p-4">
    <div class="flex items-center justify-between mb-3">
      <h2 class="text-sm font-semibold neon-cyan uppercase tracking-wider">Ficha del miembro</h2>
      <button onclick="cerrarFicha()" class="text-slate-400 hover:text-white text-2xl leading-none">×</button>
    </div>
    <div id="ficha-content"></div>
  </div>
</div>

<script>
const DATA = __DATA_JSON__;
const tierColor = {
  "Diamante":"bg-cyan-500/20 text-cyan-300 border-cyan-500/40",
  "Platino":"bg-violet-500/20 text-violet-300 border-violet-500/40",
  "Oro":"bg-amber-500/20 text-amber-300 border-amber-500/40",
  "Plata":"bg-slate-400/20 text-slate-200 border-slate-400/40",
  "Bronce":"bg-orange-500/20 text-orange-300 border-orange-500/40",
  "Sin clasificar":"bg-slate-700/40 text-slate-500 border-slate-600/40",
};
const TIER_COLORS_HEX = {
  "Diamante":"#22d3ee","Platino":"#a78bfa","Oro":"#facc15","Plata":"#cbd5e1","Bronce":"#fb923c","Sin clasificar":"#475569",
};
const TIER_ORDER = ["Diamante","Platino","Oro","Plata","Bronce","Sin clasificar"];
const fmt = n => (n||0).toLocaleString("es-CO");
function tc(s) {
  if (!s) return s;
  return s.toLowerCase().split(/(\s+)/).map(w => {
    if (!w || !w.trim()) return w;
    return w.charAt(0).toUpperCase() + w.slice(1);
  }).join('');
}

// Normaliza un nombre de país a su forma canónica: quita acentos, colapsa
// espacios y devuelve Title Case. "COLOMBIA" / "colombia" / "Colombia" → "Colombia".
// "COSTARICA" / "Costa Rica" → "Costa Rica".
const _PAIS_CANONICO = {
  "colombia":"Colombia","ecuador":"Ecuador","peru":"Perú","mexico":"México",
  "guatemala":"Guatemala","panama":"Panamá","chile":"Chile","argentina":"Argentina",
  "paraguay":"Paraguay","uruguay":"Uruguay","bolivia":"Bolivia","venezuela":"Venezuela",
  "costarica":"Costa Rica","costa rica":"Costa Rica","honduras":"Honduras",
  "elsalvador":"El Salvador","el salvador":"El Salvador","nicaragua":"Nicaragua",
  "republicadominicana":"República Dominicana","republica dominicana":"República Dominicana",
  "brasil":"Brasil","brazil":"Brasil","españa":"España","espana":"España",
  "estadosunidos":"Estados Unidos","usa":"Estados Unidos",
};
function _paisKey(p) {
  return (p||'').toString().normalize("NFD").replace(/[\u0300-\u036f]/g,'')
    .toLowerCase().trim().replace(/\s+/g,'');
}
function paisNorm(p) {
  const k = _paisKey(p);
  if (!k) return '';
  if (_PAIS_CANONICO[k]) return _PAIS_CANONICO[k];
  // Fallback Title Case si no está en el mapa
  return (p||'').toString().trim().toLowerCase()
    .split(/\s+/).map(w => w.charAt(0).toUpperCase() + w.slice(1)).join(' ');
}
function paisesUnicos(arr) {
  // Colapsa por versión normalizada (case + acento-insensitive)
  const seen = new Map();  // key → nombre canónico
  (arr||[]).forEach(p => {
    const k = _paisKey(p);
    if (k && !seen.has(k)) seen.set(k, paisNorm(p));
  });
  return [...seen.values()].sort();
}
function paisMatch(userPaises, filtro) {
  // Match case + acento-insensitive
  if (!filtro || filtro === 'Todos') return true;
  const k = _paisKey(filtro);
  return (userPaises||[]).some(p => _paisKey(p) === k);
}
// Devuelve una versión del usuario con las métricas subseteadas SOLO al país
// dado. Recalcula ped_mes, ent_mes, dev_mes, total_pedidos, top1/top2/top3,
// suma_top2/3, active y pct_dev usando ped_mes_pais/ent_mes_pais/dev_mes_pais.
// Si el usuario no tiene desglose por país o filtro='Todos' → devuelve u tal cual.
function viewByPais(u, pais) {
  if (!pais || pais === 'Todos') return u;
  const key = _paisKey(pais);
  const src = u.ped_mes_pais || {};
  // buscar la key coincidente con normalización
  const paisMatched = Object.keys(src).find(k => _paisKey(k) === key);
  if (!paisMatched) return u;   // usuario no vendió en ese país
  const pm = src[paisMatched] || {};
  const em = (u.ent_mes_pais && u.ent_mes_pais[paisMatched]) || {};
  const dm = (u.dev_mes_pais && u.dev_mes_pais[paisMatched]) || {};
  const meses = Object.keys(pm);
  const total = meses.reduce((s,m) => s + (pm[m]||0), 0);
  const totalEnt = meses.reduce((s,m) => s + (em[m]||0), 0);
  const totalDev = meses.reduce((s,m) => s + (dm[m]||0), 0);
  const sv = Object.values(pm).sort((a,b) => b - a);
  const top1 = sv[0]||0, top2 = sv[1]||0, top3 = sv[2]||0;
  const active = Object.values(pm).filter(v => v > 0).length;
  const pct = total > 0 ? +(totalDev / total * 100).toFixed(1) : 0;
  return { ...u,
    ped_mes: pm, ent_mes: em, dev_mes: dm,
    total_pedidos: total, top1, top2, top3,
    suma_top2: top1+top2, suma_top3: top1+top2+top3,
    active, pct_dev: pct,
    _filtered_pais: paisMatched,
  };
}

// Dropdown reutilizable de filtro por país. `users` = lista de la vista (cada uno
// con .paises o .paises_unicos). `current` = valor seleccionado. `cls` opcional.
function paisSelectHTML(id, users, current, cls) {
  const all = paisesUnicos((users||[]).flatMap(u => u.paises_unicos || u.paises || []));
  cls = cls || 'bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm text-slate-200 focus:outline-none focus:border-cyan-500';
  return `<select id="${id}" class="${cls}">
    <option value="Todos">Todos los países</option>
    ${all.map(p => `<option value="${p}" ${current===p?'selected':''}>${flag(p)} ${p}</option>`).join('')}
  </select>`;
}

function actualizarFrescura() {
  const gen = new Date(DATA.meta.generated_iso);
  const seg = Math.floor((Date.now() - gen.getTime())/1000);
  // Umbral de "conviene actualizar" según el contexto:
  //  - Versión PUBLICADA (GitHub Pages / embebida en GHL): la rutina publish la
  //    regenera 3x/día (09/14/18h); el mayor hueco normal es de noche (~15h), así
  //    que solo avisamos si pasaron >18h = una publicación falló de verdad.
  //  - Versión LOCAL en vivo (localhost/file): refresca cada 90s; avisamos a los
  //    10 min para detectar si el refresco se detuvo.
  const esLocal = location.protocol === 'file:' ||
                  ['localhost','127.0.0.1'].includes(location.hostname);
  const STALE = esLocal ? 600 : 18*3600;
  let txt;
  if (seg < 60) txt = `hace ${seg}s`;
  else if (seg < 3600) txt = `hace ${Math.floor(seg/60)} min`;
  else txt = `hace ${Math.floor(seg/3600)} h`;
  const color = seg < STALE ? 'text-green-400' : 'text-red-400';
  document.getElementById("header-meta").innerHTML =
    `Último mes cargado: <span class="text-cyan-400">${DATA.meta.ultimo_mes_label}</span>  ·  ` +
    `Datos generados <span class="${color}">${txt}</span>` +
    (seg >= STALE ? ` <span class="text-red-400 font-semibold">· conviene actualizar</span>` : '');
  // Resaltar el botón cuando los datos están viejos
  const btn = document.getElementById('btn-refresh');
  if (btn) {
    if (seg >= STALE) {
      btn.classList.add('active');
      btn.style.animation = 'pulse-warn 2s ease-in-out infinite';
    } else {
      btn.classList.remove('active');
      btn.style.animation = '';
    }
  }
}
actualizarFrescura();
setInterval(actualizarFrescura, 10000);

// Auto-recarga cada 1 hora: el dashboard vive embebido en GHL (iframe), y un
// iframe no se refresca solo aunque la rutina 'publish' suba HTML nuevo. Esto
// garantiza que el visitante nunca vea una versión con más de 1 hora de edad,
// sin importar cómo esté configurado el src del iframe en GHL.
// La persistencia de pestaña/scroll/filtros en localStorage hace que la recarga
// sea transparente para el usuario.
setInterval(() => location.reload(), 60 * 60 * 1000);

const CATEGORIES = [
  {id:"vip",              label:"🏆 VIP"},
  {id:"estudiantes",      label:"🎓 Estudiantes"},
  {id:"pendientes",       label:"📝 Pendientes por formulario"},
  {id:"todos_vip",        label:"⭐ Todos VIP"},
  {id:"otros",            label:"📦 Otros"},
  {id:"app_master_escala",label:"💰 App Master Escala"},
  {id:"config",           label:"📋 Reglas"},
];
// Las 4 audiencias comparten los mismos 3 sub-dashboards (resumen/clasif/alertas),
// que se calculan sobre el subconjunto de usuarios de la categoría activa.
const _AUD_TABS = [
  {id:"resumen", label:"📊 Resumen"},
  {id:"clasif",  label:"🏆 Clasificación"},
  {id:"alertas", label:"⚠ Alertas"},
];
const TABS_BY_CAT = {
  "vip":         _AUD_TABS,
  "estudiantes": [..._AUD_TABS, {id:"met_programas", label:"📊 Master vs Iniciación"}],
  "pendientes":  _AUD_TABS,
  "todos_vip":   _AUD_TABS,
  "otros": [
    {id:"otros_resumen",  label:"📊 Resumen"},
    {id:"met_dropi_ghl",  label:"👻 Lista (Dropi sin GHL)"},
    {id:"met_duplicados", label:"🔁 Posibles duplicados"},
  ],
  "app_master_escala": [
    {id:"pagos_dashboard", label:"💰 Pagos"},
  ],
  "config": [
    {id:"reglas",   label:"📋 Reglas VIP"},
    {id:"consulta", label:"🔎 Consulta individual"},
  ],
};
// Audiencias que usan los sub-dashboards genéricos (id de categoría == flag aud)
const AUDIENCE_CATS = ["vip","estudiantes","pendientes","todos_vip"];
function baseUsers() {
  // Usuarios de la categoría/audiencia activa. Para 'otros'/'config' no aplica.
  if (AUDIENCE_CATS.includes(currentCategory))
    return (DATA.usuarios || []).filter(u => (u.aud||[]).includes(currentCategory));
  return DATA.usuarios || [];
}
function computeStats(users) {
  const months = DATA.meta.ventana;
  const dist = {}; TIER_ORDER.forEach(t => dist[t] = {n:0, pedidos:0});
  let verde=0, amarillo=0, rojo=0, sinact=0;
  let master=0, iniciacion=0, ambos=0, sinprog=0;
  let clasif=0, activos2=0, desap=0, recup=0, multipais=0, totalPed=0;
  const ppm={}, apm={}; months.forEach(m=>{ppm[m]=0; apm[m]=0;});
  const LATEST=months[months.length-1], PREV=months.length>1?months[months.length-2]:null;
  users.forEach(u => {
    if (dist[u.nivel]) { dist[u.nivel].n++; dist[u.nivel].pedidos += (u.total_pedidos||0); }
    if (u.nivel !== 'Sin clasificar') clasif++;
    const sm=u.semaforo;
    if (sm==='verde') verde++; else if (sm==='amarillo'||sm==='naranja') amarillo++;
    else if (sm==='rojo') rojo++; else sinact++;
    const pr=u.programa;
    if (pr==='Master Escala') master++; else if (pr==='Iniciación Escala') iniciacion++;
    else if (pr==='Ambos') ambos++; else sinprog++;
    if ((u.paises_unicos||[]).length>1) multipais++;
    totalPed += (u.total_pedidos||0);
    months.forEach(m => { const v=(u.ped_mes&&u.ped_mes[m])||0; ppm[m]+=v; if(v>0) apm[m]++; });
    const recent=months.slice(-2);
    if (recent.some(m => (u.ped_mes&&u.ped_mes[m]>0))) activos2++;
    if (months.length>2) { const before=months.slice(0,-2);
      if (before.some(m=>u.ped_mes[m]>0) && recent.every(m=>!(u.ped_mes[m]>0))) desap++; }
    if (PREV && !(u.ped_mes[PREV]>0) && (u.ped_mes[LATEST]>0)) recup++;
  });
  return {dist, total:users.length, clasificados:clasif, total_pedidos:totalPed,
          semaforo:{verde, amarillo, rojo, sin_actividad:sinact},
          programa:{master, iniciacion, ambos, sin_programa:sinprog},
          activos_2_meses:activos2, desaparecidos:desap, recuperados:recup, multi_pais:multipais,
          pedidos_por_mes:ppm, activos_por_mes:apm};
}
// --- Persistencia de estado (sobrevive al auto-refresh cada 60s) ---
const STATE_KEY = "dashboard_vip_state_v1";
function loadState() {
  try {
    const s = JSON.parse(localStorage.getItem(STATE_KEY) || "{}");
    return s || {};
  } catch(e) { return {}; }
}
function saveState() {
  try {
    localStorage.setItem(STATE_KEY, JSON.stringify({
      cat: currentCategory,
      tab: currentTab,
      scroll: window.scrollY,
    }));
  } catch(e) {}
}
const _st = loadState();
let currentCategory = (_st.cat && TABS_BY_CAT[_st.cat]) ? _st.cat : "vip";
let _wantedTab = _st.tab || "resumen";
let currentTab = (TABS_BY_CAT[currentCategory].some(t => t.id === _wantedTab))
  ? _wantedTab
  : TABS_BY_CAT[currentCategory][0].id;

function catCount(cat) {
  if (AUDIENCE_CATS.includes(cat)) return (DATA.usuarios||[]).filter(u => (u.aud||[]).includes(cat)).length;
  if (cat === "otros") return (DATA.metricas.dropi_sin_ghl||[]).length;
  return null;
}
function renderCategories() {
  document.getElementById("categories").innerHTML = CATEGORIES.map(c => {
    const n = catCount(c.id);
    const parts = c.label.split(" ");
    const icon = parts[0];
    const text = parts.slice(1).join(" ");
    return `<button class="side-cat ${currentCategory===c.id?'active':''}" data-cat="${c.id}">`
      + `<span class="cat-icon">${icon}</span><span class="cat-text">${text}</span>`
      + (n!==null?`<span class="cat-count">${fmt(n)}</span>`:``)
      + `</button>`;
  }).join('');
  document.querySelectorAll('[data-cat]').forEach(b => b.onclick = () => {
    currentCategory = b.dataset.cat;
    currentTab = TABS_BY_CAT[currentCategory][0].id;
    saveState();
    renderCategories();
    renderTabs();
    render();
  });
}
function renderTabs() {
  const tabs = TABS_BY_CAT[currentCategory] || [];
  document.getElementById("tabs").innerHTML = tabs.map(t =>
    `<button class="tab ${currentTab===t.id?'active':''}" data-tab="${t.id}">${t.label}</button>`
  ).join('');
  document.querySelectorAll('[data-tab]').forEach(b => b.onclick = () => {
    currentTab = b.dataset.tab;
    saveState();
    renderTabs();
    render();
  });
}
// Guardar el scroll antes de cada refresh (el meta refresh dispara beforeunload)
window.addEventListener('beforeunload', saveState);
// Y restaurarlo después del render inicial
window.addEventListener('load', () => {
  if (_st.scroll) {
    setTimeout(() => window.scrollTo(0, _st.scroll), 50);
  }
});
renderCategories();
renderTabs();

function statCard(label, n, sub, accentClass) {
  return `<div class="card p-4">
    <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">${label}</div>
    <div class="text-3xl font-bold mt-1 ${accentClass||'text-slate-100'}">${fmt(n)}</div>
    <div class="text-[11px] text-slate-500 mt-1 leading-tight">${sub}</div>
  </div>`;
}

const CAT_LABEL = {vip:"🏆 VIP", estudiantes:"🎓 Estudiantes", pendientes:"📝 Pendientes por formulario", todos_vip:"⭐ Todos VIP"};
function renderResumen() {
  const users = baseUsers();
  const s = computeStats(users);
  const p = s.programa;
  const ventanaTxt = DATA.meta.ventana_labels.length > 1
    ? `${DATA.meta.ventana_labels[0]} → ${DATA.meta.ventana_labels[DATA.meta.ventana_labels.length-1]}`
    : DATA.meta.ventana_labels[0];
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">${CAT_LABEL[currentCategory]||""} · Resumen</h2>
      <div class="text-xs text-slate-500">${fmt(s.total)} contactos en esta categoría · ventana ${ventanaTxt}</div>
    </div>
    <!-- ROW 1 -->
    <div class="grid grid-cols-2 md:grid-cols-3 lg:grid-cols-6 gap-3">
      ${statCard("Usuarios totales", s.total, "En esta categoría", "neon-cyan")}
      ${statCard("Clasificados", s.clasificados, "Bronce · Plata · Oro · Platino · Diamante", "neon-yellow")}
      ${statCard("Total pedidos", s.total_pedidos, `Acumulado · ${ventanaTxt}`, "neon-green")}
      ${statCard("Multi-país", s.multi_pais, "Con ventas en +1 país", "neon-pink")}
      ${statCard("Sin alertas", s.semaforo.verde, "Clasificados y no en riesgo", "neon-green")}
      ${statCard("Activos últimos 2 meses", s.activos_2_meses, "Con pedidos recientes", "neon-blue")}
    </div>

    <!-- ROW 2 -->
    <div class="grid grid-cols-2 md:grid-cols-3 gap-3 mt-3">
      ${statCard("Desaparecidos", s.desaparecidos, "Activos antes, 0 los últimos 2 meses", "neon-red")}
      ${statCard("Recuperados", s.recuperados, "Estaban inactivos, vendieron en " + DATA.meta.ultimo_mes_label, "neon-violet")}
      ${statCard("En riesgo de salir", s.semaforo.amarillo, "3 meses consecutivos sin vender", "neon-orange")}
    </div>

    <!-- PROGRAMA FORMATIVO -->
    <h2 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mt-8 mb-3">Distribución por programa formativo (según tags GHL)</h2>
    <div class="grid grid-cols-2 md:grid-cols-4 gap-3">
      ${statCard("Master Escala", p.master, "Solo tag 'escala' (sin 'iniciacion')", "neon-violet")}
      ${statCard("Iniciación Escala", p.iniciacion, "Solo tag 'iniciacion' (sin 'escala')", "neon-pink")}
      ${statCard("Ambos programas", p.ambos, "Tienen ambos tags (escala + iniciacion)", "neon-yellow")}
      ${statCard("Sin programa", p.sin_programa, "VIPs sin tag de Master ni de Iniciación", "neon-red")}
    </div>

    <!-- CHARTS · grid 2x2 -->
    <div class="grid grid-cols-1 lg:grid-cols-2 gap-4 mt-8">
      <div class="card p-5">
        <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Distribución VIP</h3>
        <canvas id="chart-donut" height="220"></canvas>
      </div>
      <div class="card p-5">
        <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Semáforo general</h3>
        <canvas id="chart-semaforo" height="220"></canvas>
      </div>
      <div class="card p-5">
        <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Evolución pedidos por mes</h3>
        <canvas id="chart-evolucion" height="220"></canvas>
      </div>
      <div class="card p-5">
        <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Actividad en ventana de evaluación</h3>
        <canvas id="chart-actividad" height="220"></canvas>
      </div>
    </div>

    <!-- TABLA DISTRIBUCION -->
    <div class="card p-5 mt-4">
      <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Detalle por nivel</h3>
      <table class="w-full text-sm">
        <thead class="text-[10px] text-slate-500 uppercase tracking-wider">
          <tr><th class="text-left py-2">Nivel</th><th class="text-right">Usuarios</th><th class="text-right">%</th><th class="text-left pl-6">Reparto</th><th class="text-right">Pedidos acumulados</th></tr>
        </thead>
        <tbody>${renderDistRows()}</tbody>
      </table>
    </div>
  `;
}

function renderDistRows() {
  const d = computeStats(baseUsers()).dist;
  const total = Object.values(d).reduce((s,x)=>s+x.n,0);
  const maxN = Math.max(...Object.values(d).map(x=>x.n)) || 1;
  return TIER_ORDER.map(t => {
    const x = d[t];
    const pct = total ? (x.n*100/total).toFixed(1) : 0;
    const bw = (x.n/maxN*100).toFixed(1);
    const c = TIER_COLORS_HEX[t];
    return `<tr class="border-b border-white/5">
      <td class="py-2"><span class="pill ${tierColor[t]}">${t}</span></td>
      <td class="text-right font-mono text-slate-200">${x.n}</td>
      <td class="text-right text-slate-500">${pct}%</td>
      <td class="pl-6 pr-2 w-2/5"><div class="bg-white/5 h-2 rounded-full overflow-hidden"><div class="h-full rounded-full" style="width:${bw}%;background:${c}"></div></div></td>
      <td class="text-right font-mono text-slate-400">${fmt(x.pedidos)}</td>
    </tr>`;
  }).join('');
}

let currentSearch = "", currentCountry = "Todos", currentMultipais = false, currentVentas = "Todos";
// Multi-selección: Sets vacíos = "Todos". Permiten marcar varios niveles/programas a la vez.
let currentTiers = new Set();   // ej. {"Oro","Plata","Bronce"}
let currentProgs = new Set();
// Ordenamiento de la tabla de Clasificación.
//   sortCol: "nivel" (default = clasificación primero) | "total" | "pct_dev" | un mes "YYYY-MM"
//   sortDir: "desc" (mayor→menor) | "asc" (menor→mayor)
let sortCol = "nivel", sortDir = "desc";
function sortArrow(col) { return sortCol===col ? (sortDir==='desc'?' ▼':' ▲') : ''; }

const COUNTRY_FLAG = {
  "Colombia":"🇨🇴","Chile":"🇨🇱","Ecuador":"🇪🇨","México":"🇲🇽","Mexico":"🇲🇽",
  "Argentina":"🇦🇷","Perú":"🇵🇪","Peru":"🇵🇪","Guatemala":"🇬🇹","Panamá":"🇵🇦","Panama":"🇵🇦",
  "Paraguay":"🇵🇾","Costa Rica":"🇨🇷","CostaRica":"🇨🇷","España":"🇪🇸","Espana":"🇪🇸",
  "COLOMBIA":"🇨🇴","CHILE":"🇨🇱","ECUADOR":"🇪🇨","MEXICO":"🇲🇽","ARGENTINA":"🇦🇷",
  "PERU":"🇵🇪","GUATEMALA":"🇬🇹","PANAMA":"🇵🇦","PARAGUAY":"🇵🇾","COSTARICA":"🇨🇷",
};
const flag = p => COUNTRY_FLAG[p] || "🏳";
const semColor = {"verde":"bg-green-400","amarillo":"bg-yellow-400","naranja":"bg-orange-400","rojo":"bg-red-400","gris":"bg-slate-500"};
const semLabel = {"verde":"VERDE","amarillo":"AMARILLO","naranja":"NARANJA","rojo":"ROJO","gris":"SIN ACT."};
const semText  = {"verde":"text-green-400","amarillo":"text-yellow-400","naranja":"text-orange-400","rojo":"text-red-400","gris":"text-slate-500"};
const PROG_ORDER = ["Master Escala","Iniciación Escala","Ambos","Sin programa"];
const PROG_SHORT = {"Master Escala":"Master","Iniciación Escala":"Iniciación","Ambos":"Ambos","Sin programa":"Sin definir"};
const MES_ABBR = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"];
function mesShort(yyyymm) { const [y,m]=yyyymm.split('-'); return MES_ABBR[+m-1]+' '+y; }

function renderClasificacion(limit) {
  const users = baseUsers();

  // SCOPE = todos los filtros aplicados EXCEPTO el tier (para que los stats por tier sean reactivos)
  let scope = users;
  if (currentVentas === "Con ventas") scope = scope.filter(u => (u.total_pedidos||0) > 0);
  else if (currentVentas === "Sin ventas") scope = scope.filter(u => (u.total_pedidos||0) === 0);
  if (currentProgs.size) scope = scope.filter(u => currentProgs.has(u.programa));
  if (currentCountry !== "Todos") scope = scope.filter(u => paisMatch((u.paises_unicos||[]), currentCountry));
  if (currentMultipais) scope = scope.filter(u => (u.paises_unicos||[]).length > 1);
  if (currentSearch) {
    const s = currentSearch.toLowerCase();
    scope = scope.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
  }
  // counts para los botones del filtro de ventas (sobre la audiencia completa)
  const ventasCounts = {
    "Todos": users.length,
    "Con ventas": users.filter(u => (u.total_pedidos||0) > 0).length,
    "Sin ventas": users.filter(u => (u.total_pedidos||0) === 0).length,
  };

  const tierCounts = { "Todos": scope.length };
  TIER_ORDER.forEach(t => tierCounts[t] = scope.filter(u=>u.nivel===t).length);

  // counts por programa con todos los filtros excepto programa
  let scopeNoProg = users;
  if (currentCountry !== "Todos") scopeNoProg = scopeNoProg.filter(u => paisMatch((u.paises_unicos||[]), currentCountry));
  if (currentMultipais) scopeNoProg = scopeNoProg.filter(u => (u.paises_unicos||[]).length > 1);
  if (currentSearch) {
    const s = currentSearch.toLowerCase();
    scopeNoProg = scopeNoProg.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
  }
  const progCounts = { "Todos": scopeNoProg.length };
  PROG_ORDER.forEach(p => progCounts[p] = scopeNoProg.filter(u => u.programa===p).length);

  let filtered = scope;
  if (currentTiers.size) filtered = filtered.filter(u => currentTiers.has(u.nivel));
  // Si hay filtro por país: subsetear las métricas al país (ped_mes/total/top3/%dev).
  // El "nivel" no cambia (sigue siendo el del universo total del usuario).
  if (currentCountry !== "Todos") filtered = filtered.map(u => viewByPais(u, currentCountry));
  // Orden configurable por la columna que el usuario clickee.
  // dir=+1 para desc (mayor→menor), -1 para asc (menor→mayor).
  const dir = sortDir === 'asc' ? -1 : 1;
  if (sortCol === 'nivel') {
    // Escalafón primero (dir invierte Diamante↔Sin nivel); dentro de cada nivel, ventas desc.
    filtered.sort((a,b)=> dir*(TIER_ORDER.indexOf(a.nivel) - TIER_ORDER.indexOf(b.nivel))
                          || (b.total_pedidos - a.total_pedidos));
  } else if (sortCol === 'total') {
    filtered.sort((a,b)=> dir*((b.total_pedidos||0) - (a.total_pedidos||0)));
  } else if (sortCol === 'suma_top3') {
    filtered.sort((a,b)=> dir*((b.suma_top3||0) - (a.suma_top3||0)));
  } else if (sortCol === 'pct_dev') {
    filtered.sort((a,b)=> dir*((b.pct_dev||0) - (a.pct_dev||0)));
  } else {
    // un mes concreto (YYYY-MM): ordenar por pedidos de ese mes
    filtered.sort((a,b)=> dir*(((b.ped_mes&&b.ped_mes[sortCol])||0) - ((a.ped_mes&&a.ped_mes[sortCol])||0)));
  }

  const allCountries = paisesUnicos(users.flatMap(u => u.paises_unicos||u.paises||[]));
  const monthCols = DATA.meta.ventana.map(m => `<th data-sort="${m}" class="text-right text-[10px] uppercase tracking-wider cursor-pointer hover:text-cyan-300 select-none">${mesShort(m)}${sortArrow(m)}</th>`).join('');

  return `
    <!-- BARRA DE FILTROS -->
    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-4">
        <input id="search-input" type="text" placeholder="Buscar por nombre o email..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${currentSearch.replace(/"/g,'&quot;')}">
        <select id="country-select" class="bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm text-slate-200 focus:outline-none focus:border-cyan-500">
          <option value="Todos">Todos los países</option>
          ${allCountries.map(p => `<option value="${p}" ${currentCountry===p?'selected':''}>${flag(p)} ${p}</option>`).join('')}
        </select>
        <button id="clasif-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="clasif-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
        <div class="ml-auto"><span class="pill bg-violet-500/20 text-violet-200 border-violet-500/40">${fmt(scope.length)} usuarios</span></div>
      </div>

      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Nivel:</div>
        ${["Todos",...TIER_ORDER].map(t => {
          const on = t==='Todos' ? currentTiers.size===0 : currentTiers.has(t);
          return `<button data-tier="${t}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${on?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${t==='Sin clasificar'?'Sin nivel':t} <span class="ml-1 text-slate-500">${tierCounts[t]||0}</span></button>`;
        }).join('')}
      </div>

      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Programa:</div>
        ${["Todos",...PROG_ORDER].map(p => {
          const on = p==='Todos' ? currentProgs.size===0 : currentProgs.has(p);
          return `<button data-prog="${p}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${on?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${p==='Todos'?'Todos':PROG_SHORT[p]} <span class="ml-1 text-slate-500">${progCounts[p]||0}</span></button>`;
        }).join('')}
      </div>

      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Ventas:</div>
        ${["Todos","Con ventas","Sin ventas"].map(v =>
          `<button data-ventas="${v}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${currentVentas===v?'bg-emerald-600/30 text-emerald-200 border border-emerald-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${v} <span class="ml-1 text-slate-500">${ventasCounts[v]||0}</span></button>`
        ).join('')}
      </div>

      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Otros:</div>
        <button data-multipais="1" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${currentMultipais?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">🌎 Multi-País <span class="ml-1 text-slate-500">${users.filter(u=>(u.paises_unicos||[]).length>1).length}</span></button>
      </div>
    </div>

    <!-- 4 STAT CARDS por tier -->
    <div class="grid grid-cols-2 md:grid-cols-5 gap-3 mb-4">
      ${["Diamante","Platino","Oro","Plata","Bronce"].map(t => `
        <div class="card p-4">
          <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">${t}</div>
          <div class="text-3xl font-bold mt-1" style="color:${TIER_COLORS_HEX[t]}">${fmt(tierCounts[t]||0)}</div>
        </div>`).join('')}
    </div>

    <!-- TABLA -->
    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${filtered.length} de ${users.length} VIPs</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">Nombre</th>
              <th class="text-left">Teléfono</th>
              <th data-sort="nivel" class="text-center cursor-pointer hover:text-cyan-300 select-none">Nivel${sortArrow('nivel')}</th>
              <th class="text-center">Programa</th>
              <th class="text-left">Países</th>
              ${monthCols}
              <th data-sort="total" class="text-right cursor-pointer hover:text-cyan-300 select-none">Total${sortArrow('total')}</th>
              <th data-sort="suma_top3" class="text-right cursor-pointer hover:text-cyan-300 select-none" title="Suma de los 3 mejores meses — número que define el escalafón">Top-3${sortArrow('suma_top3')}</th>
              <th data-sort="pct_dev" class="text-right cursor-pointer hover:text-cyan-300 select-none">% Dev.${sortArrow('pct_dev')}</th>
              <th class="text-center">Semáforo</th>
            </tr>
          </thead>
          <tbody>
            ${filtered.slice(0,500).map(u => `
              <tr class="hover-row border-b border-white/5 cursor-pointer" data-cid="${u.cid}">
                <td class="py-2"><span class="text-cyan-300 font-medium hover:underline">${tc(u.nombre)||'—'}</span></td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||'Sin definir'}</span></td>
                <td class="text-[14px]">${u.sin_tienda?'<span class="pill bg-red-500/20 text-red-300 border-red-500/40 text-[10px]">⚠ Sin tienda</span>':((u.paises||[]).map(p => `<span title="${p}">${flag(p)}</span>`).join(' ')||'—')}${(u.paises_no_declarados||[]).length ? ` <span class="pill bg-amber-500/20 text-amber-300 border-amber-500/40 text-[9px]" title="Vende en ${(u.paises_no_declarados||[]).join(', ')} sin tener tienda declarada en GHL">⚠ ${(u.paises_no_declarados||[]).map(flag).join('')}</span>` : ''}</td>
                ${DATA.meta.ventana.map(m => `<td class="text-right font-mono text-slate-400">${fmt(u.ped_mes[m])}</td>`).join('')}
                <td class="text-right font-mono font-semibold text-slate-100">${fmt(u.total_pedidos)}</td>
                <td class="text-right font-mono font-semibold text-cyan-300" title="Suma de los 3 mejores meses (define el escalafón)">${fmt(u.suma_top3)}</td>
                <td class="text-right font-mono ${u.pct_dev>15?'text-orange-400':u.pct_dev>10?'text-yellow-400':'text-slate-400'}">${u.pct_dev}%</td>
                <td class="text-center"><span class="inline-flex items-center gap-1.5"><span class="w-2 h-2 rounded-full ${semColor[u.semaforo]||'bg-slate-600'}"></span><span class="text-[10px] ${semText[u.semaforo]||'text-slate-500'} font-semibold">${semLabel[u.semaforo]||'—'}</span></span></td>
              </tr>
            `).join('')}
            ${filtered.length>500?`<tr><td colspan="20" class="text-center text-slate-500 py-3">... y ${filtered.length-500} más</td></tr>`:''}
            ${filtered.length===0?`<tr><td colspan="20" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}

let top100_desde = null, top100_hasta = null, top100_prog = "Todos", top100_tend = "Todas", top100_pais = "Todos";

function tendencia(values) {
  // values: array de pedidos por mes (en orden cronológico) dentro del rango
  if (values.length < 2) return { sym:'—', col:'text-slate-500', key:'estable' };
  const first = values[0];
  const last = values[values.length-1];
  if (first === 0 && last === 0) return { sym:'—', col:'text-slate-500', key:'estable' };
  if (first === 0) return { sym:'▲', col:'text-green-400', key:'subiendo' };
  const delta = ((last - first) / first) * 100;
  if (delta > 10)  return { sym:'▲', col:'text-green-400', key:'subiendo' };
  if (delta < -10) return { sym:'▼', col:'text-red-400', key:'bajando' };
  return { sym:'—', col:'text-slate-500', key:'estable' };
}

function renderTop100() {
  const months = DATA.meta.ventana;
  if (!top100_desde) top100_desde = months[0];
  if (!top100_hasta) top100_hasta = months[months.length-1];
  if (months.indexOf(top100_desde) > months.indexOf(top100_hasta)) {
    [top100_desde, top100_hasta] = [top100_hasta, top100_desde];
  }
  const rangeMonths = months.slice(months.indexOf(top100_desde), months.indexOf(top100_hasta)+1);

  // Calcular sumas dentro del rango por usuario. Si hay filtro por país,
  // subseteamos las métricas del usuario al país (viewByPais).
  const list = DATA.usuarios.map(u => {
    const uu = top100_pais !== "Todos" ? viewByPais(u, top100_pais) : u;
    const ped = rangeMonths.reduce((s,m) => s + (uu.ped_mes[m]||0), 0);
    const ent = rangeMonths.reduce((s,m) => s + (uu.ent_mes[m]||0), 0);
    const dev = rangeMonths.reduce((s,m) => s + (uu.dev_mes[m]||0), 0);
    const pct = ped > 0 ? +(dev/ped*100).toFixed(1) : 0;
    const tend = tendencia(rangeMonths.map(m => uu.ped_mes[m]||0));
    return { ...u, range_ped: ped, range_ent: ent, range_dev: dev, range_pct: pct, range_tend: tend };
  });

  // Filtros
  let filtered = list.filter(u => u.range_ped > 0);
  if (top100_prog !== "Todos") filtered = filtered.filter(u => u.programa === top100_prog);
  if (top100_pais !== "Todos") filtered = filtered.filter(u => paisMatch((u.paises_unicos||u.paises||[]), top100_pais));
  if (top100_tend !== "Todas") filtered = filtered.filter(u => u.range_tend.key === top100_tend);
  filtered.sort((a,b) => b.range_ped - a.range_ped);
  const top = filtered.slice(0, 100);

  // Totales
  const tot_ped = top.reduce((s,u) => s + u.range_ped, 0);
  const tot_ent = top.reduce((s,u) => s + u.range_ent, 0);
  const tot_dev = top.reduce((s,u) => s + u.range_dev, 0);
  const tot_pct = tot_ped > 0 ? (tot_dev/tot_ped*100).toFixed(1) : 0;

  const monthOpts = months.map(m => `<option value="${m}">${mesShort(m)}</option>`).join('');

  return `
    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-4">
        <h3 class="text-base font-bold neon-cyan">Leaderboard — Top 100</h3>
        <label class="text-xs text-slate-500">Desde</label>
        <select id="t100-desde" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">${months.map(m => `<option value="${m}" ${top100_desde===m?'selected':''}>${mesShort(m)}</option>`).join('')}</select>
        <label class="text-xs text-slate-500">Hasta</label>
        <select id="t100-hasta" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">${months.map(m => `<option value="${m}" ${top100_hasta===m?'selected':''}>${mesShort(m)}</option>`).join('')}</select>
        <select id="t100-prog" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">
          <option value="Todos" ${top100_prog==='Todos'?'selected':''}>Todos los programas</option>
          ${PROG_ORDER.map(p => `<option value="${p}" ${top100_prog===p?'selected':''}>${PROG_SHORT[p]}</option>`).join('')}
        </select>
        ${paisSelectHTML('t100-pais', list, top100_pais, 'bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs')}
        <select id="t100-tend" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">
          <option value="Todas" ${top100_tend==='Todas'?'selected':''}>Todas las tendencias</option>
          <option value="subiendo" ${top100_tend==='subiendo'?'selected':''}>▲ Subiendo</option>
          <option value="bajando"  ${top100_tend==='bajando'?'selected':''}>▼ Bajando</option>
          <option value="estable"  ${top100_tend==='estable'?'selected':''}>— Estable</option>
        </select>
        <button id="t100-csv" class="text-[11px] px-3 py-1.5 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="t100-xlsx" class="text-[11px] px-3 py-1.5 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>

      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">#</th>
              <th class="text-left">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
              <th class="text-center">Programa</th>
              <th class="text-left">País</th>
              <th class="text-center">Nivel</th>
              <th class="text-center">Tend.</th>
              <th class="text-right">Pedidos VIP</th>
              <th class="text-right">Entregados</th>
              <th class="text-right">Devueltos</th>
              <th class="text-right">% Dev.</th>
              <th class="text-center">Sem.</th>
            </tr>
          </thead>
          <tbody>
            ${top.map((u,i) => `
              <tr class="hover-row border-b border-white/5 cursor-pointer" data-cid="${u.cid}">
                <td class="py-2 text-slate-500">${i+1}</td>
                <td><span class="text-cyan-300 font-medium hover:underline">${tc(u.nombre)||'—'}</span></td>
                <td class="text-slate-400">${u.email||'—'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||'Sin definir'}</span></td>
                <td class="text-slate-300">${(u.paises_unicos||[]).map(p => flag(p)+' '+p).join(', ')||'—'}</td>
                <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
                <td class="text-center font-mono ${u.range_tend.col}">${u.range_tend.sym}</td>
                <td class="text-right font-mono font-semibold text-slate-100">${fmt(u.range_ped)}</td>
                <td class="text-right font-mono text-slate-300">${fmt(u.range_ent)}</td>
                <td class="text-right font-mono text-slate-300">${fmt(u.range_dev)}</td>
                <td class="text-right font-mono ${u.range_pct>15?'text-orange-400':u.range_pct>10?'text-yellow-400':'text-green-400'}">${u.range_pct}%</td>
                <td class="text-center"><span class="w-2 h-2 rounded-full ${semColor[u.semaforo]} inline-block"></span></td>
              </tr>
            `).join('')}
            ${top.length===0?`<tr><td colspan="13" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
          <tfoot class="border-t-2 border-cyan-500/30 bg-cyan-500/5 sticky bottom-0">
            <tr class="font-semibold">
              <td colspan="2" class="py-2 text-cyan-300">TOTAL (${top.length})</td>
              <td colspan="6"></td>
              <td class="text-right font-mono text-slate-100">${fmt(tot_ped)}</td>
              <td class="text-right font-mono text-slate-200">${fmt(tot_ent)}</td>
              <td class="text-right font-mono text-slate-200">${fmt(tot_dev)}</td>
              <td class="text-right font-mono ${tot_pct>15?'text-orange-400':tot_pct>10?'text-yellow-400':'text-green-400'}">${tot_pct}%</td>
              <td></td>
            </tr>
          </tfoot>
        </table>
      </div>
    </div>
  `;
}

function wireTop100() {
  const sd = document.getElementById('t100-desde');
  const sh = document.getElementById('t100-hasta');
  const sp = document.getElementById('t100-prog');
  const st = document.getElementById('t100-tend');
  const spa = document.getElementById('t100-pais');
  if (sd) sd.onchange = e => { top100_desde = e.target.value; render(); };
  if (sh) sh.onchange = e => { top100_hasta = e.target.value; render(); };
  if (sp) sp.onchange = e => { top100_prog = e.target.value; render(); };
  if (st) st.onchange = e => { top100_tend = e.target.value; render(); };
  if (spa) spa.onchange = e => { top100_pais = e.target.value; render(); };
  document.querySelectorAll('[data-cid]').forEach(row => row.onclick = () => abrirFicha(row.dataset.cid));
  function _top100Rows() {
    const months = DATA.meta.ventana;
    let d = top100_desde || months[0], h = top100_hasta || months[months.length-1];
    if (months.indexOf(d) > months.indexOf(h)) { const t=d; d=h; h=t; }
    const rangeMonths = months.slice(months.indexOf(d), months.indexOf(h)+1);
    let list = DATA.usuarios.map(u => {
      const uu = top100_pais !== "Todos" ? viewByPais(u, top100_pais) : u;
      const ped = rangeMonths.reduce((s,m) => s + (uu.ped_mes[m]||0), 0);
      const ent = rangeMonths.reduce((s,m) => s + (uu.ent_mes[m]||0), 0);
      const dev = rangeMonths.reduce((s,m) => s + (uu.dev_mes[m]||0), 0);
      const pct = ped > 0 ? +(dev/ped*100).toFixed(1) : 0;
      const tend = tendencia(rangeMonths.map(m => uu.ped_mes[m]||0));
      return { ...u, range_ped: ped, range_ent: ent, range_dev: dev, range_pct: pct, range_tend: tend };
    }).filter(u => u.range_ped > 0);
    if (top100_prog !== "Todos") list = list.filter(u => u.programa === top100_prog);
    if (top100_pais !== "Todos") list = list.filter(u => paisMatch((u.paises_unicos||u.paises||[]), top100_pais));
    if (top100_tend !== "Todas") list = list.filter(u => u.range_tend.key === top100_tend);
    list.sort((a,b) => b.range_ped - a.range_ped);
    const top = list.slice(0, 100);
    const rows = [["#","Nombre","Email","Teléfono","Programa","Países","Nivel","Tendencia","Pedidos VIP","Entregados","Devueltos","% Dev.","Semáforo","Rango"]];
    top.forEach((u,i) => rows.push([
      i+1, u.nombre, u.email, u.telefono,
      PROG_SHORT[u.programa]||u.programa||'Sin programa',
      (u.paises_unicos||u.paises||[]).join('|'),
      u.nivel==='Sin clasificar'?'Sin nivel':u.nivel,
      u.range_tend.key, u.range_ped, u.range_ent, u.range_dev, u.range_pct,
      u.semaforo||'', `${d} → ${h}`
    ]));
    return rows;
  }
  const bc = document.getElementById('t100-csv');
  if (bc) bc.onclick = () => downloadCSV("top100_leaderboard.csv", _top100Rows());
  const bx = document.getElementById('t100-xlsx');
  if (bx) bx.onclick = () => _dlXLSX(bx, "top100_leaderboard.xlsx", _top100Rows(), "Top 100");
}

let alertaFiltroTipo = "Todas", alertaFiltroProg = "Todos", alertaSearch = "", alertaPais = "Todos";

function renderAlertas() {
  const users = baseUsers();
  const months = DATA.meta.ventana;

  // Counts globales
  const cnt = { "Eliminado":0, "Riesgo":0, "Desaparecido":0, "Crítica":0, "Huérfana":0, "Tienda no declarada":0 };
  users.forEach(u => {
    if (u.alerta_tipo) cnt[u.alerta_tipo] = (cnt[u.alerta_tipo]||0) + 1;
    // "Tienda no declarada" también cuenta a los que la tienen como flag
    // aunque tengan otra alerta primaria (ej: Eliminado + Tienda no declarada).
    if (u.tiene_no_declarada && u.alerta_tipo !== "Tienda no declarada") cnt["Tienda no declarada"]++;
  });
  const totalAlertas = users.filter(u => u.alerta_tipo || u.tiene_no_declarada).length;

  // Filtros
  let filtered = users.filter(u => u.alerta_tipo || u.tiene_no_declarada);
  if (alertaFiltroTipo === "Tienda no declarada") {
    filtered = filtered.filter(u => u.tiene_no_declarada);
  } else if (alertaFiltroTipo !== "Todas") {
    filtered = filtered.filter(u => u.alerta_tipo === alertaFiltroTipo);
  }
  if (alertaFiltroProg !== "Todos") filtered = filtered.filter(u => u.programa === alertaFiltroProg);
  if (alertaPais !== "Todos") filtered = filtered.filter(u => paisMatch((u.paises||u.paises_unicos||[]), alertaPais));
  if (alertaSearch) {
    const s = alertaSearch.toLowerCase();
    filtered = filtered.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
  }
  // Subsetear ped_mes al país filtrado (los números mostrados en la tabla)
  if (alertaPais !== "Todos") filtered = filtered.map(u => viewByPais(u, alertaPais));

  const alertColor = {
    "Crítica":     "bg-red-500/20 text-red-300 border-red-500/40",
    "Eliminado":   "bg-red-600/30 text-red-200 border-red-500/40",
    "Riesgo":      "bg-orange-500/20 text-orange-300 border-orange-500/40",
    "Desaparecido":"bg-yellow-500/20 text-yellow-300 border-yellow-500/40",
    "Huérfana":    "bg-violet-500/20 text-violet-300 border-violet-500/40",
    "Tienda no declarada": "bg-amber-500/20 text-amber-300 border-amber-500/40",
  };
  const semColor2 = {"verde":"bg-green-400","amarillo":"bg-yellow-400","naranja":"bg-orange-400","rojo":"bg-red-400","gris":"bg-slate-500"};
  const semLabel2 = {"verde":"VERDE","amarillo":"AMARILLO","naranja":"NARANJA","rojo":"ROJO","gris":"GRIS"};

  const monthCols = months.map(m => `<th class="text-right text-[10px] uppercase tracking-wider">${mesShort(m)}</th>`).join('');

  return `
    <!-- BARRA DE FILTROS -->
    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3">
        <select id="al-tipo" class="bg-black/40 border border-cyan-500/40 rounded-lg px-3 py-2 text-sm">
          <option value="Todas" ${alertaFiltroTipo==='Todas'?'selected':''}>Todas las alertas</option>
          <option value="Crítica" ${alertaFiltroTipo==='Crítica'?'selected':''}>🚨 Críticas</option>
          <option value="Desaparecido" ${alertaFiltroTipo==='Desaparecido'?'selected':''}>⚠️ Desaparecidos</option>
          <option value="Riesgo" ${alertaFiltroTipo==='Riesgo'?'selected':''}>🟠 Riesgo de eliminación</option>
          <option value="Eliminado" ${alertaFiltroTipo==='Eliminado'?'selected':''}>🔴 Eliminados</option>
          <option value="Huérfana" ${alertaFiltroTipo==='Huérfana'?'selected':''}>👻 Huérfanas (nunca vendieron)</option>
          <option value="Tienda no declarada" ${alertaFiltroTipo==='Tienda no declarada'?'selected':''}>⚠ Tienda no declarada en GHL</option>
        </select>
        <select id="al-prog" class="bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm">
          <option value="Todos" ${alertaFiltroProg==='Todos'?'selected':''}>Todos los programas</option>
          ${PROG_ORDER.map(p => `<option value="${p}" ${alertaFiltroProg===p?'selected':''}>${PROG_SHORT[p]}</option>`).join('')}
        </select>
        ${paisSelectHTML('al-pais', users, alertaPais)}
        <input id="al-search" type="text" placeholder="Buscar..." class="flex-1 min-w-[200px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500" value="${alertaSearch.replace(/"/g,'&quot;')}">
        <button id="al-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="al-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
    </div>

    <!-- STAT CARDS + TOTAL -->
    <div class="grid grid-cols-2 md:grid-cols-3 lg:grid-cols-6 gap-3 mb-4">
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">🚨 Críticas</div>
        <div class="text-3xl font-bold mt-1 text-red-400">${fmt(cnt['Crítica'])}</div>
        <div class="text-[10px] text-slate-500 mt-1">% devolución > 50%</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">⚠️ Desaparecidos</div>
        <div class="text-3xl font-bold mt-1 text-yellow-400">${fmt(cnt['Desaparecido'])}</div>
        <div class="text-[10px] text-slate-500 mt-1">0 pedidos en ${mesShort(months[months.length-1])}</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">🟠 Riesgo</div>
        <div class="text-3xl font-bold mt-1 text-orange-400">${fmt(cnt['Riesgo'])}</div>
        <div class="text-[10px] text-slate-500 mt-1">Últimos 2 meses en 0</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">🔴 Eliminados</div>
        <div class="text-3xl font-bold mt-1 text-red-500">${fmt(cnt['Eliminado'])}</div>
        <div class="text-[10px] text-slate-500 mt-1">3+ meses en 0</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">👻 Huérfanas</div>
        <div class="text-3xl font-bold mt-1 text-violet-400">${fmt(cnt['Huérfana'])}</div>
        <div class="text-[10px] text-slate-500 mt-1">Nunca han tenido ventas</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">⚠ Tienda no declarada</div>
        <div class="text-3xl font-bold mt-1 text-amber-400">${fmt(users.filter(u=>u.tiene_no_declarada).length)}</div>
        <div class="text-[10px] text-slate-500 mt-1">Venden en país sin registrar en GHL</div>
      </div>
      <div class="card p-4">
        <div class="text-[10px] font-semibold uppercase tracking-wider text-slate-500">Total alertas</div>
        <div class="text-3xl font-bold mt-1 neon-cyan">${fmt(totalAlertas)}</div>
        <div class="text-[10px] text-slate-500 mt-1">usuarios con alerta</div>
      </div>
    </div>

    <!-- TABLA -->
    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${filtered.length} de ${totalAlertas} con alerta</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
              <th class="text-center">Programa</th>
              <th class="text-left">País</th>
              <th class="text-center">Nivel</th>
              ${monthCols}
              <th class="text-center">Semáforo</th>
              <th class="text-center">Alerta</th>
            </tr>
          </thead>
          <tbody>
            ${filtered.map(u => `
              <tr class="hover-row border-b border-white/5 cursor-pointer" data-cid="${u.cid}">
                <td class="py-2"><span class="text-cyan-300 font-medium hover:underline">${tc(u.nombre)||'—'}</span></td>
                <td class="text-slate-400">${u.email||'—'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||'Sin definir'}</span></td>
                <td class="text-[14px]">${(u.paises||[]).map(p => flag(p)).join(' ')||'—'}</td>
                <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
                ${months.map(m => `<td class="text-right font-mono ${(u.ped_mes[m]||0)===0?'text-slate-700':'text-slate-400'}">${fmt(u.ped_mes[m])}</td>`).join('')}
                <td class="text-center"><span class="inline-flex items-center gap-1.5"><span class="w-2 h-2 rounded-full ${semColor2[u.semaforo]||'bg-slate-500'}"></span><span class="text-[10px] font-semibold">${semLabel2[u.semaforo]||'—'}</span></span></td>
                <td class="text-center"><span class="pill ${alertColor[u.alerta_tipo]||'bg-white/5'}">${u.alerta_tipo||'—'}</span></td>
              </tr>
            `).join('')}
            ${filtered.length===0?`<tr><td colspan="20" class="text-center text-slate-500 py-6">— sin alertas en este filtro —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}

function wireAlertas() {
  const t = document.getElementById('al-tipo');
  const p = document.getElementById('al-prog');
  const pa = document.getElementById('al-pais');
  const s = document.getElementById('al-search');
  if (t) t.onchange = e => { alertaFiltroTipo = e.target.value; render(); };
  if (p) p.onchange = e => { alertaFiltroProg = e.target.value; render(); };
  if (pa) pa.onchange = e => { alertaPais = e.target.value; render(); };
  if (s) {
    s.oninput = e => { alertaSearch = e.target.value; render(); };
    s.focus(); s.setSelectionRange(alertaSearch.length, alertaSearch.length);
  }
  document.querySelectorAll('[data-cid]').forEach(row => row.onclick = () => abrirFicha(row.dataset.cid));
  function _alRows() {
    const users = baseUsers();
    const months = DATA.meta.ventana;
    let filtered = users.filter(u => u.alerta_tipo);
    if (alertaFiltroTipo !== "Todas") filtered = filtered.filter(u => u.alerta_tipo === alertaFiltroTipo);
    if (alertaFiltroProg !== "Todos") filtered = filtered.filter(u => u.programa === alertaFiltroProg);
    if (alertaPais !== "Todos") filtered = filtered.filter(u => paisMatch((u.paises||u.paises_unicos||[]), alertaPais));
    if (alertaSearch) {
      const ss = alertaSearch.toLowerCase();
      filtered = filtered.filter(u => (u.nombre||'').toLowerCase().includes(ss) || (u.email||'').toLowerCase().includes(ss));
    }
    if (alertaPais !== "Todos") filtered = filtered.map(u => viewByPais(u, alertaPais));
    const header = ["Nombre","Email","Teléfono","Programa","Países","Nivel", ...months, "Semáforo","Alerta"];
    const rows = [header];
    filtered.forEach(u => rows.push([
      u.nombre, u.email, u.telefono,
      PROG_SHORT[u.programa]||u.programa||'Sin programa',
      (u.paises||u.paises_unicos||[]).join('|'),
      u.nivel==='Sin clasificar'?'Sin nivel':u.nivel,
      ...months.map(m => u.ped_mes[m]||0),
      u.semaforo||'', u.alerta_tipo||''
    ]));
    return rows;
  }
  const bc = document.getElementById('al-csv');
  if (bc) bc.onclick = () => downloadCSV("alertas_vip.csv", _alRows());
  const bx = document.getElementById('al-xlsx');
  if (bx) bx.onclick = () => _dlXLSX(bx, "alertas_vip.xlsx", _alRows(), "Alertas");
}

function renderPaises() {
  const pp = DATA.por_pais;
  const total = Object.values(pp).reduce((s,n)=>s+n,0);
  return `<div class="card p-5">
    <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Contactos con tienda · por país</h3>
    <table class="w-full text-sm">
      <thead class="text-[10px] text-slate-500 uppercase tracking-wider">
        <tr><th class="text-left py-2">País</th><th class="text-right">Contactos</th><th class="text-right">%</th><th class="text-left pl-6">Reparto</th></tr>
      </thead>
      <tbody>
        ${Object.entries(pp).map(([p,n]) => {
          const pct = total ? (n*100/total).toFixed(1) : 0;
          const max = Math.max(...Object.values(pp)) || 1;
          const bw = (n/max*100).toFixed(1);
          return `<tr class="border-b border-white/5">
            <td class="py-2 text-slate-200">${p}</td>
            <td class="text-right font-mono text-slate-300">${n}</td>
            <td class="text-right text-slate-500">${pct}%</td>
            <td class="pl-6 pr-2 w-2/5"><div class="bg-white/5 h-2 rounded-full overflow-hidden"><div class="h-full rounded-full bg-cyan-500" style="width:${bw}%"></div></div></td>
          </tr>`;
        }).join('')}
      </tbody></table>
  </div>`;
}

function renderReglas() {
  return `<div class="space-y-4">
    <div class="card p-5">
      <h3 class="text-sm font-semibold mb-3 neon-cyan">Reglas de Ingreso y Permanencia</h3>
      <div class="grid grid-cols-1 md:grid-cols-3 gap-4 text-sm">
        <div><h4 class="font-semibold text-slate-200 mb-2">① Ingreso</h4>
          <ul class="text-slate-400 space-y-1 text-xs">
            <li>• Mínimo 2 meses con ventas dentro de la ventana de 5 meses</li>
            <li>• Suma de tus 2 mejores meses (top-2) ≥ <span class="neon-yellow font-semibold">60 pedidos</span></li>
            <li>• Si cumples → ingresas como <span class="pill ${tierColor.Bronce}">Bronce</span></li>
          </ul></div>
        <div><h4 class="font-semibold text-slate-200 mb-2">② Escalafón</h4>
          <ul class="text-slate-400 space-y-1 text-xs">
            <li>• Para subir de Bronce: necesitas <span class="neon-yellow font-semibold">3 meses con ventas</span></li>
            <li>• Tu nivel se calcula con suma top-3</li>
            <li>• Mantienes beneficios de niveles anteriores</li>
          </ul></div>
        <div><h4 class="font-semibold text-slate-200 mb-2">③ Eliminación</h4>
          <ul class="text-slate-400 space-y-1 text-xs">
            <li>• 3 meses CONSECUTIVOS sin pedidos → eliminado</li>
            <li>• Si vendes en cualquiera de los 3, te mantienes</li>
            <li>• Aplica a todos los niveles</li>
          </ul></div>
      </div>
    </div>
    <div class="card p-5">
      <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-4">Reglas de escalafones</h3>
      <table class="w-full text-sm">
        <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10">
          <tr>
            <th class="text-left py-3">Nivel</th>
            <th class="text-left">Pedidos/mes mínimo</th>
            <th class="text-left">Meses requeridos</th>
            <th class="text-left">Sumatoria mínima</th>
          </tr>
        </thead>
        <tbody>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor.Diamante}">Diamante</span></td>
            <td class="font-mono text-slate-200">5.000+</td>
            <td class="text-slate-300">3 activos</td>
            <td class="font-mono text-slate-100">15.000</td>
          </tr>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor.Platino}">Platino</span></td>
            <td class="font-mono text-slate-200">1.000+</td>
            <td class="text-slate-300">3 activos</td>
            <td class="font-mono text-slate-100">3.000</td>
          </tr>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor.Oro}">Oro</span></td>
            <td class="font-mono text-slate-200">300+</td>
            <td class="text-slate-300">3 activos</td>
            <td class="font-mono text-slate-100">900</td>
          </tr>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor.Plata}">Plata</span></td>
            <td class="font-mono text-slate-200">100+</td>
            <td class="text-slate-300">3 activos</td>
            <td class="font-mono text-slate-100">300</td>
          </tr>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor.Bronce}">Bronce</span></td>
            <td class="font-mono text-slate-200">30+</td>
            <td class="text-slate-300">2 activos</td>
            <td class="font-mono text-slate-100">60</td>
          </tr>
          <tr class="border-b border-white/5">
            <td class="py-3"><span class="pill ${tierColor['Sin clasificar']}">Sin nivel</span></td>
            <td class="text-slate-400">&lt;30 o 1 mes</td>
            <td class="text-slate-500">—</td>
            <td class="text-slate-500">—</td>
          </tr>
        </tbody></table>
      <div class="text-[11px] text-slate-500 mt-4 leading-relaxed">
        <strong class="text-slate-400">Cómo se lee:</strong> Diamante se obtiene con 3 meses activos donde el promedio sea 5.000+ pedidos por mes
        (sumatoria de los 3 mejores meses ≥ 15.000). Bronce es el único nivel que requiere solo 2 meses activos (top-2 ≥ 60).
        Los demás requieren al menos 3 meses con ventas dentro de la ventana de 5 meses.
      </div>
    </div>
  </div>`;
}

function renderConsulta() {
  return `<div class="card p-5">
    <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Consulta individual</h3>
    <input id="consulta-input" type="text" placeholder="Pega un correo, contact_id o nombre..."
           class="w-full bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm mb-4 focus:outline-none focus:border-cyan-500">
    <div id="consulta-result" class="text-sm text-slate-500">Buscando...</div>
  </div>`;
}

// ═══════════════════════════════════════════════════════════════════════
// APP MASTER ESCALA · dashboard de pagos (suscripciones + top-ups)
// Data source: DATA.pagos, alimentado por extraer_pagos.py cada 10 min.
// ═══════════════════════════════════════════════════════════════════════
let pagoTipoFilter = "Todos";        // Todos / Top-up / Nueva suscripción
let pagoProgramaFilter = "Todos";    // Todos / Master Escala / Iniciación Escala / Ambos / Sin programa
let pagoEstFilter = "Todos";         // Todos / Sí / No
let pagoSearch = "";
let pagoSortCol = "fecha";           // fecha | neto_usd | bruto_usd | creditos | email | descripcion
let pagoSortDir = "desc";
let pagoMesDesde = null;
let pagoMesHasta = null;

function pagoMes(iso){ return (iso || "").slice(0,7); }
function pagoFmtDate(iso){ return (iso || "").slice(0,10); }
function pagoUSD(n){ return "$" + (n||0).toLocaleString("es-CO", {minimumFractionDigits:2, maximumFractionDigits:2}); }
// Etiqueta normalizada del tipo, SIEMPRE desde tipo_raw (fiable): subscription
// → Nueva suscripción, topup → Top-up, renewal → Renovación. Así el filtro y el
// gráfico tratan las renovaciones como un tipo propio (en pagos.json el `tipo`
// de las renovaciones quedó como "renewal" crudo).
function pagoTipoLabel(p){
  const tr = ((p && p.tipo_raw) || "").toLowerCase();
  if (tr === "topup") return "Top-up";
  if (tr === "subscription") return "Nueva suscripción";
  if (tr === "renewal") return "Renovación";
  return (p && p.tipo) || "";
}

function pagoFilterList() {
  let list = (DATA.pagos || []).slice();
  if (pagoTipoFilter !== "Todos") list = list.filter(p => pagoTipoLabel(p) === pagoTipoFilter);
  if (pagoProgramaFilter !== "Todos") list = list.filter(p => (p.programa || "Sin programa") === pagoProgramaFilter);
  if (pagoEstFilter === "Sí")  list = list.filter(p => p.estudiante === true);
  if (pagoEstFilter === "No")  list = list.filter(p => p.estudiante === false);
  if (pagoMesDesde) list = list.filter(p => pagoMes(p.fecha) >= pagoMesDesde);
  if (pagoMesHasta) list = list.filter(p => pagoMes(p.fecha) <= pagoMesHasta);
  if (pagoSearch) {
    const s = pagoSearch.toLowerCase();
    list = list.filter(p => (p.email||"").toLowerCase().includes(s)
                          || (p.nombre||"").toLowerCase().includes(s)
                          || (p.descripcion||"").toLowerCase().includes(s)
                          || (p.stripe_charge_id||"").toLowerCase().includes(s));
  }
  const dir = pagoSortDir === "asc" ? 1 : -1;
  list.sort((a,b) => {
    const va = a[pagoSortCol], vb = b[pagoSortCol];
    if (typeof va === "number" && typeof vb === "number") return dir * (va - vb);
    return dir * String(va||"").localeCompare(String(vb||""));
  });
  return list;
}

function pagoSortArrow(col){
  if (pagoSortCol !== col) return " <span class='text-slate-600'>⇅</span>";
  return pagoSortDir === "asc" ? " ▲" : " ▼";
}

function renderPagosDashboard() {
  const all = DATA.pagos || [];
  if (all.length === 0) {
    return `<div class="card p-8 text-center">
      <h2 class="text-lg neon-cyan mb-3">💰 App Master Escala</h2>
      <p class="text-slate-400">Aún no hay pagos registrados en <code>pagos.json</code>.<br>
      El backfill inicial puede estar corriendo. Se actualiza cada 10 min con la rutina
      <code>extraer-pagos</code>.</p></div>`;
  }
  const list = pagoFilterList();
  const meses = [...new Set(all.map(p => pagoMes(p.fecha)))].filter(Boolean).sort();
  const programas = ["Todos","Master Escala","Iniciación Escala","Ambos","Sin programa"];
  const tipos = ["Todos","Nueva suscripción","Renovación","Top-up"];
  const estOpts = ["Todos","Sí","No"];
  // Totales sobre la lista filtrada
  const tot_pagos = list.length;
  const tot_creditos = list.reduce((s,p) => s + (p.creditos||0), 0);
  const tot_bruto = list.reduce((s,p) => s + (p.bruto_usd||0), 0);
  const tot_comision = list.reduce((s,p) => s + (p.comision_usd||0), 0);
  const tot_neto = list.reduce((s,p) => s + (p.neto_usd||0), 0);
  // Distribución por tipo (para % de mix)
  const nSub = list.filter(p => pagoTipoLabel(p) === "Nueva suscripción").length;
  const nRen = list.filter(p => pagoTipoLabel(p) === "Renovación").length;
  const nTop = list.filter(p => pagoTipoLabel(p) === "Top-up").length;
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">💰 App Master Escala — Pagos</h2>
      <div class="text-xs text-slate-500">Transacciones registradas como notas GHL <code>Pago recibido App Master Escala</code>. Datos parseados por <code>extraer_pagos.py</code> · actualizado cada 10 min.</div>
    </div>

    <!-- STAT CARDS · reactivos al filtro -->
    <div class="grid grid-cols-2 md:grid-cols-5 gap-3 mb-4">
      ${statCard("Total transacciones", tot_pagos, `${nSub} nuevas · ${nRen} renovaciones · ${nTop} top-ups`, "neon-cyan")}
      ${statCard("Créditos vendidos", tot_creditos, "En transacciones filtradas", "neon-violet")}
      ${statCard("Bruto USD", pagoUSD(tot_bruto), "Antes de comisión Stripe", "neon-green")}
      ${statCard("Comisión Stripe", pagoUSD(tot_comision), `${tot_bruto>0?(tot_comision/tot_bruto*100).toFixed(1):0}% del bruto`, "neon-yellow")}
      ${statCard("Neto USD", pagoUSD(tot_neto), "Después de Stripe", "neon-pink")}
    </div>

    <!-- GRÁFICO MENSUAL -->
    <div class="card p-4 mb-4">
      <div class="flex justify-between items-baseline mb-2">
        <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500">Ingresos por mes (neto USD)</h3>
        <div class="text-[10px] text-slate-500">Barras apiladas · verde=suscripciones · ámbar=renovaciones · violeta=top-ups</div>
      </div>
      <div style="height:260px"><canvas id="pagos-chart"></canvas></div>
    </div>

    <!-- FILTROS -->
    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-3">
        <input id="pago-search" type="text" placeholder="Buscar por email, nombre, plan o Stripe charge ID..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${pagoSearch.replace(/"/g,'&quot;')}">
        <button id="pago-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="pago-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Tipo:</div>
        ${tipos.map(t => `<button data-pagotipo="${t}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${pagoTipoFilter===t?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${t}</button>`).join('')}
      </div>
      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Programa:</div>
        ${programas.map(p => `<button data-pagoprog="${p}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${pagoProgramaFilter===p?'bg-violet-600/30 text-violet-200 border border-violet-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${p==='Todos'?'Todos':(PROG_SHORT[p]||p)}</button>`).join('')}
      </div>
      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Estudiante:</div>
        ${estOpts.map(e => `<button data-pagoest="${e}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${pagoEstFilter===e?'bg-emerald-600/30 text-emerald-200 border border-emerald-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${e}</button>`).join('')}
      </div>
      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Mes:</div>
        <select id="pago-desde" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">
          <option value="">Desde (todos)</option>
          ${meses.map(m => `<option value="${m}" ${pagoMesDesde===m?'selected':''}>${m}</option>`).join('')}
        </select>
        <select id="pago-hasta" class="bg-black/40 border border-white/10 rounded-lg px-2 py-1.5 text-xs">
          <option value="">Hasta (todos)</option>
          ${meses.map(m => `<option value="${m}" ${pagoMesHasta===m?'selected':''}>${m}</option>`).join('')}
        </select>
      </div>
    </div>

    <!-- TABLA -->
    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${list.length} de ${all.length} transacciones</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th data-pagosort="fecha" class="text-left py-2 cursor-pointer hover:text-cyan-300 select-none">Fecha${pagoSortArrow('fecha')}</th>
              <th data-pagosort="email" class="text-left cursor-pointer hover:text-cyan-300 select-none">Email${pagoSortArrow('email')}</th>
              <th class="text-left">Nombre</th>
              <th class="text-center">Estudiante</th>
              <th class="text-center">Programa</th>
              <th class="text-center">Tipo</th>
              <th data-pagosort="descripcion" class="text-left cursor-pointer hover:text-cyan-300 select-none">Plan${pagoSortArrow('descripcion')}</th>
              <th data-pagosort="creditos" class="text-right cursor-pointer hover:text-cyan-300 select-none">Créditos${pagoSortArrow('creditos')}</th>
              <th data-pagosort="bruto_usd" class="text-right cursor-pointer hover:text-cyan-300 select-none">Bruto USD${pagoSortArrow('bruto_usd')}</th>
              <th class="text-right">Comisión USD</th>
              <th data-pagosort="neto_usd" class="text-right cursor-pointer hover:text-cyan-300 select-none">Neto USD${pagoSortArrow('neto_usd')}</th>
              <th class="text-left">Stripe charge ID</th>
            </tr>
          </thead>
          <tbody>
            ${list.slice(0, 2000).map(p => `
              <tr class="hover-row border-b border-white/5">
                <td class="py-2 text-slate-300 font-mono">${pagoFmtDate(p.fecha)}</td>
                <td class="text-slate-300">${p.email||'—'}</td>
                <td class="text-slate-200">${tc(p.nombre)||'—'}</td>
                <td class="text-center">${p.estudiante ? '<span class="pill bg-emerald-500/20 text-emerald-300 border-emerald-500/40">Sí</span>' : '<span class="text-slate-600">No</span>'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[p.programa]||p.programa||'—'}</span></td>
                <td class="text-center"><span class="pill ${pagoTipoLabel(p)==='Top-up'?'bg-violet-500/20 text-violet-300 border-violet-500/40':pagoTipoLabel(p)==='Renovación'?'bg-amber-500/20 text-amber-300 border-amber-500/40':'bg-cyan-500/20 text-cyan-300 border-cyan-500/40'}">${pagoTipoLabel(p)}</span></td>
                <td class="text-slate-300">${p.descripcion||'—'}</td>
                <td class="text-right font-mono text-slate-200">${fmt(p.creditos)}</td>
                <td class="text-right font-mono text-slate-200">${pagoUSD(p.bruto_usd)}</td>
                <td class="text-right font-mono text-orange-300">${pagoUSD(p.comision_usd)}</td>
                <td class="text-right font-mono font-semibold text-emerald-300">${pagoUSD(p.neto_usd)}</td>
                <td class="text-left font-mono text-slate-400">${p.stripe_charge_id ? tc(p.stripe_charge_id) : '—'}</td>
              </tr>`).join('')}
            ${list.length > 2000 ? `<tr><td colspan="12" class="text-center text-slate-500 py-3">... y ${list.length-2000} más (usa CSV para ver todos)</td></tr>` : ''}
            ${list.length === 0 ? `<tr><td colspan="12" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>` : ''}
          </tbody>
          <tfoot class="border-t-2 border-cyan-500/30 bg-cyan-500/5 sticky bottom-0">
            <tr class="font-semibold">
              <td colspan="7" class="py-2 text-cyan-300">TOTAL (${tot_pagos})</td>
              <td class="text-right font-mono text-slate-100">${fmt(tot_creditos)}</td>
              <td class="text-right font-mono text-slate-100">${pagoUSD(tot_bruto)}</td>
              <td class="text-right font-mono text-orange-200">${pagoUSD(tot_comision)}</td>
              <td class="text-right font-mono text-emerald-200">${pagoUSD(tot_neto)}</td>
              <td></td>
            </tr>
          </tfoot>
        </table>
      </div>
    </div>
  `;
}

function wirePagosDashboard() {
  document.querySelectorAll('[data-pagotipo]').forEach(b => b.onclick = () => { pagoTipoFilter = b.dataset.pagotipo; render(); });
  document.querySelectorAll('[data-pagoprog]').forEach(b => b.onclick = () => { pagoProgramaFilter = b.dataset.pagoprog; render(); });
  document.querySelectorAll('[data-pagoest]').forEach(b => b.onclick = () => { pagoEstFilter = b.dataset.pagoest; render(); });
  document.querySelectorAll('th[data-pagosort]').forEach(th => th.onclick = () => {
    const col = th.dataset.pagosort;
    if (pagoSortCol === col) pagoSortDir = pagoSortDir === 'asc' ? 'desc' : 'asc';
    else { pagoSortCol = col; pagoSortDir = 'desc'; }
    render();
  });
  const inp = document.getElementById('pago-search');
  if (inp) {
    inp.oninput = e => { pagoSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(pagoSearch.length, pagoSearch.length);
  }
  const sd = document.getElementById('pago-desde');
  const sh = document.getElementById('pago-hasta');
  if (sd) sd.onchange = e => { pagoMesDesde = e.target.value || null; render(); };
  if (sh) sh.onchange = e => { pagoMesHasta = e.target.value || null; render(); };
  // Helper: arma las filas para export (mismo formato para CSV y XLSX)
  function _pagoExportRows() {
    const list = pagoFilterList();
    // Fecha en hora Colombia (formato "YYYY-MM-DD HH:MM:SS").
    const fdate = (iso) => { if(!iso) return ""; const d=new Date(iso); if(isNaN(d)) return String(iso);
      return new Intl.DateTimeFormat('sv-SE',{timeZone:'America/Bogota',year:'numeric',month:'2-digit',day:'2-digit',hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:false}).format(d).replace('T',' '); };
    // Tipo con las etiquetas del reporte de compras.
    const tlabel = (p) => { const tr=(p.tipo_raw||'').toLowerCase();
      return tr==='topup'?'Recarga de créditos':tr==='subscription'?'Plan (suscripción)':tr==='renewal'?'Renovación de plan':(p.tipo||''); };
    // Descripción tipo "Top-up/Suscripción/Renovación Stripe {código} — {n} créditos".
    const descr = (p) => { const d=(p.descripcion||'').trim(), dl=d.toLowerCase();
      if(dl.startsWith('top-up stripe')||dl.startsWith('suscripción stripe')||dl.startsWith('renovación stripe')) return d;
      const tr=(p.tipo_raw||'').toLowerCase(), cred=p.creditos||0;
      const pref={topup:'Top-up Stripe',subscription:'Suscripción Stripe',renewal:'Renovación Stripe'}[tr];
      if(!pref) return d; const code=d.toLowerCase().replace(/ /g,'_').replace(/\+/g,'_');
      return code?(pref+' '+code+' — '+cred+' créditos'):(pref+' — '+cred+' créditos'); };
    const rows = [["Fecha","Correo","Nombre","Teléfono","Tipo","Descripción","Créditos","Pagado (USD)","Comisión Stripe (USD)","Neto recibido (USD)","Stripe charge ID"]];
    list.forEach(p => rows.push([
      fdate(p.fecha), p.email||'', p.nombre||'', p.telefono||'',
      tlabel(p), descr(p), p.creditos||0, p.bruto_usd||0, p.comision_usd||0, p.neto_usd||0, p.stripe_charge_id||''
    ]));
    return rows;
  }
  // Nombre de archivo con la fecha del día (hora Colombia): "... - YYYY-MM-DD.ext".
  const _pagoFname = (ext) => "Master Escala - Compras (planes y recargas) - " +
    new Intl.DateTimeFormat('sv-SE',{timeZone:'America/Bogota',year:'numeric',month:'2-digit',day:'2-digit'}).format(new Date()) + "." + ext;
  const btn = document.getElementById('pago-csv');
  if (btn) btn.onclick = () => downloadCSV(_pagoFname('csv'), _pagoExportRows());
  const btnX = document.getElementById('pago-xlsx');
  if (btnX) btnX.onclick = () => {
    btnX.disabled = true; const orig = btnX.textContent; btnX.textContent = 'Generando…';
    Promise.resolve()
      .then(() => downloadXLSX(_pagoFname('xlsx'), _pagoExportRows(), "Compras"))
      .finally(() => { setTimeout(() => { btnX.disabled = false; btnX.textContent = orig; }, 1500); });
  };
}

let _pagosChart = null;
function drawPagosChart() {
  const canvas = document.getElementById('pagos-chart');
  if (!canvas) return;
  if (_pagosChart) { _pagosChart.destroy(); _pagosChart = null; }
  const list = pagoFilterList();
  // Agrupar por mes + tipo
  const byMonth = {};
  list.forEach(p => {
    const m = pagoMes(p.fecha);
    if (!m) return;
    if (!byMonth[m]) byMonth[m] = { "Nueva suscripción":0, "Renovación":0, "Top-up":0 };
    const t = pagoTipoLabel(p);
    byMonth[m][t] = (byMonth[m][t]||0) + (p.neto_usd||0);
  });
  const labels = Object.keys(byMonth).sort();
  const dataSub = labels.map(m => +(byMonth[m]["Nueva suscripción"] || 0).toFixed(2));
  const dataRen = labels.map(m => +(byMonth[m]["Renovación"] || 0).toFixed(2));
  const dataTop = labels.map(m => +(byMonth[m]["Top-up"] || 0).toFixed(2));
  _pagosChart = new Chart(canvas.getContext('2d'), {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [
        { label: 'Nueva suscripción', data: dataSub, backgroundColor: 'rgba(34,197,94,.75)', borderRadius: 4 },
        { label: 'Renovación',        data: dataRen, backgroundColor: 'rgba(251,191,36,.75)', borderRadius: 4 },
        { label: 'Top-up',            data: dataTop, backgroundColor: 'rgba(167,139,250,.75)', borderRadius: 4 },
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { labels: { color: '#cbd5e1' } },
        tooltip: { callbacks: { label: ctx => ctx.dataset.label + ': $' + ctx.parsed.y.toFixed(2) } },
      },
      scales: {
        x: { stacked: true, ticks: { color: '#94a3b8' }, grid: { color: 'rgba(255,255,255,.05)' } },
        y: { stacked: true, ticks: { color: '#94a3b8', callback: v => '$' + v }, grid: { color: 'rgba(255,255,255,.05)' } },
      },
    }
  });
}

function render() {
  const main = document.getElementById("main-content");
  switch (currentTab) {
    // Audiencias (VIP / Estudiantes / Pendientes / Todos VIP) — genéricos
    case "resumen": main.innerHTML = renderResumen(); drawCharts(); break;
    case "clasif":  main.innerHTML = renderClasificacion(0); wireClasificacion(); break;
    case "alertas": main.innerHTML = renderAlertas(); wireAlertas(); break;
    // Estudiantes extra
    case "met_programas":  main.innerHTML = renderMetProgramas();   wireMetProgramas();   break;
    // Otros
    case "otros_resumen":  main.innerHTML = renderOtrosResumen();   break;
    case "met_dropi_ghl":  main.innerHTML = renderMetDropiGHL();    wireMetDropiGHL();    break;
    case "met_duplicados": main.innerHTML = renderMetDuplicados();  wireMetDuplicados();  break;
    // Configuración
    case "reglas":  main.innerHTML = renderReglas(); break;
    case "consulta":main.innerHTML = renderConsulta(); wireConsulta(); break;
    // App Master Escala
    case "pagos_dashboard": main.innerHTML = renderPagosDashboard(); wirePagosDashboard(); drawPagosChart(); break;
  }
}

function renderOtrosResumen() {
  const all = DATA.metricas.dropi_sin_ghl || [];
  const m = DATA.metricas;
  const dist = {}; TIER_ORDER.forEach(t => dist[t] = 0);
  let totalPed = 0;
  all.forEach(u => { if (dist[u.nivel] !== undefined) dist[u.nivel]++; totalPed += (u.total_pedidos||0); });
  const maxN = Math.max(...Object.values(dist), 1);
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">📦 Otros · En Dropi pero NO en GHL</h2>
      <div class="text-xs text-slate-500">Correos que Dropi reporta pero que no se crearon en GHL (ni por email ni por teléfono). Tienen ventas pero no contacto propio.</div>
    </div>
    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      ${statCard("Total Otros", all.length, `de ${fmt(m.dropi_emails_total)} correos en Dropi`, "neon-cyan")}
      ${statCard("Con ventas", m.dropi_sin_ghl_con_ventas, "Tienen ≥1 pedido en la ventana", "neon-green")}
      ${statCard("Sin ventas", m.dropi_sin_ghl_sin_ventas, "0 pedidos en la ventana", "neon-red")}
      ${statCard("Total pedidos", totalPed, "Acumulado de este grupo", "neon-yellow")}
    </div>
    <div class="card p-5">
      <h3 class="text-xs font-semibold uppercase tracking-wider text-slate-500 mb-3">Escalafón calculado (desde ventas Dropi)</h3>
      <table class="w-full text-sm"><tbody>
        ${TIER_ORDER.map(t => {
          const n = dist[t]; const bw = (n/maxN*100).toFixed(1); const c = TIER_COLORS_HEX[t];
          return `<tr class="border-b border-white/5">
            <td class="py-2"><span class="pill ${tierColor[t]}">${t}</span></td>
            <td class="text-right font-mono text-slate-200">${n}</td>
            <td class="pl-6 pr-2 w-3/5"><div class="bg-white/5 h-2 rounded-full overflow-hidden"><div class="h-full rounded-full" style="width:${bw}%;background:${c}"></div></div></td>
          </tr>`;
        }).join('')}
      </tbody></table>
      <div class="text-xs text-slate-500 mt-3">Ver el detalle completo en la pestaña <strong>👻 Lista (Dropi sin GHL)</strong>.</div>
    </div>
  `;
}

// ============================================================
// MÉTRICAS · 3 vistas independientes del VIP
// ============================================================
let metSinVipProg   = "Todos";
let metSinVipPais   = "Todos";
let metSinVipSearch = "";
let metProgFilter   = "Todos";
let metProgPais     = "Todos";
let metProgSearch   = "";
let metDropiVentas  = "Todos";   // "Todos" | "Con ventas" | "Sin ventas"
let metDropiSearch  = "";
let metDropiCountry = "Todos";
// --- Estudiantes ---
let estSubtab  = "vip";       // "vip" | "no_vip"
let estProg    = "Todos";     // "Todos" | "Master Escala" | "Iniciación Escala" | "Ambos"
let estVentas  = "Todos";     // "Todos" | "Con ventas" | "Sin ventas"
let estPais    = "Todos";
let estSearch  = "";

function downloadCSV(filename, rows) {
  const NL = String.fromCharCode(10);
  const csv = rows.map(r => r.map(c => {
    const s = (c == null ? '' : String(c));
    return (s.indexOf(',') >= 0 || s.indexOf('"') >= 0 || s.indexOf(NL) >= 0)
      ? '"' + s.replace(/"/g,'""') + '"' : s;
  }).join(',')).join(NL);
  const blob = new Blob([csv], {type:'text/csv;charset=utf-8;'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = filename; document.body.appendChild(a); a.click();
  document.body.removeChild(a); URL.revokeObjectURL(url);
}

// XLSX vía SheetJS. Se carga lazy (solo al primer click) para no inflar el
// dashboard inicial. Usa la versión Community 0.18.5 estable desde cdnjs.
let _xlsxLoading = null;
function ensureXLSX() {
  if (typeof XLSX !== 'undefined') return Promise.resolve();
  if (_xlsxLoading) return _xlsxLoading;
  _xlsxLoading = new Promise((resolve, reject) => {
    const s = document.createElement('script');
    s.src = 'https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js';
    s.onload = () => resolve();
    s.onerror = () => reject(new Error('No se pudo cargar SheetJS'));
    document.head.appendChild(s);
  });
  return _xlsxLoading;
}
// Wrapper: deshabilita el botón mientras se carga SheetJS y genera el archivo.
// Usar en el onclick: _dlXLSX(btn, "archivo.xlsx", rowsFn(), "Sheet")
function _dlXLSX(btn, filename, rows, sheetName) {
  if (!btn) return downloadXLSX(filename, rows, sheetName);
  btn.disabled = true;
  const orig = btn.textContent;
  btn.textContent = 'Generando…';
  Promise.resolve()
    .then(() => downloadXLSX(filename, rows, sheetName))
    .finally(() => setTimeout(() => { btn.disabled = false; btn.textContent = orig; }, 1200));
}

// filename: nombre incluyendo .xlsx. rows: array de arrays (misma forma que downloadCSV).
// sheetName: opcional (default "Datos"). numericCols: opcional, indices de columnas
// a formatear como número (0-based, sin contar el header).
function downloadXLSX(filename, rows, sheetName, numericCols) {
  ensureXLSX().then(() => {
    const ws = XLSX.utils.aoa_to_sheet(rows);
    // Ancho automático razonable por columna (basado en longitud del header + max
    // de las primeras 200 filas).
    const nCols = rows[0] ? rows[0].length : 0;
    const widths = [];
    for (let c = 0; c < nCols; c++) {
      let max = String(rows[0][c] || '').length;
      for (let r = 1; r < Math.min(rows.length, 201); r++) {
        const v = rows[r] && rows[r][c];
        const s = (v == null ? '' : String(v));
        if (s.length > max) max = s.length;
      }
      widths.push({ wch: Math.min(Math.max(max + 2, 8), 50) });
    }
    ws['!cols'] = widths;
    const wb = XLSX.utils.book_new();
    XLSX.utils.book_append_sheet(wb, ws, (sheetName || 'Datos').slice(0, 31));
    XLSX.writeFile(wb, filename);
  }).catch(err => {
    alert('No se pudo generar el XLSX: ' + err.message + '. Revisa tu conexión a internet.');
  });
}

// ---------- VISTA 1: No están en Comunidad VIP ----------
function renderMetSinVIP() {
  const all = DATA.metricas.sin_comunidad_vip || [];
  let list = all.slice();
  if (metSinVipProg !== "Todos") list = list.filter(u => u.programa === metSinVipProg);
  if (metSinVipPais !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metSinVipPais));
  if (metSinVipSearch) {
    const s = metSinVipSearch.toLowerCase();
    list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
  }
  const cnt = {
    Todos: all.length,
    "Master Escala": all.filter(u => u.programa==="Master Escala").length,
    "Iniciación Escala": all.filter(u => u.programa==="Iniciación Escala").length,
    "Ambos": all.filter(u => u.programa==="Ambos").length,
  };
  const listConEmail = list.filter(u => u.email && u.email.includes('@'));

  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">👥 No están en Comunidad VIP</h2>
      <div class="text-xs text-slate-500">Contactos con tag <code>escala</code> o <code>iniciacion</code> en GHL que NO tienen el tag <code>comunidad vip new</code>.</div>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      ${statCard("Total sin VIP new", cnt["Todos"], "Master o Iniciación sin la etiqueta", "neon-cyan")}
      ${statCard("Solo Master Escala", cnt["Master Escala"], "Tag 'escala' sin 'iniciacion'", "neon-violet")}
      ${statCard("Solo Iniciación", cnt["Iniciación Escala"], "Tag 'iniciacion' sin 'escala'", "neon-pink")}
      ${statCard("Ambos programas", cnt["Ambos"], "Tienen los dos tags formativos", "neon-yellow")}
    </div>

    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-3">
        <input id="met1-search" type="text" placeholder="Buscar por nombre o email..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${metSinVipSearch.replace(/"/g,'&quot;')}">
        ${paisSelectHTML('met1-pais', all, metSinVipPais)}
        <button id="met1-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="met1-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Programa:</div>
        ${["Todos","Master Escala","Iniciación Escala","Ambos"].map(p =>
          `<button data-met1prog="${p}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${metSinVipProg===p?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${p==='Todos'?'Todos':PROG_SHORT[p]} <span class="ml-1 text-slate-500">${cnt[p]||0}</span></button>`
        ).join('')}
      </div>
    </div>

    <div class="card p-4">
      <div class="flex justify-between items-center mb-2">
        <div class="text-xs text-slate-500">Mostrando ${list.length} de ${all.length} · ${listConEmail.length} con email</div>
      </div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
              <th class="text-center">Programa</th>
              <th class="text-left">País</th>
              <th class="text-left">Contact ID</th>
            </tr>
          </thead>
          <tbody>
            ${list.map(u => {
              return `
              <tr class="hover-row border-b border-white/5">
                <td class="text-slate-200">${tc(u.nombre)||'—'}</td>
                <td class="text-slate-300">${u.email||'<span class="text-slate-600">sin email</span>'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||u.programa}</span></td>
                <td class="text-[14px]">${(u.paises||[]).map(p => `<span title="${p}">${flag(p)}</span>`).join(' ')||'—'}</td>
                <td class="text-slate-500 font-mono text-[10px]">${u.cid}</td>
              </tr>`;
            }).join('')}
            ${list.length===0?`<tr><td colspan="6" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}
function wireMetSinVIP() {
  document.querySelectorAll('[data-met1prog]').forEach(b => b.onclick = () => { metSinVipProg = b.dataset.met1prog; render(); });
  const ps = document.getElementById('met1-pais');
  if (ps) ps.onchange = e => { metSinVipPais = e.target.value; render(); };
  const inp = document.getElementById('met1-search');
  if (inp) {
    inp.oninput = e => { metSinVipSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(metSinVipSearch.length, metSinVipSearch.length);
  }
  function _met1Rows() {
    const all = DATA.metricas.sin_comunidad_vip || [];
    let list = all.slice();
    if (metSinVipProg !== "Todos") list = list.filter(u => u.programa === metSinVipProg);
    if (metSinVipPais !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metSinVipPais));
    if (metSinVipSearch) {
      const s = metSinVipSearch.toLowerCase();
      list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
    }
    const rows = [["Nombre","Email","Teléfono","Programa","Países","Contact ID"]];
    list.forEach(u => rows.push([u.nombre,u.email,u.telefono,u.programa,(u.paises||[]).join('|'),u.cid]));
    return rows;
  }
  const btn = document.getElementById('met1-csv');
  if (btn) btn.onclick = () => downloadCSV("no_estan_en_comunidad_vip.csv", _met1Rows());
  const btnX = document.getElementById('met1-xlsx');
  if (btnX) btnX.onclick = () => _dlXLSX(btnX, "no_estan_en_comunidad_vip.xlsx", _met1Rows(), "Sin VIP");
}

// ---------- VISTA 2: Master vs Iniciación ----------
function renderMetProgramas() {
  const all = DATA.metricas.programas || [];
  const m = DATA.metricas;
  let list = all.slice();
  if (metProgFilter !== "Todos") list = list.filter(u => u.programa === metProgFilter);
  if (metProgPais !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metProgPais));
  if (metProgSearch) {
    const s = metProgSearch.toLowerCase();
    list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
  }
  const cnt = {
    "Todos": all.length,
    "Master Escala": m.master_total,
    "Iniciación Escala": m.iniciacion_total,
    "Ambos": m.ambos_total,
    "Sin programa": m.sin_programa_total,
  };
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">📊 Master vs Iniciación</h2>
      <div class="text-xs text-slate-500">Distribución de TODOS los contactos GHL (${fmt(m.ghl_total)}) según su tag de programa formativo.</div>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      ${statCard("🎓 Master Escala", m.master_total, "Solo tag 'escala'", "neon-violet")}
      ${statCard("🌱 Iniciación", m.iniciacion_total, "Solo tag 'iniciacion'", "neon-pink")}
      ${statCard("⚡ Ambos", m.ambos_total, "Tienen ambos tags", "neon-yellow")}
      ${statCard("⚪ Sin programa", m.sin_programa_total, "Ni Master ni Iniciación", "neon-red")}
    </div>

    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-3">
        <input id="met2-search" type="text" placeholder="Buscar por nombre o email..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${metProgSearch.replace(/"/g,'&quot;')}">
        ${paisSelectHTML('met2-pais', all, metProgPais)}
        <button id="met2-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="met2-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Programa:</div>
        ${["Todos","Master Escala","Iniciación Escala","Ambos","Sin programa"].map(p =>
          `<button data-met2prog="${p}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${metProgFilter===p?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${p==='Todos'?'Todos':(PROG_SHORT[p]||p)} <span class="ml-1 text-slate-500">${cnt[p]||0}</span></button>`
        ).join('')}
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${list.length} de ${all.length}</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
              <th class="text-center">Programa</th>
              <th class="text-left">País</th>
            </tr>
          </thead>
          <tbody>
            ${list.slice(0,1000).map(u => `
              <tr class="hover-row border-b border-white/5">
                <td class="py-2 text-slate-200">${tc(u.nombre)||'—'}</td>
                <td class="text-slate-300">${u.email||'—'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||u.programa}</span></td>
                <td class="text-[14px]">${(u.paises||[]).map(p => `<span title="${p}">${flag(p)}</span>`).join(' ')||'—'}</td>
              </tr>
            `).join('')}
            ${list.length>1000?`<tr><td colspan="5" class="text-center text-slate-500 py-3">... y ${list.length-1000} más (usa CSV para ver todos)</td></tr>`:''}
            ${list.length===0?`<tr><td colspan="5" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}
function wireMetProgramas() {
  document.querySelectorAll('[data-met2prog]').forEach(b => b.onclick = () => { metProgFilter = b.dataset.met2prog; render(); });
  const ps = document.getElementById('met2-pais');
  if (ps) ps.onchange = e => { metProgPais = e.target.value; render(); };
  const inp = document.getElementById('met2-search');
  if (inp) {
    inp.oninput = e => { metProgSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(metProgSearch.length, metProgSearch.length);
  }
  function _met2Rows() {
    const all = DATA.metricas.programas || [];
    let list = all.slice();
    if (metProgFilter !== "Todos") list = list.filter(u => u.programa === metProgFilter);
    if (metProgPais !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metProgPais));
    if (metProgSearch) {
      const s = metProgSearch.toLowerCase();
      list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
    }
    const rows = [["Nombre","Email","Teléfono","Programa","Países"]];
    list.forEach(u => rows.push([u.nombre,u.email,u.telefono,u.programa,(u.paises||[]).join('|')]));
    return rows;
  }
  const btn = document.getElementById('met2-csv');
  if (btn) btn.onclick = () => downloadCSV("master_vs_iniciacion.csv", _met2Rows());
  const btnX = document.getElementById('met2-xlsx');
  if (btnX) btnX.onclick = () => _dlXLSX(btnX, "master_vs_iniciacion.xlsx", _met2Rows(), "Programas");
}

// ---------- VISTA: Estudiantes (Escala/Iniciación + ventas Dropi) ----------
function renderEstudiantes() {
  const all = DATA.metricas.estudiantes || [];
  const m = DATA.metricas;
  const months = DATA.meta.ventana;
  // Subtab: VIP vs no-VIP
  let base = all.filter(u => estSubtab === "vip" ? u.tiene_vip_new : !u.tiene_vip_new);
  let list = base.slice();
  if (estProg !== "Todos")   list = list.filter(u => u.programa === estProg);
  if (estPais !== "Todos")   list = list.filter(u => paisMatch((u.paises||[]), estPais));
  if (estVentas === "Con ventas") list = list.filter(u => u.tiene_ventas);
  else if (estVentas === "Sin ventas") list = list.filter(u => !u.tiene_ventas);
  if (estSearch) {
    const s = estSearch.toLowerCase();
    list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s) || (u.telefono||'').includes(s));
  }
  // Subsetear ped_mes al país filtrado
  if (estPais !== "Todos") list = list.map(u => viewByPais(u, estPais));
  const monthCols = months.map(m2 => `<th class="text-right text-[10px] uppercase tracking-wider">${mesShort(m2)}</th>`).join('');
  // Conteos por programa dentro del subtab activo
  const progCnt = {
    "Todos": base.length,
    "Master Escala": base.filter(u=>u.programa==="Master Escala").length,
    "Iniciación Escala": base.filter(u=>u.programa==="Iniciación Escala").length,
    "Ambos": base.filter(u=>u.programa==="Ambos").length,
  };
  const conVentasBase = base.filter(u=>u.tiene_ventas).length;
  const ventasCnt = { "Todos": base.length, "Con ventas": conVentasBase, "Sin ventas": base.length - conVentasBase };
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">🎓 Estudiantes</h2>
      <div class="text-xs text-slate-500">Contactos con programa formativo (Escala / Iniciación) y sus pedidos de Dropi en la ventana. Separados según tengan o no la etiqueta <span class="text-cyan-400">Comunidad VIP</span>.</div>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      ${statCard("🎓 Total estudiantes", m.estudiantes_total, "Con tag Escala o Iniciación", "neon-violet")}
      ${statCard("🏆 En Comunidad VIP", m.estudiantes_vip, "Estudiantes con etiqueta VIP", "neon-green")}
      ${statCard("🚫 Sin VIP", m.estudiantes_no_vip, "Estudiantes sin etiqueta VIP", "neon-red")}
      ${statCard("💰 Con ventas Dropi", m.estudiantes_con_ventas, "Estudiantes con ≥1 pedido", "neon-yellow")}
    </div>

    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-2 mb-3">
        <button data-estsub="vip" class="text-xs px-4 py-2 rounded-lg font-semibold ${estSubtab==='vip'?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">🏆 En Comunidad VIP <span class="ml-1 text-slate-500">${m.estudiantes_vip}</span></button>
        <button data-estsub="no_vip" class="text-xs px-4 py-2 rounded-lg font-semibold ${estSubtab==='no_vip'?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">🚫 Sin etiqueta VIP <span class="ml-1 text-slate-500">${m.estudiantes_no_vip}</span></button>
      </div>
      <div class="flex flex-wrap items-center gap-3 mb-3">
        <input id="est-search" type="text" placeholder="Buscar por nombre, email o teléfono..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${estSearch.replace(/"/g,'&quot;')}">
        ${paisSelectHTML('est-pais', base, estPais)}
        <button id="est-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="est-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
      <div class="flex flex-wrap items-center gap-2 mb-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Programa:</div>
        ${["Todos","Master Escala","Iniciación Escala","Ambos"].map(p =>
          `<button data-estprog="${p}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${estProg===p?'bg-violet-600/30 text-violet-200 border border-violet-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${p==='Todos'?'Todos':(PROG_SHORT[p]||p)} <span class="ml-1 text-slate-500">${progCnt[p]||0}</span></button>`
        ).join('')}
      </div>
      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Ventas:</div>
        ${["Todos","Con ventas","Sin ventas"].map(f =>
          `<button data-estventas="${f}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${estVentas===f?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${f} <span class="ml-1 text-slate-500">${ventasCnt[f]||0}</span></button>`
        ).join('')}
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${list.length} de ${base.length} ${estSubtab==='vip'?'(con etiqueta VIP)':'(sin etiqueta VIP)'}</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
              <th class="text-center">Programa</th>
              <th class="text-left">Países</th>
              <th class="text-center">Escalafón</th>
              ${monthCols}
              <th class="text-right">Total ped.</th>
            </tr>
          </thead>
          <tbody>
            ${list.slice(0,1500).map(u => `
              <tr class="hover-row border-b border-white/5">
                <td class="py-2 text-slate-200">${tc(u.nombre)||'—'}</td>
                <td class="text-slate-300">${u.email||'—'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||u.programa}</span></td>
                <td class="text-[14px]">${(u.paises||[]).map(p => `<span title="${p}">${flag(p)}</span>`).join(' ')||'—'}</td>
                <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
                ${months.map(m2 => `<td class="text-right font-mono ${(u.ped_mes[m2]||0)===0?'text-slate-700':'text-slate-400'}">${fmt(u.ped_mes[m2])}</td>`).join('')}
                <td class="text-right font-mono font-semibold ${u.total_pedidos>0?'text-slate-100':'text-slate-600'}">${fmt(u.total_pedidos)}</td>
              </tr>
            `).join('')}
            ${list.length>1500?`<tr><td colspan="${6+months.length+1}" class="text-center text-slate-500 py-3">... y ${list.length-1500} más (usa CSV para ver todos)</td></tr>`:''}
            ${list.length===0?`<tr><td colspan="${6+months.length+1}" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}
function wireEstudiantes() {
  document.querySelectorAll('[data-estsub]').forEach(b => b.onclick = () => { estSubtab = b.dataset.estsub; render(); });
  document.querySelectorAll('[data-estprog]').forEach(b => b.onclick = () => { estProg = b.dataset.estprog; render(); });
  document.querySelectorAll('[data-estventas]').forEach(b => b.onclick = () => { estVentas = b.dataset.estventas; render(); });
  const ps = document.getElementById('est-pais');
  if (ps) ps.onchange = e => { estPais = e.target.value; render(); };
  const inp = document.getElementById('est-search');
  if (inp) {
    inp.oninput = e => { estSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(estSearch.length, estSearch.length);
  }
  function _estRows() {
    const all = DATA.metricas.estudiantes || [];
    const months = DATA.meta.ventana;
    let base = all.filter(u => estSubtab === "vip" ? u.tiene_vip_new : !u.tiene_vip_new);
    let list = base.slice();
    if (estProg !== "Todos") list = list.filter(u => u.programa === estProg);
    if (estPais !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), estPais));
    if (estVentas === "Con ventas") list = list.filter(u => u.tiene_ventas);
    else if (estVentas === "Sin ventas") list = list.filter(u => !u.tiene_ventas);
    if (estSearch) {
      const s = estSearch.toLowerCase();
      list = list.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s) || (u.telefono||'').includes(s));
    }
    if (estPais !== "Todos") list = list.map(u => viewByPais(u, estPais));
    const header = ["Nombre","Email","Teléfono","Programa","VIP","Países","Escalafón", ...months, "Total pedidos"];
    const rows = [header];
    list.forEach(u => rows.push([u.nombre,u.email,u.telefono,u.programa,u.tiene_vip_new?"Sí":"No",(u.paises||[]).join('|'),u.nivel, ...months.map(mo=>u.ped_mes[mo]||0), u.total_pedidos]));
    return rows;
  }
  const btn = document.getElementById('est-csv');
  if (btn) btn.onclick = () => downloadCSV(`estudiantes_${estSubtab}.csv`, _estRows());
  const btnX = document.getElementById('est-xlsx');
  if (btnX) btnX.onclick = () => _dlXLSX(btnX, `estudiantes_${estSubtab}.xlsx`, _estRows(), "Estudiantes");
}

// ---------- VISTA 3: En Dropi sin GHL ----------
function renderMetDropiGHL() {
  const all = DATA.metricas.dropi_sin_ghl || [];
  const m = DATA.metricas;
  const months = DATA.meta.ventana;
  let list = all.slice();
  if (metDropiVentas === "Con ventas") list = list.filter(u => u.tiene_ventas);
  else if (metDropiVentas === "Sin ventas") list = list.filter(u => !u.tiene_ventas);
  if (metDropiCountry !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metDropiCountry));
  if (metDropiSearch) {
    const s = metDropiSearch.toLowerCase();
    list = list.filter(u => (u.email||'').toLowerCase().includes(s) || (u.nombre||'').toLowerCase().includes(s) || (u.telefono||'').includes(s));
  }
  const allCountries = paisesUnicos(all.flatMap(u => u.paises||[]));
  const monthCols = months.map(m2 => `<th class="text-right text-[10px] uppercase tracking-wider">${mesShort(m2)}</th>`).join('');
  const cnt = {
    "Todos":      all.length,
    "Con ventas": m.dropi_sin_ghl_con_ventas,
    "Sin ventas": m.dropi_sin_ghl_sin_ventas,
  };
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">👻 En Dropi pero NO están en GHL</h2>
      <div class="text-xs text-slate-500">Correos que aparecen en los Excels de Dropi pero no existen como contacto principal ni como tienda en ningún contacto de GHL.</div>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mb-4">
      ${statCard("Total Dropi sin GHL", all.length, `de ${fmt(m.dropi_emails_total)} correos en Dropi`, "neon-cyan")}
      ${statCard("Con ventas", m.dropi_sin_ghl_con_ventas, "Tienen ≥1 pedido en la ventana", "neon-green")}
      ${statCard("Sin ventas", m.dropi_sin_ghl_sin_ventas, "0 pedidos en toda la ventana", "neon-red")}
      ${statCard("Universo GHL", m.ghl_emails_universo, "Emails únicos en GHL (principal + tiendas)", "neon-violet")}
    </div>

    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3 mb-3">
        <input id="met3-search" type="text" placeholder="Buscar por email, nombre o teléfono..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${metDropiSearch.replace(/"/g,'&quot;')}">
        <select id="met3-country" class="bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm text-slate-200 focus:outline-none focus:border-cyan-500">
          <option value="Todos">Todos los países</option>
          ${allCountries.map(p => `<option value="${p}" ${metDropiCountry===p?'selected':''}>${flag(p)} ${p}</option>`).join('')}
        </select>
        <button id="met3-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="met3-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
      <div class="flex flex-wrap items-center gap-2">
        <div class="text-[10px] uppercase tracking-wider text-slate-500 w-20">Ventas:</div>
        ${["Todos","Con ventas","Sin ventas"].map(f =>
          `<button data-met3ventas="${f}" class="text-[11px] px-3 py-1.5 rounded-lg font-medium ${metDropiVentas===f?'bg-cyan-600/30 text-cyan-200 border border-cyan-500/40':'bg-white/5 text-slate-400 border border-white/5 hover:text-slate-200'}">${f} <span class="ml-1 text-slate-500">${cnt[f]||0}</span></button>`
        ).join('')}
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${list.length} de ${all.length}</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2">Email</th>
              <th class="text-left">Nombre</th>
              <th class="text-left">Teléfono</th>
              <th class="text-left">Países</th>
              <th class="text-center">Programa</th>
              <th class="text-center">Escalafón</th>
              ${monthCols}
              <th class="text-right">Total ped.</th>
              <th class="text-center">Meses act.</th>
            </tr>
          </thead>
          <tbody>
            ${list.slice(0,1500).map(u => `
              <tr class="hover-row border-b border-white/5">
                <td class="py-2 text-slate-200">${u.email}</td>
                <td class="text-slate-300">${tc(u.nombre)||'—'}</td>
                <td class="text-slate-400 font-mono">${u.telefono||'—'}</td>
                <td class="text-[14px]">${(u.paises||[]).map(p => `<span title="${p}">${flag(p)}</span>`).join(' ')||'—'}</td>
                <td class="text-center"><span class="pill bg-white/5 border-white/10 text-slate-300">${PROG_SHORT[u.programa]||u.programa||'—'}</span></td>
                <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
                ${months.map(m2 => `<td class="text-right font-mono ${(u.ped_mes[m2]||0)===0?'text-slate-700':'text-slate-400'}">${fmt(u.ped_mes[m2])}</td>`).join('')}
                <td class="text-right font-mono font-semibold ${u.total_pedidos>0?'text-slate-100':'text-slate-600'}">${fmt(u.total_pedidos)}</td>
                <td class="text-center font-mono text-slate-400">${u.n_meses_activos}/${months.length}</td>
              </tr>
            `).join('')}
            ${list.length>1500?`<tr><td colspan="${6+months.length+2}" class="text-center text-slate-500 py-3">... y ${list.length-1500} más (usa CSV para ver todos)</td></tr>`:''}
            ${list.length===0?`<tr><td colspan="${6+months.length+2}" class="text-center text-slate-500 py-6">— sin resultados —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}
// ---------- VISTA 4: Posibles duplicados ----------
let metDupSearch = "";
let metDupPais = "Todos";
function renderMetDuplicados() {
  const all = DATA.metricas.duplicados || [];
  const m = DATA.metricas;
  let list = all.slice();
  if (metDupPais !== "Todos") list = list.filter(d => d.tienda_pais === metDupPais);
  if (metDupSearch) {
    const s = metDupSearch.toLowerCase();
    list = list.filter(d =>
      (d.nombre||'').toLowerCase().includes(s) ||
      (d.email_principal||'').toLowerCase().includes(s) ||
      (d.tienda_email||'').toLowerCase().includes(s) ||
      (d.otro_nombre||'').toLowerCase().includes(s)
    );
  }
  const dupPaises = [...new Set(all.map(d => d.tienda_pais).filter(Boolean))].sort();
  return `
    <div class="card p-4 mb-4">
      <h2 class="text-base font-bold neon-cyan mb-1">🔁 Posibles duplicados</h2>
      <div class="text-xs text-slate-500 leading-relaxed">
        Contactos cuyo correo de tienda coincide con el <strong>email principal de otro contacto distinto</strong>.<br>
        Ejemplo: <span class="text-slate-300">"Diego Adolfo"</span> tiene como Tienda 3 el correo <code>diego@gmail.com</code>,
        que es el email principal de <span class="text-slate-300">"Diego Forero"</span> → posible duplicado o cuenta compartida.
      </div>
    </div>

    <div class="grid grid-cols-2 md:grid-cols-3 gap-3 mb-4">
      ${statCard("Coincidencias totales", m.duplicados_total, "Pares contacto ↔ tienda detectados", "neon-cyan")}
      ${statCard("Contactos involucrados", m.duplicados_contactos_unicos, "Contactos únicos con al menos un cruce", "neon-violet")}
      ${statCard("Universo GHL", m.ghl_total, "Sobre los que se hace el cruce", "neon-yellow")}
    </div>

    <div class="card p-4 mb-4">
      <div class="flex flex-wrap items-center gap-3">
        <input id="met4-search" type="text" placeholder="Buscar por nombre o correo..."
               class="flex-1 min-w-[260px] bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm focus:outline-none focus:border-cyan-500"
               value="${metDupSearch.replace(/"/g,'&quot;')}">
        <select id="met4-pais" class="bg-black/40 border border-white/10 rounded-lg px-3 py-2 text-sm text-slate-200 focus:outline-none focus:border-cyan-500">
          <option value="Todos">Todos los países</option>
          ${dupPaises.map(p => `<option value="${p}" ${metDupPais===p?'selected':''}>${flag(p)} ${p}</option>`).join('')}
        </select>
        <button id="met4-csv" class="text-[11px] px-3 py-2 rounded-lg bg-cyan-600/30 text-cyan-200 border border-cyan-500/40 hover:bg-cyan-600/40">⬇ CSV</button>
        <button id="met4-xlsx" class="text-[11px] px-3 py-2 rounded-lg bg-emerald-600/30 text-emerald-200 border border-emerald-500/40 hover:bg-emerald-600/40">⬇ XLSX</button>
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-slate-500 mb-2">Mostrando ${list.length} de ${all.length}</div>
      <div class="overflow-x-auto scrollable">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10 sticky top-0 bg-[#06091a] z-10">
            <tr>
              <th class="text-left py-2" colspan="4">Contacto con la tienda</th>
              <th class="text-left border-l border-white/10 pl-3" colspan="3">↔ Coincide con email principal de</th>
            </tr>
            <tr>
              <th class="text-left">Nombre</th>
              <th class="text-left">Email principal</th>
              <th class="text-left">Slot</th>
              <th class="text-left">Email tienda</th>
              <th class="text-left border-l border-white/10 pl-3">Nombre</th>
              <th class="text-left">Email</th>
              <th class="text-left">Teléfono</th>
            </tr>
          </thead>
          <tbody>
            ${list.map(d => `
              <tr class="hover-row border-b border-white/5">
                <td class="py-2 text-slate-200">${tc(d.nombre)||'—'}</td>
                <td class="text-slate-300">${d.email_principal||'<span class="text-slate-600">—</span>'}</td>
                <td class="text-center"><span class="pill bg-violet-500/20 text-violet-300 border-violet-500/40">${d.tienda_slot}${d.tienda_pais?' · '+flag(d.tienda_pais)+' '+d.tienda_pais:''}</span></td>
                <td class="text-amber-300 font-mono">${d.tienda_email}</td>
                <td class="text-slate-200 border-l border-white/10 pl-3">${tc(d.otro_nombre)||'—'}</td>
                <td class="text-amber-300 font-mono">${d.tienda_email}</td>
                <td class="text-slate-400 font-mono">${d.otro_telefono||'—'}</td>
              </tr>
            `).join('')}
            ${list.length===0?`<tr><td colspan="7" class="text-center text-slate-500 py-6">— no hay coincidencias —</td></tr>`:''}
          </tbody>
        </table>
      </div>
    </div>
  `;
}
function wireMetDuplicados() {
  const ps = document.getElementById('met4-pais');
  if (ps) ps.onchange = e => { metDupPais = e.target.value; render(); };
  const inp = document.getElementById('met4-search');
  if (inp) {
    inp.oninput = e => { metDupSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(metDupSearch.length, metDupSearch.length);
  }
  function _met4Rows() {
    const all = DATA.metricas.duplicados || [];
    let list = all.slice();
    if (metDupPais !== "Todos") list = list.filter(d => d.tienda_pais === metDupPais);
    if (metDupSearch) {
      const s = metDupSearch.toLowerCase();
      list = list.filter(d =>
        (d.nombre||'').toLowerCase().includes(s) ||
        (d.email_principal||'').toLowerCase().includes(s) ||
        (d.tienda_email||'').toLowerCase().includes(s) ||
        (d.otro_nombre||'').toLowerCase().includes(s)
      );
    }
    const rows = [["Contacto B (tiene la tienda)","Email principal B","Teléfono B","Slot tienda","País tienda","Email tienda","Contacto A (email coincide)","Email principal A","Teléfono A","CID B","CID A"]];
    list.forEach(d => rows.push([
      d.nombre, d.email_principal, d.telefono, d.tienda_slot, d.tienda_pais,
      d.tienda_email, d.otro_nombre, d.tienda_email, d.otro_telefono, d.cid, d.otro_cid
    ]));
    return rows;
  }
  const btn = document.getElementById('met4-csv');
  if (btn) btn.onclick = () => downloadCSV("posibles_duplicados.csv", _met4Rows());
  const btnX = document.getElementById('met4-xlsx');
  if (btnX) btnX.onclick = () => _dlXLSX(btnX, "posibles_duplicados.xlsx", _met4Rows(), "Duplicados");
}

function wireMetDropiGHL() {
  document.querySelectorAll('[data-met3ventas]').forEach(b => b.onclick = () => { metDropiVentas = b.dataset.met3ventas; render(); });
  const cs = document.getElementById('met3-country');
  if (cs) cs.onchange = e => { metDropiCountry = e.target.value; render(); };
  const inp = document.getElementById('met3-search');
  if (inp) {
    inp.oninput = e => { metDropiSearch = e.target.value; render(); };
    inp.focus(); inp.setSelectionRange(metDropiSearch.length, metDropiSearch.length);
  }
  function _met3Rows() {
    const all = DATA.metricas.dropi_sin_ghl || [];
    const months = DATA.meta.ventana;
    let list = all.slice();
    if (metDropiVentas === "Con ventas") list = list.filter(u => u.tiene_ventas);
    else if (metDropiVentas === "Sin ventas") list = list.filter(u => !u.tiene_ventas);
    if (metDropiCountry !== "Todos") list = list.filter(u => paisMatch((u.paises||[]), metDropiCountry));
    if (metDropiSearch) {
      const s = metDropiSearch.toLowerCase();
      list = list.filter(u => (u.email||'').toLowerCase().includes(s) || (u.nombre||'').toLowerCase().includes(s) || (u.telefono||'').includes(s));
    }
    const header = ["Email","Nombre","Teléfono","Países","Programa","Escalafón", ...months, "Total pedidos","Meses activos"];
    const rows = [header];
    list.forEach(u => rows.push([u.email,u.nombre,u.telefono,(u.paises||[]).join('|'),u.programa||'Sin programa',u.nivel, ...months.map(m=>u.ped_mes[m]||0), u.total_pedidos, u.n_meses_activos]));
    return rows;
  }
  const btn = document.getElementById('met3-csv');
  if (btn) btn.onclick = () => downloadCSV("dropi_sin_ghl.csv", _met3Rows());
  const btnX = document.getElementById('met3-xlsx');
  if (btnX) btnX.onclick = () => _dlXLSX(btnX, "dropi_sin_ghl.xlsx", _met3Rows(), "Dropi sin GHL");
}

function wireClasificacion() {
  document.querySelectorAll('th[data-sort]').forEach(th => th.onclick = () => {
    const col = th.dataset.sort;
    if (sortCol === col) { sortDir = (sortDir === 'desc' ? 'asc' : 'desc'); }
    else { sortCol = col; sortDir = 'desc'; }
    render();
  });
  document.querySelectorAll('[data-tier]').forEach(b => b.onclick = () => {
    const t = b.dataset.tier;
    if (t === 'Todos') currentTiers.clear();
    else { currentTiers.has(t) ? currentTiers.delete(t) : currentTiers.add(t); }
    render();
  });
  document.querySelectorAll('[data-prog]').forEach(b => b.onclick = () => {
    const p = b.dataset.prog;
    if (p === 'Todos') currentProgs.clear();
    else { currentProgs.has(p) ? currentProgs.delete(p) : currentProgs.add(p); }
    render();
  });
  document.querySelectorAll('[data-ventas]').forEach(b => b.onclick = () => { currentVentas = b.dataset.ventas; render(); });
  document.querySelectorAll('[data-multipais]').forEach(b => b.onclick = () => { currentMultipais = !currentMultipais; render(); });
  document.querySelectorAll('[data-cid]').forEach(row => row.onclick = () => abrirFicha(row.dataset.cid));
  const cs = document.getElementById('country-select');
  if (cs) cs.onchange = e => { currentCountry = e.target.value; render(); };
  const inp = document.getElementById('search-input');
  if (inp) {
    inp.oninput = e => { currentSearch = e.target.value; render(); };
    inp.focus();
    inp.setSelectionRange(currentSearch.length, currentSearch.length);
  }
  // Export: respeta filtros + orden actuales de la tabla de Clasificación
  function _clasifRows() {
    const users = baseUsers();
    let scope = users;
    if (currentVentas === "Con ventas") scope = scope.filter(u => (u.total_pedidos||0) > 0);
    else if (currentVentas === "Sin ventas") scope = scope.filter(u => (u.total_pedidos||0) === 0);
    if (currentProgs.size) scope = scope.filter(u => currentProgs.has(u.programa));
    if (currentCountry !== "Todos") scope = scope.filter(u => paisMatch((u.paises_unicos||[]), currentCountry));
    if (currentMultipais) scope = scope.filter(u => (u.paises_unicos||[]).length > 1);
    if (currentSearch) {
      const s = currentSearch.toLowerCase();
      scope = scope.filter(u => (u.nombre||'').toLowerCase().includes(s) || (u.email||'').toLowerCase().includes(s));
    }
    let filtered = scope;
    if (currentTiers.size) filtered = filtered.filter(u => currentTiers.has(u.nivel));
    const dir = sortDir === 'asc' ? -1 : 1;
    if (sortCol === 'nivel') filtered.sort((a,b)=> dir*(TIER_ORDER.indexOf(a.nivel) - TIER_ORDER.indexOf(b.nivel)) || (b.total_pedidos - a.total_pedidos));
    else if (sortCol === 'total') filtered.sort((a,b)=> dir*((b.total_pedidos||0) - (a.total_pedidos||0)));
    else if (sortCol === 'suma_top3') filtered.sort((a,b)=> dir*((b.suma_top3||0) - (a.suma_top3||0)));
    else if (sortCol === 'pct_dev') filtered.sort((a,b)=> dir*((b.pct_dev||0) - (a.pct_dev||0)));
    else filtered.sort((a,b)=> dir*(((b.ped_mes&&b.ped_mes[sortCol])||0) - ((a.ped_mes&&a.ped_mes[sortCol])||0)));
    const months = DATA.meta.ventana;
    const header = ["Nombre","Email","Teléfono","Nivel","Programa","Países", ...months, "Total pedidos","Top-3 (escalafón)","% Dev.","Semáforo","Alerta"];
    const rows = [header];
    filtered.forEach(u => rows.push([
      u.nombre, u.email, u.telefono,
      u.nivel==='Sin clasificar'?'Sin nivel':u.nivel,
      PROG_SHORT[u.programa]||u.programa||'Sin programa',
      (u.paises_unicos||u.paises||[]).join('|'),
      ...months.map(m => u.ped_mes[m]||0),
      u.total_pedidos, u.suma_top3, u.pct_dev,
      u.semaforo||'', u.alerta_tipo||''
    ]));
    return rows;
  }
  const btnC = document.getElementById('clasif-csv');
  if (btnC) btnC.onclick = () => downloadCSV("clasificacion_vip.csv", _clasifRows());
  const btnCX = document.getElementById('clasif-xlsx');
  if (btnCX) btnCX.onclick = () => _dlXLSX(btnCX, "clasificacion_vip.xlsx", _clasifRows(), "Clasificación VIP");
}

function initials(name) {
  const parts = (name||'').trim().split(/\s+/).slice(0,2);
  return parts.map(p => p.charAt(0).toUpperCase()).join('') || '??';
}

function cerrarFicha() {
  document.getElementById('ficha-modal').classList.add('hidden');
}

let fichaChart = null;
function abrirFicha(cid) {
  const u = DATA.usuarios.find(x => x.cid === cid);
  if (!u) return;
  const months = DATA.meta.ventana;
  const months_labels = months.map(mesShort);
  const totalEntregados = months.reduce((s,m) => s + (u.ent_mes[m]||0), 0);
  const totalDevoluciones = months.reduce((s,m) => s + (u.dev_mes[m]||0), 0);
  const totalPedidos = u.total_pedidos;
  const pctDevTotal = totalEntregados > 0 ? ((totalDevoluciones/(totalEntregados+totalDevoluciones))*100).toFixed(1) : 0;
  // Últimos 3 meses (cronológicos)
  const last3 = months.slice(-3);
  const last3_ent = last3.reduce((s,m) => s + (u.ent_mes[m]||0), 0);
  const last3_dev = last3.reduce((s,m) => s + (u.dev_mes[m]||0), 0);
  const last3_ped = last3.reduce((s,m) => s + (u.ped_mes[m]||0), 0);
  const last3_pct = last3_ent > 0 ? ((last3_dev/(last3_ent+last3_dev))*100).toFixed(1) : 0;

  const rows = months.map((m, i) => {
    const ent = u.ent_mes[m]||0, dev = u.dev_mes[m]||0, ped = u.ped_mes[m]||0;
    const pct = ped > 0 ? ((dev/ped)*100).toFixed(1) : 0;
    let tend = '—', tendCol = 'text-slate-500';
    if (i > 0) {
      const prevPed = u.ped_mes[months[i-1]]||0;
      if (prevPed > 0) {
        const delta = ((ped - prevPed) / prevPed * 100);
        if (Math.abs(delta) >= 10) {
          tend = (delta > 0 ? '▲' : '▼') + Math.abs(delta).toFixed(0) + '%';
          tendCol = delta > 0 ? 'text-green-400' : 'text-red-400';
        }
      }
    }
    return `<tr class="border-b border-white/5">
      <td class="py-1.5 text-slate-200">${mesShort(m)}</td>
      <td class="text-right font-mono">${fmt(ent)}</td>
      <td class="text-right font-mono">${fmt(dev)}</td>
      <td class="text-right font-mono font-semibold">${fmt(ped)}</td>
      <td class="text-right font-mono ${pct>15?'text-orange-400':pct>10?'text-yellow-400':'text-slate-400'}">${pct}%</td>
      <td class="text-center"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></td>
      <td class="text-center font-mono text-xs ${tendCol}">${tend}</td>
    </tr>`;
  }).join('');

  // Agrupar tiendas por país (una entrada por país, con lista de correos)
  const tiendasPorPais = {};
  (u.tiendas_detalle||[]).forEach(t => {
    const key = (t.pais && t.pais.trim()) || 'Sin país';
    if (!tiendasPorPais[key]) tiendasPorPais[key] = [];
    tiendasPorPais[key].push(t);
  });
  const paisesOrdenados = Object.keys(tiendasPorPais).sort((a,b) => a.localeCompare(b));
  // Set de países declarados en GHL (canónicos) — se usa para marcar ⚠ los
  // países donde el email vende en Dropi pero NO hay tienda registrada en GHL.
  const _declaradosSet = new Set((u.paises_declarados||[]).map(_paisKey));
  const paisEsDeclarado = pais => _declaradosSet.has(_paisKey(pais));

  document.getElementById('ficha-content').innerHTML = `
    <div class="card p-4 mb-3">
      <div class="flex items-start gap-3">
        <div class="w-12 h-12 rounded-full bg-gradient-to-br from-cyan-500/40 to-blue-700/40 border border-cyan-500/30 flex items-center justify-center text-sm font-bold flex-shrink-0">${initials(u.nombre)}</div>
        <div class="flex-1">
          <h3 class="text-lg font-bold">${tc(u.nombre)||'—'}</h3>
          <div class="text-xs text-slate-400">${u.email||'—'}</div>
          <div class="mt-1.5"><span class="pill ${tierColor[u.nivel]}">${u.nivel==='Sin clasificar'?'Sin nivel':u.nivel}</span></div>
        </div>
      </div>
      <div class="grid grid-cols-2 md:grid-cols-5 gap-3 mt-4">
        <div><div class="text-xl font-bold neon-cyan">${fmt(last3_ped)}</div><div class="text-[9px] uppercase tracking-wider text-slate-500">Total ventana</div></div>
        <div><div class="text-xl font-bold">${fmt(u.ped_mes[months[months.length-1]]||0)}</div><div class="text-[9px] uppercase tracking-wider text-slate-500">${mesShort(months[months.length-1])}</div></div>
        <div><div class="text-xl font-bold ${last3_pct>15?'text-orange-400':last3_pct>10?'text-yellow-400':'neon-green'}">${last3_pct}%</div><div class="text-[9px] uppercase tracking-wider text-slate-500">% Dev.</div></div>
        <div><div class="text-xl font-bold">${u.n_tiendas}</div><div class="text-[9px] uppercase tracking-wider text-slate-500">Tiendas</div></div>
        <div><div class="text-sm font-bold leading-tight">${(u.paises_unicos||[]).map(p=>flag(p)+' '+p).join('<br>')||'—'}</div><div class="text-[9px] uppercase tracking-wider text-slate-500 mt-0.5">Países</div></div>
      </div>
    </div>

    <!-- TIENDAS: agrupadas por país + lista de correos. Visible desde arriba. -->
    <div class="card p-4 mb-3">
      <div class="flex items-baseline justify-between mb-2">
        <h4 class="text-sm font-semibold">Tiendas por país <span class="text-[10px] text-slate-500">(${(u.tiendas_detalle||[]).length} correo${(u.tiendas_detalle||[]).length===1?'':'s'} en ${paisesOrdenados.length} país${paisesOrdenados.length===1?'':'es'})</span></h4>
        ${(u.tiendas_detalle||[]).length ? `<button onclick="navigator.clipboard.writeText(${JSON.stringify((u.tiendas_detalle||[]).map(t=>t.email).join('\\n'))}); this.textContent='✓ Copiado';" class="text-[10px] px-2 py-1 rounded bg-cyan-600/20 text-cyan-300 border border-cyan-500/30 hover:bg-cyan-600/30">📋 Copiar correos</button>` : ''}
      </div>
      ${u.sin_tienda ? '<div class="text-sm text-red-400">⚠ Este contacto no tiene ninguna Tienda 1..10 cargada en GHL.</div>' : `
      <div class="grid grid-cols-1 md:grid-cols-2 gap-2">
        ${paisesOrdenados.map(pais => `
          <div class="rounded-lg border border-white/10 bg-black/30 p-2.5">
            <div class="flex items-center gap-2 mb-1.5">
              <span class="text-base leading-none">${flag(pais==='Sin país'?'':pais)}</span>
              <span class="text-xs font-semibold text-slate-200">${pais}</span>
              <span class="text-[10px] text-slate-500">· ${tiendasPorPais[pais].length} correo${tiendasPorPais[pais].length===1?'':'s'}</span>
            </div>
            <ul class="space-y-0.5 text-[11px] font-mono text-slate-300 pl-1">
              ${tiendasPorPais[pais].map(t => `<li class="truncate" title="${t.email}${t.primera_vez?' · '+t.primera_vez:''}">• ${t.email}${t.primera_vez?` <span class="text-slate-500 font-sans">(${t.primera_vez})</span>`:''}</li>`).join('')}
            </ul>
          </div>
        `).join('')}
      </div>`}
    </div>

    ${(() => {
      // Ventas por país mes a mes (usa u.ped_mes_pais / ent_mes_pais / dev_mes_pais).
      // Muestra una columna por país + totales, para comparar el rendimiento
      // de cada tienda del contacto durante la ventana.
      const pmp = u.ped_mes_pais || {};
      const emp = u.ent_mes_pais || {};
      const dmp = u.dev_mes_pais || {};
      const paisesConVentas = Object.keys(pmp).filter(p => Object.values(pmp[p]||{}).some(v => v > 0)).sort();
      if (!paisesConVentas.length) return '';
      // Totales por país (para header)
      const totPais = {};
      paisesConVentas.forEach(p => totPais[p] = months.reduce((s,m) => s + (pmp[p][m]||0), 0));
      const huerfanos = paisesConVentas.filter(p => !paisEsDeclarado(p));
      return `
      <div class="card p-4 mb-3">
        <h4 class="text-sm font-semibold mb-1">Ventas por país (mes a mes)</h4>
        <div class="text-[11px] text-slate-500 mb-2">Desglose de pedidos según la tienda del país. Pedidos = Entregados + Devoluciones.${huerfanos.length ? ` <span class="text-amber-400">⚠ ${huerfanos.map(flag).join(' ')} vende sin tienda declarada en GHL.</span>` : ''}</div>
        <div class="overflow-x-auto">
        <table class="w-full text-xs">
          <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10">
            <tr>
              <th class="text-left py-2">Mes</th>
              ${paisesConVentas.map(p => {
                const declarado = paisEsDeclarado(p);
                const mark = declarado ? '' : ' <span class="text-amber-400" title="Sin tienda declarada en GHL">⚠</span>';
                const cls = declarado ? '' : ' text-amber-300';
                return `<th class="text-right px-2${cls}"><span class="text-base leading-none">${flag(p)}</span> ${p}${mark}</th>`;
              }).join('')}
              <th class="text-right px-2 border-l border-white/10">Total</th>
            </tr>
          </thead>
          <tbody>
            ${months.map(m => {
              const total_m = paisesConVentas.reduce((s,p) => s + (pmp[p][m]||0), 0);
              return `<tr class="border-b border-white/5">
                <td class="py-1.5 text-slate-300">${mesShort(m)}</td>
                ${paisesConVentas.map(p => `<td class="text-right font-mono ${(pmp[p][m]||0)===0?'text-slate-600':'text-slate-200'}">${fmt(pmp[p][m]||0)}</td>`).join('')}
                <td class="text-right font-mono font-semibold text-cyan-300 border-l border-white/10">${fmt(total_m)}</td>
              </tr>`;
            }).join('')}
            <tr class="border-t border-white/20 bg-white/[0.03]">
              <td class="py-2 font-semibold neon-cyan">Total</td>
              ${paisesConVentas.map(p => `<td class="text-right font-mono font-semibold">${fmt(totPais[p])}</td>`).join('')}
              <td class="text-right font-mono font-semibold text-cyan-300 border-l border-white/10">${fmt(Object.values(totPais).reduce((s,v)=>s+v,0))}</td>
            </tr>
            <tr class="text-[10px] text-slate-500">
              <td class="pt-1">Entregados</td>
              ${paisesConVentas.map(p => `<td class="text-right font-mono">${fmt(months.reduce((s,m) => s + ((emp[p]||{})[m]||0), 0))}</td>`).join('')}
              <td class="text-right font-mono border-l border-white/10">${fmt(paisesConVentas.reduce((s,p) => s + months.reduce((s2,m) => s2 + ((emp[p]||{})[m]||0), 0), 0))}</td>
            </tr>
            <tr class="text-[10px] text-slate-500">
              <td>Devoluciones</td>
              ${paisesConVentas.map(p => `<td class="text-right font-mono">${fmt(months.reduce((s,m) => s + ((dmp[p]||{})[m]||0), 0))}</td>`).join('')}
              <td class="text-right font-mono border-l border-white/10">${fmt(paisesConVentas.reduce((s,p) => s + months.reduce((s2,m) => s2 + ((dmp[p]||{})[m]||0), 0), 0))}</td>
            </tr>
          </tbody>
        </table>
        </div>
      </div>`;
    })()}

    <div class="card p-4 mb-3">
      <h4 class="text-sm font-semibold mb-1">Historial por mes</h4>
      <div class="text-[11px] text-slate-500 mb-2">Pedidos VIP = Entregados + Devoluciones</div>
      <table class="w-full text-xs">
        <thead class="text-[10px] text-slate-500 uppercase tracking-wider border-b border-white/10">
          <tr><th class="text-left py-2">Mes</th><th class="text-right">Entregados</th><th class="text-right">Devoluciones</th><th class="text-right">Pedidos VIP</th><th class="text-right">% Dev.</th><th class="text-center">Nivel mes</th><th class="text-center">Tend.</th></tr>
        </thead>
        <tbody>
          ${rows}
          <tr class="border-b border-white/10 bg-white/[0.02]">
            <td class="py-2 font-semibold neon-cyan">Total general</td>
            <td class="text-right font-mono font-semibold">${fmt(totalEntregados)}</td>
            <td class="text-right font-mono font-semibold">${fmt(totalDevoluciones)}</td>
            <td class="text-right font-mono font-semibold">${fmt(totalPedidos)}</td>
            <td class="text-right font-mono font-semibold">${pctDevTotal}%</td>
            <td></td><td></td>
          </tr>
          <tr class="bg-amber-500/10">
            <td class="py-2 font-semibold neon-yellow">Últimos 3 meses activos</td>
            <td class="text-right font-mono font-semibold neon-yellow">${fmt(last3_ent)}</td>
            <td class="text-right font-mono font-semibold neon-yellow">${fmt(last3_dev)}</td>
            <td class="text-right font-mono font-semibold neon-yellow">${fmt(last3_ped)}</td>
            <td class="text-right font-mono font-semibold neon-yellow">${last3_pct}%</td>
            <td></td><td></td>
          </tr>
        </tbody>
      </table>
    </div>

    <div class="card p-4">
      <h4 class="text-[11px] font-semibold uppercase tracking-wider text-slate-500 mb-2">Evolución pedidos VIP</h4>
      <canvas id="ficha-chart" height="140"></canvas>
    </div>
  `;
  document.getElementById('ficha-modal').classList.remove('hidden');

  if (!u.sin_tienda) {
    if (fichaChart) fichaChart.destroy();
    fichaChart = new Chart(document.getElementById('ficha-chart'), {
      type: 'line',
      data: {
        labels: months_labels,
        datasets: [{
          label: 'Pedidos VIP',
          data: months.map(m => u.ped_mes[m]||0),
          borderColor: '#a78bfa', backgroundColor: 'rgba(167,139,250,0.15)',
          borderWidth: 2, tension: 0.3, fill: true, pointRadius: 5, pointBackgroundColor: '#a78bfa',
        }]
      },
      options: {
        plugins: { legend: { labels: { color:'#cbd5e1' } } },
        scales: {
          y: { ticks: {color:'#94a3b8'}, grid:{color:'rgba(255,255,255,0.05)'} },
          x: { ticks: {color:'#94a3b8'}, grid:{display:false} }
        }
      }
    });
  }
}

document.addEventListener('keydown', e => { if (e.key === 'Escape') cerrarFicha(); });

function wireConsulta() {
  const inp = document.getElementById("consulta-input");
  const out = document.getElementById("consulta-result");
  out.innerText = "Escribe algo para buscar.";
  inp.oninput = e => {
    const q = e.target.value.toLowerCase().trim();
    if (!q) { out.innerText = "Escribe algo para buscar."; return; }
    const hits = DATA.usuarios.filter(u =>
      (u.nombre||'').toLowerCase().includes(q) ||
      (u.email||'').toLowerCase().includes(q) ||
      (u.cid||'').toLowerCase().includes(q)
    ).slice(0,5);
    if (!hits.length) { out.innerText = "Sin coincidencias."; return; }
    out.innerHTML = hits.map(u => `
      <div class="card p-4 mb-3">
        <div class="flex justify-between items-baseline mb-2">
          <div><div class="font-semibold text-slate-200">${tc(u.nombre)||'—'}</div>
               <div class="text-xs text-slate-500">${u.email||''} · <code class="text-[10px]">${u.cid}</code></div></div>
          <span class="pill ${tierColor[u.nivel]}">${u.nivel}</span>
        </div>
        <div class="grid grid-cols-2 md:grid-cols-5 gap-3 text-xs">
          ${DATA.meta.ventana.map(m => `<div><div class="text-slate-500">${m}</div><div class="font-mono text-slate-200">${fmt(u.ped_mes[m])}</div></div>`).join('')}
        </div>
        <div class="grid grid-cols-2 md:grid-cols-5 gap-3 text-xs mt-3">
          <div><div class="text-slate-500">Top-2</div><div class="font-mono text-slate-200">${fmt(u.suma_top2)}</div></div>
          <div><div class="text-slate-500">Top-3</div><div class="font-mono text-slate-200">${fmt(u.suma_top3)}</div></div>
          <div><div class="text-slate-500">Meses con ventas</div><div class="font-mono text-slate-200">${u.active}/${DATA.meta.ventana.length}</div></div>
          <div><div class="text-slate-500">N° tiendas</div><div class="font-mono text-slate-200">${u.n_tiendas}</div></div>
          <div><div class="text-slate-500">Programa</div><div class="font-mono text-slate-200 text-[10px]">${u.programa}</div></div>
        </div>
      </div>`).join('');
  };
  inp.focus();
}

function drawCharts() {
  const _st = computeStats(baseUsers());
  const d = _st.dist;
  // 1. Distribución VIP (donut)
  new Chart(document.getElementById("chart-donut"), {
    type: 'doughnut',
    data: {
      labels: TIER_ORDER.filter(t=>d[t].n>0).map(t => t==='Sin clasificar'?'Sin nivel':t),
      datasets: [{
        data: TIER_ORDER.filter(t=>d[t].n>0).map(t=>d[t].n),
        backgroundColor: TIER_ORDER.filter(t=>d[t].n>0).map(t=>TIER_COLORS_HEX[t]),
        borderWidth: 0,
      }]
    },
    options: { plugins: { legend: { position:'right', labels:{ color:'#cbd5e1', font:{size:11} } } } }
  });

  // 2. Semáforo general (bar con 4 barras)
  new Chart(document.getElementById("chart-semaforo"), {
    type: 'bar',
    data: {
      labels: ['Verde', 'Amarillo', 'Rojo', 'Sin actividad'],
      datasets: [{
        data: [_st.semaforo.verde, _st.semaforo.amarillo, _st.semaforo.rojo, _st.semaforo.sin_actividad],
        backgroundColor: ['#4ade80', '#facc15', '#f87171', '#64748b'],
        borderRadius: 4,
      }]
    },
    options: {
      plugins:{ legend:{display:false} },
      scales: {
        y: { ticks: {color:'#94a3b8'}, grid:{color:'rgba(255,255,255,0.05)'} },
        x: { ticks: {color:'#94a3b8'}, grid:{display:false} }
      }
    }
  });

  // 3. Evolución pedidos por mes (line + área)
  const months = DATA.meta.ventana;
  new Chart(document.getElementById("chart-evolucion"), {
    type: 'line',
    data: {
      labels: months.map(mesShort),
      datasets: [{
        label: 'Pedidos VIP',
        data: months.map(m => _st.pedidos_por_mes[m] || 0),
        borderColor: '#818cf8',
        backgroundColor: 'rgba(129,140,248,0.15)',
        borderWidth: 2,
        tension: 0.35,
        fill: true,
        pointRadius: 4,
        pointBackgroundColor: '#818cf8',
      }]
    },
    options: {
      plugins: { legend: { labels: { color:'#cbd5e1' } } },
      scales: {
        y: { ticks: {color:'#94a3b8'}, grid:{color:'rgba(255,255,255,0.05)'} },
        x: { ticks: {color:'#94a3b8'}, grid:{display:false} }
      }
    }
  });

  // 4. Actividad en ventana (bar — usuarios con ventas por mes)
  new Chart(document.getElementById("chart-actividad"), {
    type: 'bar',
    data: {
      labels: months.map(mesShort),
      datasets: [{
        data: months.map(m => _st.activos_por_mes[m] || 0),
        backgroundColor: '#a5b4fc',
        borderRadius: 4,
      }]
    },
    options: {
      plugins:{ legend:{display:false} },
      scales: {
        y: { ticks: {color:'#94a3b8'}, grid:{color:'rgba(255,255,255,0.05)'} },
        x: { ticks: {color:'#94a3b8'}, grid:{display:false} }
      }
    }
  });
}

render();
</script>

<!-- Twemoji: renderiza emojis (banderas) como imágenes para que se vean
     idéntico en Mac/Windows/Linux/móvil. Sin esto, Windows muestra los
     emojis de banderas como letras "Regional Indicator" (ej. "🇨🇴" → "CO"). -->
<script src="https://cdn.jsdelivr.net/npm/@twemoji/api@15.1.0/dist/twemoji.min.js" crossorigin="anonymous"></script>
<style>img.twemoji{height:1em;width:auto;vertical-align:-0.125em;display:inline-block;margin:0 1px}</style>
<script>
(function(){
  if (!window.twemoji) return;
  const opts = {className: 'twemoji', folder: 'svg', ext: '.svg'};
  const parse = node => { try { twemoji.parse(node, opts); } catch(e){} };
  parse(document.body);
  // Re-parsear contenido dinámico (cambios de tab, filtros, búsqueda)
  new MutationObserver(muts => {
    for (const m of muts) for (const n of m.addedNodes)
      if (n.nodeType === 1) parse(n);
  }).observe(document.body, {childList: true, subtree: true});
})();
</script>
</body>
</html>
"""


def export_escalafon_json(data, out_path):
    """API de consulta para la landing: dict { sha256(email): {nivel, nombre} }.
    Los correos NO se publican en plano (la landing hashea el input igual y busca).
    Si el correo no está en la base, la landing recibe `null` (key inexistente).
    Orden de prioridad: VIP > Estudiantes > Dropi-sin-GHL (gana el primer match)."""
    escalafon = {}
    def add(email, nivel, nombre):
        em = (email or "").strip().lower()
        if not em or "@" not in em or not nivel:
            return
        h = hashlib.sha256(em.encode("utf-8")).hexdigest()
        if h not in escalafon:
            escalafon[h] = {"nivel": nivel, "nombre": (nombre or "").strip().title()}
    for u in data.get("usuarios", []):
        add(u.get("email"), u.get("nivel"), u.get("nombre"))
    for u in data.get("metricas", {}).get("estudiantes", []):
        add(u.get("email"), u.get("nivel"), u.get("nombre"))
    for u in data.get("metricas", {}).get("dropi_sin_ghl", []):
        add(u.get("email"), u.get("nivel"), u.get("nombre"))
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(escalafon, f, ensure_ascii=False, separators=(",", ":"))
    return len(escalafon)


def main():
    print("Generando dashboard.html...")
    data = compute_all()
    html_str = render_html(data).replace("__DATA_JSON__", json.dumps(data, ensure_ascii=False, default=str))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html_str)
    print(f"✓ {OUT}")
    n = export_escalafon_json(data, os.path.join(HERE, "escalafon.json"))
    print(f"✓ {os.path.join(HERE, 'escalafon.json')} · {n} correos consultables")
    print(f"  Usuarios totales: {data['stats']['usuarios_totales']}")
    print(f"  Clasificados:     {data['stats']['clasificados_vip']}")
    print(f"  Multi-país:       {data['stats']['multi_pais']}")
    print(f"  Activos 2 meses:  {data['stats']['activos_2_meses']}")
    print(f"  Desaparecidos:    {data['stats']['desaparecidos']}")
    print(f"  Recuperados:      {data['stats']['recuperados']}")


if __name__ == "__main__":
    main()
