#!/usr/bin/env python3
"""
Re-clasifica los contactos GHL aplicando la regla completa:
- 0-1 meses con ventas → Sin clasificar
- 2 meses con ventas → Bronce (cap, sin importar sumatoria)
- 3+ meses con ventas → tier según sumatoria de los 2 mejores meses
- Si sumatoria < 30 → Sin clasificar (no llega ni a Bronce)

Usa el dump de contactos ya descargados (ghl_contacts_raw.json) + maestro_emails.xlsx.
NO re-llama a la API.
"""
import os, json
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(__file__)
RAW = os.path.join(ROOT, "ghl_contacts_raw.json")
MAESTRO = os.path.join(ROOT, "maestro_emails.xlsx")
OUT = os.path.join(ROOT, "clasificacion_usuarios.xlsx")

TIENDA_EMAIL_IDS = {
    "Tienda 1":  "CQk3UpeEwUnbegiqR2Q3",
    "Tienda 2":  "P3jZOcralEFKIg4XpYho",
    "Tienda 3":  "p7TjCy0lVm6fP9xEbS3l",
    "Tienda 4":  "CZSUn21ycO4tr4LkNrbj",
    "Tienda 5":  "Mir5XAqxPoCrfT3fkgRF",
    "Tienda 6":  "riVtpCpiQPJvASPdEkKd",
    "Tienda 7":  "ThOpku1erpbHGCCBei6Z",
    "Tienda 8":  "83RmxTxcA8gkkWUsADls",
    "Tienda 9":  "2bl6kY6oQEIbPRRKpmaq",
    "Tienda 10": "75tlJ2SQeSdymOTxoScY",
}
TIENDA_PAIS_IDS = {
    "Tienda 1":  "0HWT1wbaaadgxxBPODUH",
    "Tienda 2":  "yJyX6eZUzkgnpBmAslde",
    "Tienda 3":  "IsI0hZEBHczVZ0itmPmV",
    "Tienda 4":  "CqZvpz0gtfu4bvCZwNqA",
    "Tienda 5":  "yWAvExtJrOJXTdjDnQuj",
    "Tienda 6":  "u7iiKtYTqJSKiFpioMdv",
    "Tienda 7":  "28jQePKJQGIZ198U0R6Z",
    "Tienda 8":  "gCaIQVieS9PqEU5AI8Uh",
    "Tienda 9":  "pEkrMm5ahV6PPow8PYQW",
    "Tienda 10": "cnShQcBSUMbU1WAFVqQx",
}
def _discover_months():
    """Lee maestro_emails.xlsx y devuelve los últimos 5 meses cronológicos con data."""
    wb = openpyxl.load_workbook(MAESTRO, read_only=True, data_only=True)
    ws = wb["MAESTRO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]; idx = hdr.index("mes")
    months = sorted({r[idx] for r in rows[1:] if r[idx]})
    return months[-5:]

MONTHS = _discover_months()
print(f"Ventana de evaluación: {MONTHS}")

def classify(active_months, suma_top2, suma_top3):
    """
    Regla completa (cada nivel usa la métrica que corresponde a sus meses):
      • INGRESO: 2 meses con ventas + top-2 ≥ 60
      • BRONCE:  top-2 ≥ 60     (2 meses bastan)
      • PLATA+:  3 meses con ventas, threshold sobre top-3
          Plata     →  top-3 ≥    300   (100/mes × 3)
          Oro       →  top-3 ≥    900   (300/mes × 3)
          Platino   →  top-3 ≥  3.000   (1.000/mes × 3)
          Diamante  →  top-3 ≥ 15.000   (5.000/mes × 3)
    """
    if active_months < 2 or suma_top2 < 60:
        return "Sin clasificar"
    if active_months < 3:
        return "Bronce"  # cap: necesitas 3 meses con ventas para escalar
    # active_months >= 3 → puede escalar
    if suma_top3 >= 15000: return "Diamante"
    if suma_top3 >=  3000: return "Platino"
    if suma_top3 >=   900: return "Oro"
    if suma_top3 >=   300: return "Plata"
    return "Bronce"

def is_at_risk(monthly_values_in_order):
    """Riesgo de eliminación = los ÚLTIMOS 3 meses están en cero Y había actividad
    antes. Excluye:
      - usuarios sin actividad histórica  ([0,0,0,0])
      - usuarios recién ingresados        ([0,0,0,X])
      - usuarios que se recuperaron        ([X,0,0,X])
    """
    if len(monthly_values_in_order) < 3:
        return False
    last_3 = monthly_values_in_order[-3:]
    earlier = monthly_values_in_order[:-3]
    return all(v == 0 for v in last_3) and any(v > 0 for v in earlier)

def extract_tiendas(contact):
    cf = {f["id"]: f.get("value") for f in contact.get("customFields", [])}
    out = []
    for label, fid in TIENDA_EMAIL_IDS.items():
        em = cf.get(fid)
        if em and isinstance(em, str) and "@" in em:
            out.append({
                "label": label,
                "email": em.strip().lower(),
                "pais": cf.get(TIENDA_PAIS_IDS[label]) or "",
            })
    return out

def load_maestro():
    wb = openpyxl.load_workbook(MAESTRO, read_only=True, data_only=True)
    ws = wb["MAESTRO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]; idx = {n:i for i,n in enumerate(hdr)}
    db = defaultdict(list)
    for r in rows[1:]:
        em = r[idx["email"]]
        if not em: continue
        db[em].append({
            "pais": r[idx["pais"]],
            "mes":  r[idx["mes"]],
            "pedidos": r[idx["pedidos"]] or 0,
        })
    return db

def main():
    print("Cargando dump GHL + maestro...")
    with open(RAW) as fp:
        contacts = json.load(fp)
    maestro = load_maestro()
    print(f"  {len(contacts)} contactos · {len(maestro)} correos en maestro\n")

    rows_out = []
    huerfanas = []
    duplicados = []  # contactos con el mismo correo repetido en varias tiendas
    sin_tienda = 0

    for c in contacts:
        tiendas_raw = extract_tiendas(c)
        if not tiendas_raw:
            sin_tienda += 1
            continue
        cid = c.get("id")
        primary = (c.get("email") or "").strip().lower()
        nombre = c.get("contactName") or f"{c.get('firstName','')} {c.get('lastName','')}".strip()
        tel = c.get("phone") or ""

        # SUMA: dedup por correo único (maestro[email] ya devuelve filas de
        # todos los países, sumarlo varias veces = doble conteo).
        # FLAG de duplicado: solo si (correo, país) coincide. Mismo correo en
        # diferentes países es legal (un usuario con Dropi en varios países).
        seen_emails = {}
        seen_pairs = {}  # (email, pais_normalizado) -> primer slot que lo usó
        for t in tiendas_raw:
            pais_norm = (t["pais"] or "").strip().lower()
            pair_key = (t["email"], pais_norm)
            if pair_key in seen_pairs and pais_norm:
                # Mismo correo Y mismo país → error de captura
                first = seen_pairs[pair_key]
                duplicados.append({
                    "contact_id":cid, "contact_email":primary, "contact_nombre":nombre,
                    "correo_duplicado":t["email"],
                    "pais":t["pais"],
                    "slot_original":first["label"],
                    "slot_duplicado":t["label"],
                })
            else:
                seen_pairs[pair_key] = t
            # Para la suma: un solo lookup por correo único
            if t["email"] not in seen_emails:
                seen_emails[t["email"]] = t
        tiendas = list(seen_emails.values())  # 1 entrada por correo único (todos los países)

        ped_mes = {m:0 for m in MONTHS}
        for t in tiendas:
            if t["email"] not in maestro:
                huerfanas.append({"contact_id":cid,"contact_email":primary,"contact_nombre":nombre,
                                  "tienda_label":t["label"],"tienda_email":t["email"],"tienda_pais_ghl":t["pais"]})
                continue
            for mr in maestro[t["email"]]:
                if mr["mes"] in ped_mes:
                    ped_mes[mr["mes"]] += mr["pedidos"]

        active = sum(1 for v in ped_mes.values() if v > 0)
        sorted_v = sorted(ped_mes.values(), reverse=True)
        top1 = sorted_v[0]
        top2 = sorted_v[1] if len(sorted_v) > 1 else 0
        top3 = sorted_v[2] if len(sorted_v) > 2 else 0
        suma_top2 = top1 + top2
        suma_top3 = top1 + top2 + top3
        nivel = classify(active, suma_top2, suma_top3)
        ordered_vals = [ped_mes[m] for m in MONTHS]
        eliminado = is_at_risk(ordered_vals)
        row = {
            "contact_id":cid, "primary_email":primary, "nombre":nombre, "telefono":tel,
            "n_tiendas":len(tiendas),
            "tiendas_emails":"; ".join(t["email"] for t in tiendas),
            "tiendas_paises":"; ".join(t["pais"] or "—" for t in tiendas),
            "meses_con_ventas":active,
            "total_ventana":sum(ped_mes.values()),
            "top1":top1, "top2":top2, "top3":top3,
            "suma_top2":suma_top2, "suma_top3":suma_top3,
            "nivel":nivel,
            "riesgo_eliminacion":"SÍ" if eliminado else "",
        }
        # Columnas dinámicas por mes: ped_<yyyy_mm>
        for m in MONTHS:
            row[f"ped_{m.replace('-','_')}"] = ped_mes[m]
        rows_out.append(row)

    rows_out.sort(key=lambda r: (r["nivel"]!="Diamante", r["nivel"]!="Platino", r["nivel"]!="Oro",
                                  r["nivel"]!="Plata", r["nivel"]!="Bronce", -r["suma_top3"]))

    wb_out = openpyxl.Workbook()
    ws = wb_out.active; ws.title = "USUARIOS_CLASIFICADOS"
    month_cols = [f"ped_{m.replace('-','_')}" for m in MONTHS]
    headers = ["contact_id","primary_email","nombre","telefono","n_tiendas",
               "tiendas_emails","tiendas_paises",
               *month_cols,
               "meses_con_ventas","total_ventana","top1","top2","top3","suma_top2","suma_top3","nivel","riesgo_eliminacion"]
    ws.append(headers)
    for r in rows_out:
        ws.append([r[h] for h in headers])

    ws2 = wb_out.create_sheet("TIENDAS_NO_ENCONTRADAS")
    ws2.append(["contact_id","contact_email","contact_nombre","tienda_label","tienda_email","tienda_pais_ghl"])
    for o in huerfanas:
        ws2.append([o["contact_id"],o["contact_email"],o["contact_nombre"],o["tienda_label"],o["tienda_email"],o["tienda_pais_ghl"]])

    ws3 = wb_out.create_sheet("RESUMEN_POR_NIVEL")
    ws3.append(["nivel","cantidad_usuarios","total_pedidos_acumulados"])
    by_tier = defaultdict(lambda: {"n":0,"total":0})
    for r in rows_out:
        by_tier[r["nivel"]]["n"] += 1
        by_tier[r["nivel"]]["total"] += r["total_ventana"]
    for n in ["Diamante","Platino","Oro","Plata","Bronce","Sin clasificar"]:
        b = by_tier.get(n, {"n":0,"total":0})
        ws3.append([n, b["n"], b["total"]])

    # Hoja: contactos con (correo, mismo país) repetido en >1 slot de Tienda.
    # Excluye casos legales (mismo correo en diferentes países).
    ws_dup = wb_out.create_sheet("CORREOS_DUPLICADOS_GHL")
    ws_dup.append(["contact_id","contact_email","contact_nombre",
                   "correo_duplicado","pais","slot_original","slot_duplicado"])
    for d in duplicados:
        ws_dup.append([d["contact_id"],d["contact_email"],d["contact_nombre"],
                       d["correo_duplicado"],d["pais"],d["slot_original"],d["slot_duplicado"]])

    # Hoja extra: capeados en Bronce por solo 2 meses con ventas
    ws_cap = wb_out.create_sheet("CAPEADOS_2_MESES")
    ws_cap.append(["contact_id","nombre","primary_email","meses_con_ventas",
                   "suma_top3","nivel_que_hubiera_tenido_con_3_meses"])
    def hypothetical_with_3m(s):
        if s>=5000: return "Diamante"
        if s>=1000: return "Platino"
        if s>=300:  return "Oro"
        if s>=100:  return "Plata"
        return "Bronce"
    for r in rows_out:
        if r["nivel"]=="Bronce" and r["meses_con_ventas"]<3 and r["suma_top3"]>=100:
            ws_cap.append([r["contact_id"],r["nombre"],r["primary_email"],
                           r["meses_con_ventas"],r["suma_top3"],
                           hypothetical_with_3m(r["suma_top3"])])

    # Hoja extra: usuarios en RIESGO de eliminación (3 meses consecutivos sin pedidos)
    ws4 = wb_out.create_sheet("RIESGO_ELIMINACION")
    ws4.append(["contact_id","nombre","primary_email","nivel_actual", *month_cols])
    for r in rows_out:
        if r["riesgo_eliminacion"]:
            ws4.append([r["contact_id"],r["nombre"],r["primary_email"],r["nivel"],
                        *[r[c] for c in month_cols]])

    wb_out.save(OUT)
    print(f"✅ Escrito: {OUT}")
    print(f"   Contactos analizados: {len(rows_out)}")
    print(f"   Sin tienda registrada (excluidos): {sin_tienda}")
    print(f"   Duplicaciones de correo en slots Tienda: {len(duplicados)}\n")
    print("Distribución por nivel:")
    for n in ["Diamante","Platino","Oro","Plata","Bronce","Sin clasificar"]:
        b = by_tier.get(n, {"n":0,"total":0})
        print(f"   {n:18}  {b['n']:>5}  ({b['total']:>10,} pedidos)")

if __name__ == "__main__":
    main()
