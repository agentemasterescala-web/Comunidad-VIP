#!/usr/bin/env python3
"""
Escribe la clasificación VIP a GHL: tags + custom fields + historial.

Uso:
  GHL_TOKEN=... GHL_LOCATION=... python3 escribir_a_ghl.py            # TODOS los contactos
  GHL_TOKEN=... GHL_LOCATION=... python3 escribir_a_ghl.py --dry-run  # simula sin escribir
  GHL_TOKEN=... GHL_LOCATION=... python3 escribir_a_ghl.py --contact-id ABC --dry-run
  GHL_TOKEN=... GHL_LOCATION=... python3 escribir_a_ghl.py --contact-id ABC

Lee:
  - ghl_contacts_raw.json    (dump de contactos GHL ya descargados)
  - maestro_emails.xlsx      (consolidado de pedidos por email/país/mes)

Lógica de clasificación:
  - Ingreso: top-2 ≥ 30
  - Cap Bronce: <3 meses con ventas
  - Tiers (suma top-3): Bronce 30-99, Plata 100-299, Oro 300-999,
                       Platino 1.000-4.999, Diamante 15.000+ (Opción A)
"""
import os, json, sys, time, argparse, urllib.request, urllib.error
from collections import defaultdict
import openpyxl

# Carga .env del mismo directorio si existe
_HERE = os.path.dirname(os.path.abspath(__file__))
_ENV = os.path.join(_HERE, ".env")
if os.path.isfile(_ENV):
    for line in open(_ENV):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line: continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

TOK = os.environ.get("GHL_TOKEN")
LOC = os.environ.get("GHL_LOCATION")
if not TOK or not LOC:
    sys.exit("Faltan GHL_TOKEN y/o GHL_LOCATION (env o .env)")

ROOT = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(ROOT, "ghl_contacts_raw.json")
MAESTRO = os.path.join(ROOT, "maestro_emails.xlsx")

# IDs de los 10 slots de Tienda (correo + país)
TIENDA_EMAIL_IDS = {
    "Tienda 1":"CQk3UpeEwUnbegiqR2Q3","Tienda 2":"P3jZOcralEFKIg4XpYho",
    "Tienda 3":"p7TjCy0lVm6fP9xEbS3l","Tienda 4":"CZSUn21ycO4tr4LkNrbj",
    "Tienda 5":"Mir5XAqxPoCrfT3fkgRF","Tienda 6":"riVtpCpiQPJvASPdEkKd",
    "Tienda 7":"ThOpku1erpbHGCCBei6Z","Tienda 8":"83RmxTxcA8gkkWUsADls",
    "Tienda 9":"2bl6kY6oQEIbPRRKpmaq","Tienda 10":"75tlJ2SQeSdymOTxoScY",
}
TIENDA_PAIS_IDS = {
    "Tienda 1":"0HWT1wbaaadgxxBPODUH","Tienda 2":"yJyX6eZUzkgnpBmAslde",
    "Tienda 3":"IsI0hZEBHczVZ0itmPmV","Tienda 4":"CqZvpz0gtfu4bvCZwNqA",
    "Tienda 5":"yWAvExtJrOJXTdjDnQuj","Tienda 6":"u7iiKtYTqJSKiFpioMdv",
    "Tienda 7":"28jQePKJQGIZ198U0R6Z","Tienda 8":"gCaIQVieS9PqEU5AI8Uh",
    "Tienda 9":"pEkrMm5ahV6PPow8PYQW","Tienda 10":"cnShQcBSUMbU1WAFVqQx",
}

# IDs de campos del bloque "Comunidad VIP"
F = {
    "escalafon_vip":     "evyetA9K7plkYMDd3tCQ",  # SINGLE_OPTIONS
    "pedidos_vip":       "YAVJHSdLoFnTKbUxUtLK",  # NUMERICAL
    "mes_escalafon":     "tXNrCxLvidhkNyK85T4T",  # TEXT
    "cantidad_ult_mes":  "XIoj5twBfJzJ6irOxraV",  # NUMERICAL
    "ventas_ult_1":      "bgQhOLdDMJUmcxUgXv89",  # TEXT
    "ventas_ult_2":      "OUH451COVuZeMl6BD3lo",  # TEXT
    "ventas_ult_3":      "ogVSepUDzQxqzv6U3ACw",  # TEXT
    "historial":         "SbrJjfBouQa52aSuH64P",  # LARGE_TEXT
}

def discover_months_from_maestro(path=None):
    """Devuelve los últimos 5 meses (orden cronológico) que aparecen en maestro_emails.xlsx."""
    p = path or MAESTRO
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["MAESTRO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]; idx = hdr.index("mes")
    months = sorted({r[idx] for r in rows[1:] if r[idx]})
    return months[-5:]

MONTHS = discover_months_from_maestro()   # últimos 5 meses con data
LATEST = MONTHS[-1]                       # mes de corte
print(f"Ventana de evaluación: {MONTHS}  (corte = {LATEST})")

ES_MONTH_SHORT = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def month_short(yyyy_mm):
    """'2026-03' → 'Mar'26'  (formato del historial existente en GHL)"""
    y, m = yyyy_mm.split("-")
    return f"{ES_MONTH_SHORT[int(m)-1]}'{y[-2:]}"

def historial_label(nivel):
    """'Sin clasificar' → 'Sin nivel' (lo demás se queda igual)"""
    return "Sin nivel" if nivel == "Sin clasificar" else nivel

def parse_historial(s):
    """Lee 'Mar'26:Platino | Feb'26:Bronce | ...' a [(mes, nivel), ...].
    Solo lee la primera línea para tolerar entradas legacy en líneas siguientes."""
    if not s: return []
    first = s.split("\n")[0].strip()
    out = []
    for p in [x.strip() for x in first.split("|") if x.strip()]:
        if ":" in p:
            mes, niv = p.split(":", 1)
            out.append((mes.strip(), niv.strip()))
    return out

def update_historial_str(existing, yyyy_mm, nivel):
    """Devuelve historial actualizado, dedupe por mes (queda solo primera ocurrencia).
    El nuevo entry va siempre al frente."""
    new_mes = month_short(yyyy_mm)
    new_niv = historial_label(nivel)
    entries = parse_historial(existing)
    # Dedupe: keep first occurrence of each month
    seen = set()
    deduped = []
    for m, n in entries:
        if m not in seen:
            seen.add(m); deduped.append((m, n))
    # Remove existing entry for new_mes (será reemplazado por el nuevo al frente)
    deduped = [(m, n) for (m, n) in deduped if m != new_mes]
    entries = [(new_mes, new_niv)] + deduped
    return " | ".join(f"{m}:{n}" for m, n in entries)

TIER_TAG = {
    "Bronce":   "vip bronce",
    "Plata":    "vip plata",
    "Oro":      "vip oro",
    "Platino":  "vip platino",
    "Diamante": "vip diamante",
    "Sin clasificar": "vip sin clasificar",
}
TIER_FIELD_VALUE = {
    "Bronce":   "Bronce",
    "Plata":    "Plata",
    "Oro":      "Oro",
    "Platino":  "Platino",
    "Diamante": "Diamante",
    "Sin clasificar": "Sin Nivel",  # opción del dropdown en GHL
}
ALL_VIP_TAGS = set(TIER_TAG.values())


def classify(active_months, suma_top2, suma_top3):
    """Bronce usa top-2 (2 meses), Plata-Diamante usan top-3 (3 meses)."""
    if active_months < 2 or suma_top2 < 60:
        return "Sin clasificar"
    if active_months < 3:
        return "Bronce"
    if suma_top3 >= 15000: return "Diamante"
    if suma_top3 >=  3000: return "Platino"
    if suma_top3 >=   900: return "Oro"
    if suma_top3 >=   300: return "Plata"
    return "Bronce"


def http(method, url, body=None, retries=4):
    data = json.dumps(body).encode() if body is not None else None
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, data=data, method=method, headers={
                "Authorization": f"Bearer {TOK}",
                "Version": "2021-07-28",
                "Accept": "application/json",
                "Content-Type": "application/json",
                "User-Agent": "ProyectoClaude/1.0",
            })
            with urllib.request.urlopen(req, timeout=60) as r:
                txt = r.read()
                return json.loads(txt) if txt else {}
        except urllib.error.HTTPError as e:
            err_body = ""
            try: err_body = e.read().decode()
            except: pass
            last = RuntimeError(f"HTTP {e.code} {method} {url}: {err_body}")
            if e.code in (429, 502, 503, 504):
                time.sleep(2 ** attempt); continue
            raise last
        except Exception as e:
            last = e; time.sleep(2 ** attempt)
    raise last


def update_custom_fields(cid, fields_array):
    return http("PUT", f"https://services.leadconnectorhq.com/contacts/{cid}",
                {"customFields": fields_array})

def add_tags(cid, tags):
    return http("POST", f"https://services.leadconnectorhq.com/contacts/{cid}/tags",
                {"tags": list(tags)})

def remove_tags(cid, tags):
    return http("DELETE", f"https://services.leadconnectorhq.com/contacts/{cid}/tags",
                {"tags": list(tags)})


def extract_tiendas(contact):
    cf = {f["id"]: f.get("value") for f in contact.get("customFields", [])}
    seen = {}
    for label, fid in TIENDA_EMAIL_IDS.items():
        em = cf.get(fid)
        if em and isinstance(em, str) and "@" in em:
            em_low = em.strip().lower()
            if em_low not in seen:
                seen[em_low] = {"label": label, "email": em_low,
                                "pais": cf.get(TIENDA_PAIS_IDS[label]) or ""}
    return list(seen.values())


def load_maestro():
    wb = openpyxl.load_workbook(MAESTRO, read_only=True, data_only=True)
    ws = wb["MAESTRO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]; idx = {n:i for i,n in enumerate(hdr)}
    db = defaultdict(list)
    for r in rows[1:]:
        em = r[idx["email"]]
        if not em: continue
        db[em].append({"pais": r[idx["pais"]], "mes": r[idx["mes"]],
                       "pedidos": r[idx["pedidos"]] or 0})
    return db


def calc_user(contact, maestro):
    tiendas = extract_tiendas(contact)
    ped_mes = {m: 0 for m in MONTHS}
    for t in tiendas:
        for mr in maestro.get(t["email"], []):
            if mr["mes"] in ped_mes:
                ped_mes[mr["mes"]] += mr["pedidos"]
    active = sum(1 for v in ped_mes.values() if v > 0)
    sv = sorted(ped_mes.values(), reverse=True)
    top1 = sv[0]; top2 = sv[1] if len(sv) > 1 else 0; top3 = sv[2] if len(sv) > 2 else 0
    nivel = classify(active, top1+top2, top1+top2+top3)
    return {
        "nivel": nivel,
        "ped_mes": ped_mes,
        "active": active,
        "suma_top3": top1+top2+top3,
        "n_tiendas": len(tiendas),
    }


def build_updates(contact, calc):
    """Devuelve dict con plan de cambios. No ejecuta. Marca is_noop=True
    cuando todos los valores nuevos coinciden con los actuales en GHL."""
    cid = contact["id"]
    cf_now = {f["id"]: f.get("value") for f in contact.get("customFields", [])}

    current_escalafon = cf_now.get(F["escalafon_vip"]) or ""
    current_pedidos   = cf_now.get(F["pedidos_vip"])
    current_mes_esc   = cf_now.get(F["mes_escalafon"]) or ""
    current_hist      = cf_now.get(F["historial"]) or ""
    current_cant      = cf_now.get(F["cantidad_ult_mes"])
    current_v1        = cf_now.get(F["ventas_ult_1"])
    current_v2        = cf_now.get(F["ventas_ult_2"])
    current_v3        = cf_now.get(F["ventas_ult_3"])
    current_tags      = list(contact.get("tags") or [])

    new_escalafon_value = TIER_FIELD_VALUE[calc["nivel"]]
    tier_changed = current_escalafon != new_escalafon_value
    mes_esc_new = LATEST if (tier_changed or not current_mes_esc) else current_mes_esc

    new_hist = update_historial_str(current_hist, LATEST, calc["nivel"])
    entry = f"{month_short(LATEST)}:{historial_label(calc['nivel'])}"

    pm = calc["ped_mes"]
    def _back(n):
        return pm[MONTHS[-1-n]] if len(MONTHS) > n else 0
    new_cant = _back(0); new_v1 = _back(1); new_v2 = _back(2); new_v3 = _back(3)

    custom_fields_payload = [
        {"id": F["escalafon_vip"],    "field_value": new_escalafon_value},
        {"id": F["pedidos_vip"],      "field_value": calc["suma_top3"]},
        {"id": F["mes_escalafon"],    "field_value": mes_esc_new},
        {"id": F["cantidad_ult_mes"], "field_value": new_cant},
        {"id": F["ventas_ult_1"],     "field_value": str(new_v1)},
        {"id": F["ventas_ult_2"],     "field_value": str(new_v2)},
        {"id": F["ventas_ult_3"],     "field_value": str(new_v3)},
        {"id": F["historial"],        "field_value": new_hist},
    ]

    new_tag = TIER_TAG[calc["nivel"]]
    tags_to_remove = [t for t in current_tags if t in ALL_VIP_TAGS and t != new_tag]
    tags_to_add = [new_tag] if new_tag not in current_tags else []

    # No-op detection: si TODOS los campos coinciden y no hay cambios de tags,
    # podemos saltarnos las llamadas a la API.
    def _same_num(a, b):
        try: return float(a) == float(b)
        except (TypeError, ValueError): return str(a or "") == str(b or "")
    fields_match = (
        current_escalafon == new_escalafon_value
        and _same_num(current_pedidos, calc["suma_top3"])
        and (current_mes_esc or "") == mes_esc_new
        and (current_hist or "") == new_hist
        and _same_num(current_cant, new_cant)
        and str(current_v1 or "0") == str(new_v1)
        and str(current_v2 or "0") == str(new_v2)
        and str(current_v3 or "0") == str(new_v3)
    )
    tags_match = (not tags_to_add) and (not tags_to_remove)
    is_noop = fields_match and tags_match

    return {
        "contact_id": cid,
        "current_escalafon": current_escalafon,
        "new_escalafon": new_escalafon_value,
        "tier_changed": tier_changed,
        "mes_esc_old": current_mes_esc,
        "mes_esc_new": mes_esc_new,
        "historial_entry": entry,
        "custom_fields": custom_fields_payload,
        "tags_to_add": tags_to_add,
        "tags_to_remove": tags_to_remove,
        "ped_mes": pm,
        "is_noop": is_noop,
    }


def apply_updates(plan):
    cid = plan["contact_id"]
    update_custom_fields(cid, plan["custom_fields"])
    if plan["tags_to_remove"]:
        remove_tags(cid, plan["tags_to_remove"])
    if plan["tags_to_add"]:
        add_tags(cid, plan["tags_to_add"])


def fmt_plan(contact, calc, plan):
    nombre = contact.get("contactName") or "(sin nombre)"
    cid = contact["id"]
    pm = plan["ped_mes"]
    arrow = "🔄" if plan["tier_changed"] else "  "
    print(f"\n──────────────────────────────────────────────────────────────")
    print(f"  {arrow}  {nombre}  ({cid})")
    print(f"      Nivel: {plan['current_escalafon'] or '(vacío)':12} → {plan['new_escalafon']}")
    print(f"      Top-3: {calc['suma_top3']:,}  ·  meses con ventas: {calc['active']}/{len(MONTHS)}  ·  tiendas: {calc['n_tiendas']}")
    ped_str = "  ".join(f"{m[5:]}/{m[2:4]}={pm[m]:,}" for m in MONTHS)
    print(f"      Pedidos: {ped_str}")
    print(f"      Mes Escalafón: {plan['mes_esc_old'] or '(vacío)'} → {plan['mes_esc_new']}")
    print(f"      Tags ➖ remove: {plan['tags_to_remove'] or '—'}")
    print(f"      Tags ➕ add:    {plan['tags_to_add'] or '—'}")
    print(f"      Historial entry: {plan['historial_entry']}")
    print(f"      Historial completo: {plan['custom_fields'][7]['field_value']}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--contact-id", help="Procesar solo este contacto")
    ap.add_argument("--limit", type=int, default=0, help="Procesar solo N contactos (debug)")
    ap.add_argument("--verbose", action="store_true", help="Mostrar también los no-op")
    args = ap.parse_args()

    print("Cargando contactos + maestro...")
    with open(RAW) as fp:
        contacts = json.load(fp)
    maestro = load_maestro()

    if args.contact_id:
        contacts = [c for c in contacts if c.get("id") == args.contact_id]
        if not contacts:
            sys.exit(f"No se encontró el contact_id {args.contact_id}")

    print(f"  {len(contacts)} contactos a procesar  ·  modo: {'DRY-RUN' if args.dry_run else 'LIVE'}\n")

    processed = 0; updated = 0; noops = 0; errors = 0
    for c in contacts:
        if args.limit and processed >= args.limit:
            break
        tiendas = extract_tiendas(c)
        if not tiendas:
            continue
        calc = calc_user(c, maestro)
        plan = build_updates(c, calc)

        if plan["is_noop"]:
            noops += 1
            if args.verbose or args.contact_id:
                print(f"  ⊝  {(c.get('contactName') or c.get('id'))[:40]:42} no-op ({plan['new_escalafon']})")
        else:
            fmt_plan(c, calc, plan)
            if not args.dry_run:
                try:
                    apply_updates(plan)
                    updated += 1
                    time.sleep(0.2)
                except Exception as e:
                    errors += 1
                    print(f"      ❌ ERROR: {e}")
            else:
                updated += 1
        processed += 1

    print(f"\n────────────────────────────────────────")
    print(f"  Procesados: {processed}  ·  actualizados: {updated}  ·  no-op: {noops}  ·  errores: {errors}")
    print(f"  Modo: {'DRY-RUN' if args.dry_run else 'LIVE'}")


if __name__ == "__main__":
    main()
