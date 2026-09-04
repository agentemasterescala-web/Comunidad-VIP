#!/usr/bin/env python3
"""
Consolida los 19 archivos Excel (Dic 2025, Ene/Feb/Mar 2026) en una tabla
maestra: email -> {mes: pedidos, pais, nombre, telefono}.

Métrica de pedidos = ENTREGADOS + DEVOLUCIONES (decisión del usuario).
Salida: maestro_emails.xlsx con una fila por (email, país, mes, pedidos).

Resiliencia (para no perder meses en silencio):
  · Cada Excel se materializa desde Google Drive (lee los bytes) y se reintenta
    con backoff — un archivo "solo en la nube" ya no tira la corrida.
  · Guarda de regresión: si un archivo falla o un mes que el maestro anterior
    tenía desaparece o cae >50%, NO se sobrescribe el maestro (falla cerrado,
    sale con código 2) para conservar el último bueno.
"""
import os, re, glob, unicodedata, io, time
from collections import defaultdict, Counter
import openpyxl

# Paths portables — todo se resuelve relativo a la ubicación de este script.
# Estructura esperada (cualquier profundidad — busca recursivamente):
#   <parent>/originales/<Mes><Año>/*.xlsx
#   <parent>/originales/<Año>/<Mes><Año>/*.xlsx     (ej. layout Google Drive)
#   <parent>/procesados/<este script>
_HERE = os.path.dirname(os.path.abspath(__file__))
_PARENT = os.path.dirname(_HERE)
ROOT = os.path.join(_PARENT, "originales")
OUT_DIR = _HERE

ES_MONTHS = {
    "enero":"01","febrero":"02","marzo":"03","abril":"04","mayo":"05","junio":"06",
    "julio":"07","agosto":"08","septiembre":"09","octubre":"10","noviembre":"11","diciembre":"12"
}

def _strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def auto_discover_months():
    """Walk recursivo bajo ROOT buscando carpetas '<MesEspañol><Año>'
    (ej. 'Enero2026'). Devuelve {ruta_absoluta: 'YYYY-MM'}.
    Si el mismo mes aparece en varias rutas (p. ej. 2025/Diciembre2025 y
    2026/Diciembre2025 en Drive), solo se conserva la primera para evitar
    doble conteo en la agregación."""
    found = {}             # ruta -> 'YYYY-MM'
    seen_labels = set()    # 'YYYY-MM' ya tomado
    if not os.path.isdir(ROOT):
        return found
    pat = re.compile(r"^([A-Za-zÁÉÍÓÚáéíóúñÑ]+)(\d{4})$")
    for dirpath, dirnames, _ in os.walk(ROOT, followlinks=True):
        # ordenar para que el primer match sea determinista
        dirnames.sort()
        for d in dirnames:
            m = pat.match(d)
            if not m: continue
            mes_name = _strip_accents(m.group(1).lower())
            if mes_name not in ES_MONTHS: continue
            year = m.group(2)
            label = f"{year}-{ES_MONTHS[mes_name]}"
            if label in seen_labels:
                # evitar doble conteo de la misma etiqueta (mismo mes en
                # múltiples ubicaciones del árbol). Conservamos la primera.
                continue
            seen_labels.add(label)
            found[os.path.join(dirpath, d)] = label
    return found

# Mapping ruta_absoluta -> mes etiqueta (autodetectado, recursivo)
MONTH_MAP = auto_discover_months()

def norm(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower()

COUNTRY_NAMES = {"ARGENTINA","CHILE","COLOMBIA","ECUADOR","GUATEMALA",
                 "MEXICO","PANAMA","PARAGUAY","PERU","COSTARICA","COSTARRICA"}

def country_from_filename(fn):
    fn = os.path.basename(fn).upper()
    for c in ("ARGENTINA","CHILE","COLOMBIA","ECUADOR","GUATEMALA","MEXICO","PANAMA","PARAGUAY","PERU","COSTARICA","COSTA RICA"):
        if c in fn:
            return c.replace("COSTA RICA","COSTARICA")
    return "DESCONOCIDO"

def is_consolidated_workbook(wb, fn):
    """Decide si un workbook es 'multi-país en una sola hoja por país'.
    Verdadero si el nombre del archivo contiene 'paises' (convención original)
    o si el workbook tiene >=2 hojas con nombre de país (ej. mayo 2026, donde
    el archivo se llama solo 'IVAN CAICEDO MAYO.xlsx' sin 'paises')."""
    if "paises" in os.path.basename(fn).lower():
        return True
    country_sheets = 0
    for sn in wb.sheetnames:
        sn_up = sn.upper().replace(" ","")
        if any(c in sn_up for c in COUNTRY_NAMES):
            country_sheets += 1
    return country_sheets >= 2

def find_header_row(ws, max_scan=10):
    """Encuentra la fila que contiene 'email' (o 'USUARIOS'/'USUARIO') y métricas.
    Acepta 'entregados' (formato hasta Jun 2026) y 'entregadas' (formato Jul 2026+)."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_scan), start=1):
        norms = [norm(c) for c in row]
        if "email" in norms or "usuarios" in norms or "usuario" in norms:
            if any(n in ("entregados","entregadas") for n in norms) \
               or any("ord. ing." in n or "ord ing" in n for n in norms):
                return i, row
    return None, None

def col_index(header_row, target_norms):
    """Devuelve índice de la primera columna cuya cabecera normalizada esté en target_norms."""
    for idx, c in enumerate(header_row):
        if norm(c) in target_norms:
            return idx
    return None

def parse_sheet(ws, country, month_label):
    hdr_row_idx, hdr = find_header_row(ws)
    if hdr is None:
        return []

    # Elegir la columna del email del vendedor. Formatos que hemos visto:
    #   • Hasta Abr 2026:  solo "USUARIOS" con el correo.
    #   • May 2026:        "USUARIOS" con el correo + "Email" con el correo del
    #                      líder/gestor del grupo (varias filas comparten Email,
    #                      así que usar "Email" mete doble/multi-conteo).
    #   • Jun 2026:        solo "email" con el correo.
    #   • Jul 2026+:       "Usuario" con el NOMBRE + "Email" con el correo.
    # Regla: si existen ambas columnas, muestreamos las primeras filas y usamos
    # la que tenga MÁS emails únicos válidos (el correo del vendedor cambia
    # fila a fila; el email del líder se repite).
    col_email = None
    cand_usuarios = col_index(hdr, {"usuarios","usuario"})
    cand_email = col_index(hdr, {"email"})
    if cand_usuarios is None:
        col_email = cand_email
    elif cand_email is None:
        col_email = cand_usuarios
    else:
        # Ambas presentes → muestrear filas siguientes al header
        sample_rows = list(ws.iter_rows(min_row=hdr_row_idx+1, max_row=hdr_row_idx+60, values_only=True))
        def unique_valid_emails(col_ix):
            s = set()
            for row in sample_rows:
                if col_ix < len(row):
                    v = row[col_ix]
                    if v and isinstance(v, str) and "@" in v and "." in v.split("@")[-1]:
                        s.add(v.strip().lower())
            return len(s)
        n_u = unique_valid_emails(cand_usuarios)
        n_e = unique_valid_emails(cand_email)
        # Se prefiere la columna con más correos únicos (el "USUARIOS" en mayo
        # es único por fila; el "Email" del líder se repite decenas de veces).
        col_email = cand_usuarios if n_u >= n_e else cand_email
    col_nombre = col_index(hdr, {"nombre"})
    col_tel = col_index(hdr, {"telefono"})
    # Si nombre/telefono no estan en el header principal, pueden venir como sub-headers
    # en la siguiente fila (caso Feb). Detectar esto:
    next_row = None
    rows_iter = ws.iter_rows(values_only=True)
    rows_list = list(rows_iter)
    if hdr_row_idx < len(rows_list):
        next_row = rows_list[hdr_row_idx]  # fila siguiente al header (0-indexed)
    if (col_nombre is None or col_tel is None) and next_row is not None:
        if col_nombre is None:
            col_nombre = col_index(next_row, {"nombre"})
        if col_tel is None:
            col_tel = col_index(next_row, {"telefono"})

    # Métricas: ENTREGADOS y DEVOLUCIONES.
    # Formato hasta Jun 2026: columnas 'entregados' y 'devoluciones' (masc plural).
    # Formato Jul 2026+:      columnas 'entregadas' y 'movilizadas' — donde
    #                         movilizadas = entregadas + devoluciones (equivalente
    #                         a "pedidos" del sistema viejo). Devoluciones se
    #                         calcula como (movilizadas - entregadas).
    col_entregados = None
    col_devoluciones = None
    col_movilizadas = None
    for idx, c in enumerate(hdr):
        n = norm(c)
        if n in ("entregados","entregadas"):
            col_entregados = idx
        elif n in ("devoluciones","devueltas"):
            col_devoluciones = idx
        elif n == "movilizadas":
            col_movilizadas = idx

    if col_email is None or col_entregados is None:
        print(f"  ⚠ {country} {month_label}: header incompleto (falta email/entregados), skip")
        return []
    if col_devoluciones is None and col_movilizadas is None:
        print(f"  ⚠ {country} {month_label}: header sin devoluciones ni movilizadas, skip")
        return []

    out = []
    # Empezar a leer desde la fila siguiente al header (saltando posible sub-header / Total)
    start_idx = hdr_row_idx  # 0-indexed, empieza en la fila que sigue al header
    for r_idx in range(start_idx, len(rows_list)):
        row = rows_list[r_idx]
        if row is None: continue
        email = row[col_email] if col_email < len(row) else None
        if email is None: continue
        email = str(email).strip()
        if not email or "@" not in email:
            continue  # Salta 'Total' y filas vacías
        email_norm = email.lower()
        nombre = row[col_nombre] if col_nombre is not None and col_nombre < len(row) else None
        tel = row[col_tel] if col_tel is not None and col_tel < len(row) else None
        ent = row[col_entregados] if col_entregados < len(row) else 0
        try:
            ent = int(float(ent)) if ent not in (None, "") else 0
        except (TypeError, ValueError):
            ent = 0
        # Devoluciones: si existe columna, usarla directamente. Si no, derivar de
        # (movilizadas - entregadas) — equivale a "devoluciones" del formato viejo,
        # ya que movilizadas = entregadas + devoluciones en el formato Jul 2026+.
        if col_devoluciones is not None:
            dev = row[col_devoluciones] if col_devoluciones < len(row) else 0
            try:
                dev = int(float(dev)) if dev not in (None, "") else 0
            except (TypeError, ValueError):
                dev = 0
        else:
            mov = row[col_movilizadas] if col_movilizadas < len(row) else 0
            try:
                mov = int(float(mov)) if mov not in (None, "") else 0
            except (TypeError, ValueError):
                mov = 0
            dev = max(0, mov - ent)  # nunca negativo
        pedidos = ent + dev
        out.append({
            "email": email_norm,
            "nombre": str(nombre).strip() if nombre else "",
            "telefono": str(tel).strip() if tel else "",
            "pais": country,
            "mes": month_label,
            "entregados": ent,
            "devoluciones": dev,
            "pedidos": pedidos,
        })
    return out

def load_workbook_resiliente(fn, retries=4):
    """Abre un .xlsx tolerando archivos 'solo en la nube' de Google Drive.

    Fuerza la materializacion leyendo TODOS los bytes (dispara la descarga del
    File Provider de Drive) y carga desde memoria; reintenta con backoff si falla
    (descarga a medias / I/O transitorio). Sin esto, un archivo no materializado
    tira openpyxl y el mes se caia del maestro en silencio."""
    last = None
    for intento in range(retries):
        try:
            with open(fn, "rb") as f:
                data = f.read()
            return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(2 ** intento)
    raise last


def meses_conteo_maestro(path):
    """{mes: nº de filas} del maestro existente (hoja MAESTRO). {} si no existe
    o no se puede leer. Sirve de baseline para la guarda de regresion."""
    if not os.path.isfile(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["MAESTRO"] if "MAESTRO" in wb.sheetnames else wb.active
        it = ws.iter_rows(values_only=True)
        hdr = next(it, None)
        idx = None
        if hdr:
            for i, h in enumerate(hdr):
                if str(h).strip().lower() == "mes":
                    idx = i
                    break
        c = Counter()
        if idx is not None:
            for r in it:
                if idx < len(r) and r[idx]:
                    c[r[idx]] += 1
        wb.close()
        return dict(c)
    except Exception:  # noqa: BLE001
        return {}


def main():
    all_records = []
    failed_files = []   # (basename, mes, error) de Excels que no se pudieron abrir
    files_processed = 0
    for path, month_label in MONTH_MAP.items():
        files = sorted(glob.glob(os.path.join(path, "*.xlsx")))
        files = [f for f in files if not os.path.basename(f).startswith("~$")]
        for fn in files:
            files_processed += 1
            print(f"[{files_processed}] {os.path.basename(fn)}")
            try:
                wb = load_workbook_resiliente(fn)
            except Exception as e:
                print(f"  ✗ error tras reintentos: {e}")
                failed_files.append((os.path.basename(fn), month_label, str(e)))
                continue
            consolidated = is_consolidated_workbook(wb, fn)
            if consolidated:
                # Una hoja por país. Las versiones nuevas (Abril 2026+) traen 3
                # hojas por país: <PAIS> USUARIOS, PRODUCTOS <PAIS>, PROVEEDORES
                # <PAIS>. Solo procesamos las que contengan datos por usuario.
                for sn in wb.sheetnames:
                    sn_up = sn.upper().replace(" ","").replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U")
                    if any(s in sn_up for s in ("PRODUCTO","PROVEEDOR")):
                        continue  # skip non-user sheets
                    # Detectar país en el nombre de la hoja
                    country = None
                    for cn in COUNTRY_NAMES:
                        if cn in sn_up:
                            country = cn; break
                    if country is None:
                        # fallback: usar nombre limpio
                        country = sn_up.replace("USUARIOS","").replace("USUARIO","").strip()
                        if not country: continue
                    ws = wb[sn]
                    recs = parse_sheet(ws, country, month_label)
                    print(f"    └ {sn} → {country}: {len(recs)} usuarios")
                    all_records.extend(recs)
            else:
                country = country_from_filename(fn)
                # Hoja "Original" (con duplicados) o "Export" o la primera
                pref = ["Original","Export"]
                target = None
                for p in pref:
                    if p in wb.sheetnames:
                        target = p; break
                if target is None:
                    target = wb.sheetnames[0]
                ws = wb[target]
                recs = parse_sheet(ws, country, month_label)
                print(f"    └ {target}: {len(recs)} usuarios")
                all_records.extend(recs)
            wb.close()

    print(f"\nTotal registros (email, pais, mes): {len(all_records)}")
    # Consolidar por email+pais+mes (sumar por si hay duplicados)
    agg = defaultdict(lambda: {"pedidos":0,"entregados":0,"devoluciones":0,"nombre":"","telefono":""})
    for r in all_records:
        k = (r["email"], r["pais"], r["mes"])
        agg[k]["pedidos"] += r["pedidos"]
        agg[k]["entregados"] += r["entregados"]
        agg[k]["devoluciones"] += r["devoluciones"]
        if not agg[k]["nombre"] and r["nombre"]:
            agg[k]["nombre"] = r["nombre"]
        if not agg[k]["telefono"] and r["telefono"]:
            agg[k]["telefono"] = r["telefono"]

    # Guarda anti-borrado: si no se leyó ningún registro (p. ej. el proceso
    # de launchd no tiene acceso a Google Drive / Full Disk Access), NO
    # sobrescribir el maestro existente con datos vacíos. Aborta con código
    # de error para que el pipeline se detenga y conserve el último maestro bueno.
    if not all_records:
        print("\n❌ 0 registros leídos de los Excels de Dropi.")
        print(f"   Verifica que '{ROOT}' sea accesible (Google Drive montado y con")
        print("   Full Disk Access para el proceso que ejecuta el pipeline).")
        print("   maestro_emails.xlsx NO se sobrescribe para no perder datos.")
        raise SystemExit(1)

    out_path = os.path.join(OUT_DIR, "maestro_emails.xlsx")

    # ── Guarda de regresión: no sobrescribir un maestro bueno con datos
    # incompletos. Compara el build nuevo contra el maestro anterior; si algún
    # Excel falló al abrir, o un mes que existía desaparece / cae >50%, aborta y
    # conserva el último maestro bueno (falla cerrado). Típico cuando un archivo
    # de Drive quedó "solo en la nube" al correr el agente.
    nuevo_por_mes = Counter(mes for (_e, _p, mes) in agg.keys())
    prev_por_mes = meses_conteo_maestro(out_path)
    problemas = []
    if failed_files:
        problemas.append(
            f"{len(failed_files)} archivo(s) no se pudieron abrir: "
            + ", ".join(f"{n} [{m}]" for n, m, _ in failed_files))
    for mes, prev_n in sorted(prev_por_mes.items()):
        nuevo_n = nuevo_por_mes.get(mes, 0)
        if nuevo_n == 0:
            problemas.append(f"mes {mes}: estaba con {prev_n} filas y ahora 0 (desapareció)")
        elif prev_n >= 20 and nuevo_n < prev_n * 0.5:
            problemas.append(f"mes {mes}: cayó de {prev_n} a {nuevo_n} filas (<50%)")
    if problemas and prev_por_mes:
        print("\n❌ Consolidación INCOMPLETA — NO se sobrescribe el maestro:")
        for p in problemas:
            print(f"   • {p}")
        print("   Causa típica: Excels de Google Drive 'solo en la nube' no")
        print("   materializados al correr. Se conserva el último maestro bueno.")
        raise SystemExit(2)
    elif problemas:
        print("\n⚠ Advertencias (sin maestro previo que proteger, se continúa):")
        for p in problemas:
            print(f"   • {p}")

    # Escribir maestro
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "MAESTRO"
    ws.append(["email","pais","mes","entregados","devoluciones","pedidos","nombre","telefono"])
    for (email, pais, mes), v in sorted(agg.items()):
        ws.append([email, pais, mes, v["entregados"], v["devoluciones"], v["pedidos"], v["nombre"], v["telefono"]])

    # Pivot: una fila por email+pais con columnas mes
    pivot = defaultdict(lambda: {m:0 for m in MONTH_MAP.values()})
    meta = {}
    for (email, pais, mes), v in agg.items():
        pivot[(email, pais)][mes] += v["pedidos"]
        if (email,pais) not in meta:
            meta[(email,pais)] = (v["nombre"], v["telefono"])

    ws2 = wb_out.create_sheet("PIVOT_USUARIO_PAIS")
    months = sorted(MONTH_MAP.values())
    ws2.append(["email","pais","nombre","telefono"] + months + ["total_4m","top1","top2","suma_top2"])
    for (email,pais), monthly in sorted(pivot.items()):
        nombre, tel = meta[(email,pais)]
        vals = [monthly[m] for m in months]
        sorted_v = sorted(vals, reverse=True)
        top1, top2 = sorted_v[0], sorted_v[1] if len(sorted_v)>1 else 0
        ws2.append([email, pais, nombre, tel] + vals + [sum(vals), top1, top2, top1+top2])

    # Escritura atómica: guardar a un temporal y reemplazar de golpe, para que
    # un lector concurrente (reclasificar.py corre en su propio agente) nunca
    # vea un .xlsx a medio escribir.
    tmp_path = out_path + ".tmp"
    wb_out.save(tmp_path)
    os.replace(tmp_path, out_path)
    print(f"\n✅ Maestro escrito en: {out_path}")
    print(f"   - Hoja MAESTRO: {len(agg)} filas (email × pais × mes)")
    print(f"   - Hoja PIVOT_USUARIO_PAIS: {len(pivot)} filas (email × pais)")

    # Resumen rápido
    distinct_emails = set(e for (e,_) in pivot.keys())
    print(f"   - Correos únicos (sin importar país): {len(distinct_emails)}")

if __name__ == "__main__":
    main()
